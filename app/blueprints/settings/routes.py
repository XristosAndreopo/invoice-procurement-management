"""
app/blueprints/settings/routes.py

Settings & Master Data routes.

Enterprise scope:
- Theme selection (all logged-in users)
- Feedback form (all logged-in users)
- Feedback admin (admin-only)
- ServiceUnits CRUD (admin-only)
- Suppliers CRUD (admin-only)
- OptionValue pages (enterprise dropdown master data)
- Income Tax Rules (admin-only)
- Withholding Profiles (admin-only)
- Procurement Committees (manager+admin)

NEW:
- ALE–KAE master list (admin-only) + Excel import
- CPV master list (admin-only) + Excel import

REPORT FIELDS (V4.3):
- ServiceUnit.address, ServiceUnit.phone
- Supplier.email, Supplier.emba
These are required for report headers (e.g., Προτιμολόγιο).

SECURITY:
- UI is never trusted. All permissions are enforced server-side.
- Viewer read-only guard exists globally, but routes still enforce explicit decorators.
"""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from typing import Dict, Optional, Set
import unicodedata

from flask import Blueprint, flash, redirect, render_template, request, url_for, abort
from flask_login import current_user, login_required

from ...audit import log_action, serialize_model
from ...extensions import db
from ...models import (
    Feedback,
    OptionCategory,
    OptionValue,
    Personnel,
    ServiceUnit,
    Supplier,
    IncomeTaxRule,
    WithholdingProfile,
    ProcurementCommittee,
    AleKae,
    Cpv,
)
from ...security import admin_required, manager_required

settings_bp = Blueprint("settings", __name__, url_prefix="/settings")


# ----------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------
def _parse_optional_int(value: str) -> Optional[int]:
    """Parse optional int from string; returns None for empty/invalid."""
    if not value:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _parse_decimal(value: str | None) -> Optional[Decimal]:
    """Parse decimal supporting comma/dot; returns None if empty/invalid."""
    if value is None:
        return None
    raw = str(value).strip().replace(",", ".")
    if raw == "":
        return None
    try:
        return Decimal(raw)
    except (InvalidOperation, ValueError):
        return None


def _active_personnel_for_dropdown(service_unit_id: Optional[int] = None):
    """Active personnel list for dropdown selection (optionally filtered by service unit)."""
    q = Personnel.query.filter_by(is_active=True)
    if service_unit_id:
        q = q.filter_by(service_unit_id=service_unit_id)
    return q.order_by(Personnel.last_name.asc(), Personnel.first_name.asc()).all()


def _get_or_create_category(key: str, label: str) -> OptionCategory:
    """
    Ensure an OptionCategory exists.

    Safe even if seed wasn't run yet.
    """
    category = OptionCategory.query.filter_by(key=key).first()
    if category:
        if category.label != label:
            category.label = label
            db.session.commit()
        return category

    category = OptionCategory(key=key, label=label)
    db.session.add(category)
    db.session.commit()
    return category


def _option_values_for_category(category: OptionCategory):
    """Return OptionValue list ordered consistently."""
    return (
        OptionValue.query.filter_by(category_id=category.id)
        .order_by(OptionValue.sort_order.asc(), OptionValue.value.asc())
        .all()
    )


def _ensure_committee_scope_or_403(service_unit_id: int):
    """
    Manager/Deputy can manage committees ONLY for their own ServiceUnit.
    Admin can manage all.
    """
    if current_user.is_admin:
        return
    if not current_user.service_unit_id or current_user.service_unit_id != service_unit_id:
        abort(403)
    if not current_user.can_manage():
        abort(403)


def _normalize_header(text: str) -> str:
    """
    Normalize Excel headers:
    - lowercase
    - trim spaces
    - remove diacritics (e.g. Περιγραφή == Περιγραφη)
    """
    if text is None:
        return ""
    s = " ".join(str(text).strip().lower().split())
    s = "".join(ch for ch in unicodedata.normalize("NFD", s) if unicodedata.category(ch) != "Mn")
    return s


def _safe_str(v) -> str:
    """Convert an excel cell to trimmed string (safe)."""
    if v is None:
        return ""
    return str(v).strip()


# ----------------------------------------------------------------------
# THEME (all users)
# ----------------------------------------------------------------------
@settings_bp.route("/theme", methods=["GET", "POST"])
@login_required
def theme():
    """Allow any logged-in user to select their theme."""
    themes = {
        "default": ("Προεπιλογή", "Φωτεινό θέμα με ουδέτερα χρώματα."),
        "dark": ("Σκούρο", "Σκούρο θέμα κατάλληλο για χαμηλό φωτισμό."),
        "ocean": ("Ocean", "Απαλό μπλε θέμα."),
    }

    if request.method == "POST":
        selected = request.form.get("theme")
        if selected not in themes:
            flash("Μη έγκυρο θέμα.", "danger")
            return redirect(url_for("settings.theme"))

        current_user.theme = selected
        db.session.commit()
        flash("Το θέμα ενημερώθηκε.", "success")
        return redirect(url_for("settings.theme"))

    return render_template("settings/theme.html", themes=themes)


# ----------------------------------------------------------------------
# FEEDBACK (all users)
# ----------------------------------------------------------------------
@settings_bp.route("/feedback", methods=["GET", "POST"])
@login_required
def feedback():
    """Feedback / complaint form."""
    categories = [
        ("complaint", "Παράπονο"),
        ("suggestion", "Πρόταση"),
        ("bug", "Σφάλμα"),
        ("other", "Άλλο"),
    ]

    if request.method == "POST":
        category = request.form.get("category") or None
        subject = (request.form.get("subject") or "").strip()
        message = (request.form.get("message") or "").strip()
        related_procurement_id_raw = (request.form.get("related_procurement_id") or "").strip()

        if not subject:
            flash("Ο τίτλος είναι υποχρεωτικός.", "danger")
            return redirect(url_for("settings.feedback"))
        if not message:
            flash("Το κείμενο είναι υποχρεωτικό.", "danger")
            return redirect(url_for("settings.feedback"))

        related_procurement_id = _parse_optional_int(related_procurement_id_raw)
        if related_procurement_id_raw and related_procurement_id is None:
            flash("Μη έγκυρο Α/Α προμήθειας.", "danger")
            return redirect(url_for("settings.feedback"))

        fb = Feedback(
            user_id=current_user.id,
            category=category,
            subject=subject,
            message=message,
            related_procurement_id=related_procurement_id,
            status="new",
        )
        db.session.add(fb)
        db.session.commit()

        flash("Το μήνυμά σας καταχωρήθηκε.", "success")
        return redirect(url_for("settings.feedback"))

    recent_feedback = (
        Feedback.query.filter_by(user_id=current_user.id)
        .order_by(Feedback.created_at.desc())
        .limit(5)
        .all()
    )
    return render_template("settings/feedback.html", categories=categories, recent_feedback=recent_feedback)


# ----------------------------------------------------------------------
# FEEDBACK ADMIN (admin only)
# ----------------------------------------------------------------------
@settings_bp.route("/feedback/admin", methods=["GET", "POST"])
@login_required
@admin_required
def feedback_admin():
    """Admin-only page to review and manage all feedback."""
    status_choices: Dict[str, str] = {
        "new": "Νέο",
        "in_progress": "Σε εξέλιξη",
        "resolved": "Επιλυμένο",
        "closed": "Κλειστό",
    }

    category_labels: Dict[Optional[str], str] = {
        "complaint": "Παράπονο",
        "suggestion": "Πρόταση",
        "bug": "Σφάλμα",
        "other": "Άλλο",
        None: "—",
    }

    status_filter = (request.args.get("status") or "").strip() or None
    category_filter = (request.args.get("category") or "").strip() or None

    if request.method == "POST":
        fb_id_raw = (request.form.get("feedback_id") or "").strip()
        new_status = (request.form.get("status") or "").strip()

        fb_id = _parse_optional_int(fb_id_raw)
        if fb_id is None or new_status not in status_choices:
            flash("Μη έγκυρη ενημέρωση κατάστασης.", "danger")
            return redirect(url_for("settings.feedback_admin"))

        fb = Feedback.query.get(fb_id)
        if not fb:
            flash("Το συγκεκριμένο παράπονο δεν βρέθηκε.", "danger")
            return redirect(url_for("settings.feedback_admin"))

        fb.status = new_status
        db.session.commit()
        flash("Η κατάσταση ενημερώθηκε.", "success")
        return redirect(url_for("settings.feedback_admin"))

    q = Feedback.query

    if status_filter and status_filter in status_choices:
        q = q.filter(Feedback.status == status_filter)

    if category_filter and category_filter in {"complaint", "suggestion", "bug", "other"}:
        q = q.filter(Feedback.category == category_filter)

    feedback_items = q.order_by(Feedback.created_at.desc()).all()

    return render_template(
        "settings/feedback_admin.html",
        feedback_items=feedback_items,
        status_choices=status_choices,
        category_labels=category_labels,
        status_filter=status_filter,
        category_filter=category_filter,
    )


# ----------------------------------------------------------------------
# SERVICE UNITS (admin-only)
# ----------------------------------------------------------------------
@settings_bp.route("/service-units")
@login_required
@admin_required
def service_units_list():
    """List ServiceUnits basic info (admin-only)."""
    units = ServiceUnit.query.order_by(ServiceUnit.description.asc()).all()
    return render_template("settings/service_units_list.html", units=units)


@settings_bp.route("/service-units/roles")
@login_required
@admin_required
def service_units_roles_list():
    """List ServiceUnits with Manager/Deputy (admin-only)."""
    units = ServiceUnit.query.order_by(ServiceUnit.description.asc()).all()
    return render_template("settings/service_units_roles_list.html", units=units)


@settings_bp.route("/service-units/new", methods=["GET", "POST"])
@login_required
@admin_required
def service_unit_create():
    """
    Create ServiceUnit (admin-only).

    IMPORTANT:
    - Basic fields only.
    - Manager/Deputy assignment is done in /service-units/<id>/edit.

    REPORT FIELDS:
    - address, phone are used by reports (e.g., Προτιμολόγιο header).
    """
    if request.method == "POST":
        description = (request.form.get("description") or "").strip()
        code = (request.form.get("code") or "").strip()
        short_name = (request.form.get("short_name") or "").strip()
        aahit = (request.form.get("aahit") or "").strip()
        commander = (request.form.get("commander") or "").strip()
        curator = (request.form.get("curator") or "").strip()
        supply_officer = (request.form.get("supply_officer") or "").strip()

        # NEW: report header fields
        address = (request.form.get("address") or "").strip()
        phone = (request.form.get("phone") or "").strip()

        if not description:
            flash("Η περιγραφή είναι υποχρεωτική.", "danger")
            return redirect(url_for("settings.service_unit_create"))

        unit = ServiceUnit(
            description=description,
            code=code or None,
            short_name=short_name or None,
            aahit=aahit or None,
            commander=commander or None,
            curator=curator or None,
            supply_officer=supply_officer or None,
            address=address or None,
            phone=phone or None,
            manager_personnel_id=None,
            deputy_personnel_id=None,
        )

        db.session.add(unit)
        db.session.flush()
        log_action(entity=unit, action="CREATE", after=serialize_model(unit))
        db.session.commit()

        flash("Η υπηρεσία δημιουργήθηκε.", "success")
        return redirect(url_for("settings.service_units_list"))

    return render_template(
        "settings/service_unit_form.html",
        unit=None,
        form_title="Νέα Υπηρεσία",
        is_create=True,
    )


@settings_bp.route("/service-units/import", methods=["POST"])
@login_required
@admin_required
def service_units_import():
    """
    Import ServiceUnits from Excel (admin-only).

    Expected headers (Greek or English, accents are ignored):
    - Κωδικός / code
    - Περιγραφή / description  (required)
    - Συντομογραφία / short_name (or "short name")

    NOTE:
    - This import currently does NOT set address/phone.
      Those are report fields and can be edited from the form.

    Behavior:
    - Creates new ServiceUnits from rows with a non-empty description.
    - Skips empty rows.
    - Does NOT set Manager/Deputy here.
    - AUDIT: logs CREATE per inserted ServiceUnit (id exists after flush).
    """
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Δεν επιλέχθηκε αρχείο.", "danger")
        return redirect(url_for("settings.service_units_list"))

    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        flash("Επιτρέπεται μόνο αρχείο .xlsx", "danger")
        return redirect(url_for("settings.service_units_list"))

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception:
        flash("Αποτυχία ανάγνωσης Excel. Ελέγξτε το αρχείο.", "danger")
        return redirect(url_for("settings.service_units_list"))

    header_cells = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = [str(h).strip() if h is not None else "" for h in header_cells]
    idx_map = {_normalize_header(h): i for i, h in enumerate(headers) if _normalize_header(h)}

    code_idx = idx_map.get("κωδικος", idx_map.get("code"))
    desc_idx = idx_map.get("περιγραφη", idx_map.get("description"))
    short_idx = idx_map.get("συντομογραφια", idx_map.get("short name", idx_map.get("short_name")))

    if desc_idx is None:
        flash("Το Excel πρέπει να έχει στήλη 'Περιγραφή' (ή 'description').", "danger")
        return redirect(url_for("settings.service_units_list"))

    inserted_units: list[ServiceUnit] = []
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        desc_val = row[desc_idx] if desc_idx < len(row) else None
        description = (str(desc_val).strip() if desc_val is not None else "")
        if not description:
            skipped += 1
            continue

        code = None
        if code_idx is not None and code_idx < len(row):
            v = row[code_idx]
            code = (str(v).strip() if v is not None else "") or None

        short_name = None
        if short_idx is not None and short_idx < len(row):
            v = row[short_idx]
            short_name = (str(v).strip() if v is not None else "") or None

        unit = ServiceUnit(
            description=description,
            code=code,
            short_name=short_name,
            address=None,
            phone=None,
            manager_personnel_id=None,
            deputy_personnel_id=None,
        )
        db.session.add(unit)
        inserted_units.append(unit)

    if not inserted_units:
        flash("Δεν βρέθηκαν έγκυρες γραμμές προς εισαγωγή.", "warning")
        return redirect(url_for("settings.service_units_list"))

    db.session.flush()

    for unit in inserted_units:
        log_action(entity=unit, action="CREATE", before=None, after=serialize_model(unit))

    db.session.commit()

    flash(f"Εισαγωγή ολοκληρώθηκε: {len(inserted_units)} νέες υπηρεσίες, {skipped} γραμμές αγνοήθηκαν.", "success")
    return redirect(url_for("settings.service_units_list"))


@settings_bp.route("/service-units/<int:unit_id>/edit-info", methods=["GET", "POST"])
@login_required
@admin_required
def service_unit_edit_info(unit_id: int):
    """
    Edit ServiceUnit basic fields (admin-only).

    REPORT FIELDS:
    - address, phone used by reports (e.g., Προτιμολόγιο).
    """
    unit = ServiceUnit.query.get_or_404(unit_id)

    if request.method == "POST":
        before = serialize_model(unit)

        description = (request.form.get("description") or "").strip()
        code = (request.form.get("code") or "").strip()
        short_name = (request.form.get("short_name") or "").strip()
        aahit = (request.form.get("aahit") or "").strip()
        commander = (request.form.get("commander") or "").strip()
        curator = (request.form.get("curator") or "").strip()
        supply_officer = (request.form.get("supply_officer") or "").strip()

        # NEW: report header fields
        address = (request.form.get("address") or "").strip()
        phone = (request.form.get("phone") or "").strip()

        if not description:
            flash("Η περιγραφή είναι υποχρεωτική.", "danger")
            return redirect(url_for("settings.service_unit_edit_info", unit_id=unit_id))

        unit.description = description
        unit.code = code or None
        unit.short_name = short_name or None
        unit.aahit = aahit or None
        unit.commander = commander or None
        unit.curator = curator or None
        unit.supply_officer = supply_officer or None
        unit.address = address or None
        unit.phone = phone or None

        db.session.flush()
        log_action(entity=unit, action="UPDATE", before=before, after=serialize_model(unit))
        db.session.commit()

        flash("Η υπηρεσία ενημερώθηκε.", "success")
        return redirect(url_for("settings.service_units_list"))

    return render_template(
        "settings/service_unit_form.html",
        unit=unit,
        form_title="Επεξεργασία Υπηρεσίας",
        is_create=False,
    )


@settings_bp.route("/service-units/<int:unit_id>/edit", methods=["GET", "POST"])
@login_required
@admin_required
def service_unit_edit(unit_id: int):
    """Assign Manager/Deputy (admin-only)."""
    unit = ServiceUnit.query.get_or_404(unit_id)
    personnel_list = _active_personnel_for_dropdown()

    if request.method == "POST":
        before = serialize_model(unit)

        manager_pid = _parse_optional_int((request.form.get("manager_personnel_id") or "").strip())
        deputy_pid = _parse_optional_int((request.form.get("deputy_personnel_id") or "").strip())

        if manager_pid and deputy_pid and manager_pid == deputy_pid:
            flash("Ο ίδιος/η ίδια δεν μπορεί να είναι και Manager και Deputy.", "danger")
            return redirect(url_for("settings.service_unit_edit", unit_id=unit_id))

        active_ids: Set[int] = {p.id for p in personnel_list}
        if manager_pid and manager_pid not in active_ids:
            flash("Μη έγκυρος Manager. Επιτρέπεται μόνο ενεργό προσωπικό.", "danger")
            return redirect(url_for("settings.service_unit_edit", unit_id=unit_id))
        if deputy_pid and deputy_pid not in active_ids:
            flash("Μη έγκυρος Deputy. Επιτρέπεται μόνο ενεργό προσωπικό.", "danger")
            return redirect(url_for("settings.service_unit_edit", unit_id=unit_id))

        unit.manager_personnel_id = manager_pid
        unit.deputy_personnel_id = deputy_pid

        db.session.flush()
        log_action(entity=unit, action="UPDATE", before=before, after=serialize_model(unit))
        db.session.commit()

        flash("Οι ρόλοι Manager/Deputy ενημερώθηκαν.", "success")
        return redirect(url_for("settings.service_units_roles_list"))

    return render_template(
        "settings/service_unit_roles_form.html",
        unit=unit,
        personnel_list=personnel_list,
        form_title="Ορισμός Deputy/Manager",
    )


@settings_bp.route("/service-units/<int:unit_id>/delete", methods=["POST"])
@login_required
@admin_required
def service_unit_delete(unit_id: int):
    """Delete ServiceUnit (admin-only)."""
    unit = ServiceUnit.query.get_or_404(unit_id)
    before = serialize_model(unit)

    db.session.delete(unit)
    db.session.flush()
    log_action(entity=unit, action="DELETE", before=before, after=None)
    db.session.commit()

    flash("Η υπηρεσία διαγράφηκε.", "success")
    return redirect(url_for("settings.service_units_list"))


# ----------------------------------------------------------------------
# SUPPLIERS CRUD (admin only)
# ----------------------------------------------------------------------
@settings_bp.route("/suppliers")
@login_required
@admin_required
def suppliers_list():
    """List suppliers (admin-only)."""
    suppliers = Supplier.query.order_by(Supplier.name.asc()).all()
    return render_template("settings/suppliers_list.html", suppliers=suppliers)


@settings_bp.route("/suppliers/new", methods=["GET", "POST"])
@login_required
@admin_required
def supplier_create():
    """
    Create supplier (admin-only).

    REPORT FIELDS:
    - email, emba used in reports (winner supplier section).
    """
    if request.method == "POST":
        afm = (request.form.get("afm") or "").strip()
        name = (request.form.get("name") or "").strip()
        doy = (request.form.get("doy") or "").strip()
        # NEW: report fields
        email = (request.form.get("email") or "").strip()
        emba = (request.form.get("emba") or "").strip()

        address = (request.form.get("address") or "").strip()
        city = (request.form.get("city") or "").strip()
        postal_code = (request.form.get("postal_code") or "").strip()
        country = (request.form.get("country") or "").strip()
        bank_name = (request.form.get("bank_name") or "").strip()
        iban = (request.form.get("iban") or "").strip()

        if not afm or len(afm) != 9 or not afm.isdigit():
            flash("Το ΑΦΜ πρέπει να είναι 9 ψηφία.", "danger")
            return redirect(url_for("settings.supplier_create"))
        if not name:
            flash("Η επωνυμία είναι υποχρεωτική.", "danger")
            return redirect(url_for("settings.supplier_create"))

        supplier = Supplier(
            afm=afm,
            name=name,
            doy=doy or None,
            email=email or None,
            emba=emba or None,
            address=address or None,
            city=city or None,
            postal_code=postal_code or None,
            country=country or None,
            bank_name=bank_name or None,
            iban=iban or None,
        )

        db.session.add(supplier)
        db.session.flush()
        log_action(entity=supplier, action="CREATE", after=serialize_model(supplier))
        db.session.commit()

        flash("Ο προμηθευτής δημιουργήθηκε.", "success")
        return redirect(url_for("settings.suppliers_list"))

    return render_template("settings/supplier_form.html", supplier=None, form_title="Νέος Προμηθευτής")


@settings_bp.route("/suppliers/<int:supplier_id>/edit", methods=["GET", "POST"])
@login_required
@admin_required
def supplier_edit(supplier_id: int):
    """
    Edit supplier (admin-only).

    REPORT FIELDS:
    - email, emba used in reports (winner supplier section).
    """
    supplier = Supplier.query.get_or_404(supplier_id)

    if request.method == "POST":
        before = serialize_model(supplier)

        afm = (request.form.get("afm") or "").strip()
        name = (request.form.get("name") or "").strip()
        doy = (request.form.get("doy") or "").strip()
        # NEW: report fields
        email = (request.form.get("email") or "").strip()
        emba = (request.form.get("emba") or "").strip()

        address = (request.form.get("address") or "").strip()
        city = (request.form.get("city") or "").strip()
        postal_code = (request.form.get("postal_code") or "").strip()
        country = (request.form.get("country") or "").strip()
        bank_name = (request.form.get("bank_name") or "").strip()
        iban = (request.form.get("iban") or "").strip()

        if not afm or len(afm) != 9 or not afm.isdigit():
            flash("Το ΑΦΜ πρέπει να είναι 9 ψηφία.", "danger")
            return redirect(url_for("settings.supplier_edit", supplier_id=supplier_id))
        if not name:
            flash("Η επωνυμία είναι υποχρεωτική.", "danger")
            return redirect(url_for("settings.supplier_edit", supplier_id=supplier_id))

        supplier.afm = afm
        supplier.name = name
        supplier.doy = doy or None
        supplier.email = email or None
        supplier.emba = emba or None
        supplier.address = address or None
        supplier.city = city or None
        supplier.postal_code = postal_code or None
        supplier.country = country or None
        supplier.bank_name = bank_name or None
        supplier.iban = iban or None

        db.session.flush()
        log_action(entity=supplier, action="UPDATE", before=before, after=serialize_model(supplier))
        db.session.commit()

        flash("Ο προμηθευτής ενημερώθηκε.", "success")
        return redirect(url_for("settings.suppliers_list"))

    return render_template("settings/supplier_form.html", supplier=supplier, form_title="Επεξεργασία Προμηθευτή")


@settings_bp.route("/suppliers/<int:supplier_id>/delete", methods=["POST"])
@login_required
@admin_required
def supplier_delete(supplier_id: int):
    """Delete supplier (admin-only)."""
    supplier = Supplier.query.get_or_404(supplier_id)
    before = serialize_model(supplier)

    db.session.delete(supplier)
    db.session.flush()
    log_action(entity=supplier, action="DELETE", before=before)
    db.session.commit()

    flash("Ο προμηθευτής διαγράφηκε.", "success")
    return redirect(url_for("settings.suppliers_list"))


@settings_bp.route("/suppliers/import", methods=["POST"])
@login_required
@admin_required
def suppliers_import():
    """
    Admin-only Excel import for Suppliers.

    Headers (Greek preferred, accents ignored):
    - ΑΦΜ (required) / afm
    - ΕΠΩΝΥΜΙΑ (required) / name / ονομασια
    - ΔΟΥ (optional) / doy / δ.ο.υ.
    - EMAIL, ΕΜΠΑ, ΔΙΕΥΘΥΝΣΗ, ΤΟΠΟΣ, ΤΚ, ΧΩΡΑ, ΤΡΑΠΕΖΑ, IBAN (optional)

    Behavior:
    - Skip duplicates by AFM (no update).
    """
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Δεν επιλέχθηκε αρχείο.", "danger")
        return redirect(url_for("settings.suppliers_list"))

    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        flash("Επιτρέπεται μόνο αρχείο .xlsx", "danger")
        return redirect(url_for("settings.suppliers_list"))

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception:
        flash("Αποτυχία ανάγνωσης Excel. Ελέγξτε το αρχείο.", "danger")
        return redirect(url_for("settings.suppliers_list"))

    try:
        header_cells = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    except StopIteration:
        flash("Το Excel είναι κενό.", "danger")
        return redirect(url_for("settings.suppliers_list"))

    # ✅ improve header normalize to support Δ.Ο.Υ.
    def _norm(h: str) -> str:
        s = _normalize_header(h)
        return s.replace(".", "")

    headers = [str(h).strip() if h is not None else "" for h in header_cells]
    idx_map = {_norm(h): i for i, h in enumerate(headers) if _norm(h)}

    afm_idx = idx_map.get("αφμ", idx_map.get("afm"))
    name_idx = idx_map.get("επωνυμια", idx_map.get("name", idx_map.get("ονομασια")))
    doy_idx = idx_map.get("δου", idx_map.get("doy", idx_map.get("δοy", idx_map.get("δοϋ"))))
    email_idx = idx_map.get("email")
    emba_idx = idx_map.get("εμπα", idx_map.get("emba"))
    addr_idx = idx_map.get("διευθυνση", idx_map.get("address"))
    city_idx = idx_map.get("τοπος", idx_map.get("city"))
    pc_idx = idx_map.get("τκ", idx_map.get("tk", idx_map.get("postal_code")))
    country_idx = idx_map.get("χωρα", idx_map.get("country"))
    bank_idx = idx_map.get("τραπεζα", idx_map.get("bank_name"))
    iban_idx = idx_map.get("iban")

    if afm_idx is None or name_idx is None:
        flash("Το Excel πρέπει να έχει στήλες 'ΑΦΜ' και 'ΕΠΩΝΥΜΙΑ' (ή 'name').", "danger")
        return redirect(url_for("settings.suppliers_list"))

    inserted: list[Supplier] = []
    skipped_missing = 0
    skipped_invalid_afm = 0
    skipped_duplicate = 0

    def _cell(row_vals, i: Optional[int]):
        if i is None or i >= len(row_vals):
            return None
        return row_vals[i]

    for row in ws.iter_rows(min_row=2, values_only=True):
        afm_raw = _safe_str(_cell(row, afm_idx))
        name_raw = _safe_str(_cell(row, name_idx))

        if not afm_raw or not name_raw:
            skipped_missing += 1
            continue

        afm = "".join(ch for ch in afm_raw if ch.isdigit())
        if len(afm) != 9:
            skipped_invalid_afm += 1
            continue

        if Supplier.query.filter_by(afm=afm).first():
            skipped_duplicate += 1
            continue

        supplier = Supplier(
            afm=afm,
            name=name_raw,
            doy=_safe_str(_cell(row, doy_idx)) or None,
            email=_safe_str(_cell(row, email_idx)) or None,
            emba=_safe_str(_cell(row, emba_idx)) or None,
            address=_safe_str(_cell(row, addr_idx)) or None,
            city=_safe_str(_cell(row, city_idx)) or None,
            postal_code=_safe_str(_cell(row, pc_idx)) or None,
            country=_safe_str(_cell(row, country_idx)) or None,
            bank_name=_safe_str(_cell(row, bank_idx)) or None,
            iban=_safe_str(_cell(row, iban_idx)) or None,
        )
        db.session.add(supplier)
        inserted.append(supplier)

    if not inserted:
        flash("Δεν εισήχθησαν εγγραφές. Ελέγξτε required πεδία/διπλότυπα/ΑΦΜ.", "warning")
        return redirect(url_for("settings.suppliers_list"))

    db.session.flush()
    for s in inserted:
        log_action(entity=s, action="CREATE", before=None, after=serialize_model(s))
    db.session.commit()

    flash(
        f"Εισαγωγή ολοκληρώθηκε: {len(inserted)} νέοι προμηθευτές. "
        f"Παραλείφθηκαν: {skipped_missing} (ελλιπή), {skipped_invalid_afm} (μη έγκυρο ΑΦΜ), {skipped_duplicate} (διπλότυπα).",
        "success",
    )
    return redirect(url_for("settings.suppliers_list"))

# ----------------------------------------------------------------------
# NEW: ALE–KAE (admin-only) + Excel import
# ----------------------------------------------------------------------
@settings_bp.route("/ale-kae", methods=["GET", "POST"])
@login_required
@admin_required
def ale_kae():
    """
    Admin-only CRUD page for ALE–KAE.

    POST action:
    - create, update, delete
    """
    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        if action == "create":
            ale = (request.form.get("ale") or "").strip()
            old_kae = (request.form.get("old_kae") or "").strip() or None
            description = (request.form.get("description") or "").strip() or None
            responsibility = (request.form.get("responsibility") or "").strip() or None

            if not ale:
                flash("Το ΑΛΕ είναι υποχρεωτικό.", "danger")
                return redirect(url_for("settings.ale_kae"))

            if AleKae.query.filter_by(ale=ale).first():
                flash("Υπάρχει ήδη εγγραφή με αυτό το ΑΛΕ.", "danger")
                return redirect(url_for("settings.ale_kae"))

            row = AleKae(
                ale=ale,
                old_kae=old_kae,
                description=description,
                responsibility=responsibility,
            )
            db.session.add(row)
            db.session.flush()
            log_action(entity=row, action="CREATE", before=None, after=serialize_model(row))
            db.session.commit()

            flash("Η εγγραφή ΑΛΕ-ΚΑΕ προστέθηκε.", "success")
            return redirect(url_for("settings.ale_kae"))

        if action == "update":
            rid = _parse_optional_int((request.form.get("id") or "").strip())
            if rid is None:
                flash("Μη έγκυρη εγγραφή.", "danger")
                return redirect(url_for("settings.ale_kae"))

            row = AleKae.query.get_or_404(rid)
            before = serialize_model(row)

            ale = (request.form.get("ale") or "").strip()
            old_kae = (request.form.get("old_kae") or "").strip() or None
            description = (request.form.get("description") or "").strip() or None
            responsibility = (request.form.get("responsibility") or "").strip() or None

            if not ale:
                flash("Το ΑΛΕ είναι υποχρεωτικό.", "danger")
                return redirect(url_for("settings.ale_kae"))

            exists = AleKae.query.filter(AleKae.ale == ale, AleKae.id != row.id).first()
            if exists:
                flash("Υπάρχει ήδη άλλη εγγραφή με αυτό το ΑΛΕ.", "danger")
                return redirect(url_for("settings.ale_kae"))

            row.ale = ale
            row.old_kae = old_kae
            row.description = description
            row.responsibility = responsibility

            db.session.flush()
            log_action(entity=row, action="UPDATE", before=before, after=serialize_model(row))
            db.session.commit()

            flash("Η εγγραφή ενημερώθηκε.", "success")
            return redirect(url_for("settings.ale_kae"))

        if action == "delete":
            rid = _parse_optional_int((request.form.get("id") or "").strip())
            if rid is None:
                flash("Μη έγκυρη εγγραφή.", "danger")
                return redirect(url_for("settings.ale_kae"))

            row = AleKae.query.get_or_404(rid)
            before = serialize_model(row)

            db.session.delete(row)
            db.session.flush()
            log_action(entity=row, action="DELETE", before=before, after=None)
            db.session.commit()

            flash("Η εγγραφή διαγράφηκε.", "success")
            return redirect(url_for("settings.ale_kae"))

        flash("Μη έγκυρη ενέργεια.", "danger")
        return redirect(url_for("settings.ale_kae"))

    rows = AleKae.query.order_by(AleKae.ale.asc()).all()
    return render_template("settings/ale_kae.html", rows=rows)


@settings_bp.route("/ale-kae/import", methods=["POST"])
@login_required
@admin_required
def ale_kae_import():
    """
    Admin-only Excel import for ALE–KAE.

    Headers (Greek preferred, accents ignored):
    - ΑΛΕ (required)
    - ΠΑΛΙΟΣ ΚΑΕ
    - ΠΕΡΙΓΡΑΦΗ
    - ΑΡΜΟΔΙΟΤΗΤΑΣ
    """
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Δεν επιλέχθηκε αρχείο.", "danger")
        return redirect(url_for("settings.ale_kae"))

    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        flash("Επιτρέπεται μόνο αρχείο .xlsx", "danger")
        return redirect(url_for("settings.ale_kae"))

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception:
        flash("Αποτυχία ανάγνωσης Excel. Ελέγξτε το αρχείο.", "danger")
        return redirect(url_for("settings.ale_kae"))

    try:
        header_cells = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    except StopIteration:
        flash("Το Excel είναι κενό.", "danger")
        return redirect(url_for("settings.ale_kae"))

    headers = [str(h).strip() if h is not None else "" for h in header_cells]
    idx_map = {_normalize_header(h): i for i, h in enumerate(headers) if _normalize_header(h)}

    ale_idx = idx_map.get("αλε", idx_map.get("ale"))
    old_kae_idx = idx_map.get("παλιος καε", idx_map.get("old kae", idx_map.get("old_kae")))
    desc_idx = idx_map.get("περιγραφη", idx_map.get("description"))
    resp_idx = idx_map.get("αρμοδιοτητας", idx_map.get("responsibility"))

    if ale_idx is None:
        flash("Το Excel πρέπει να έχει στήλη 'ΑΛΕ'.", "danger")
        return redirect(url_for("settings.ale_kae"))

    inserted: list[AleKae] = []
    skipped_missing = 0
    skipped_duplicate = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        def _cell(i: Optional[int]):
            if i is None or i >= len(row):
                return None
            return row[i]

        ale = _safe_str(_cell(ale_idx))
        if not ale:
            skipped_missing += 1
            continue

        if AleKae.query.filter_by(ale=ale).first():
            skipped_duplicate += 1
            continue

        obj = AleKae(
            ale=ale,
            old_kae=_safe_str(_cell(old_kae_idx)) or None,
            description=_safe_str(_cell(desc_idx)) or None,
            responsibility=_safe_str(_cell(resp_idx)) or None,
        )
        db.session.add(obj)
        inserted.append(obj)

    if not inserted:
        flash("Δεν εισήχθησαν εγγραφές. Ελέγξτε required πεδία/διπλότυπα.", "warning")
        return redirect(url_for("settings.ale_kae"))

    db.session.flush()
    for obj in inserted:
        log_action(entity=obj, action="CREATE", before=None, after=serialize_model(obj))
    db.session.commit()

    flash(
        f"Εισαγωγή ολοκληρώθηκε: {len(inserted)} νέες εγγραφές. "
        f"Παραλείφθηκαν: {skipped_missing} (ελλιπή), {skipped_duplicate} (διπλότυπα).",
        "success",
    )
    return redirect(url_for("settings.ale_kae"))


# ----------------------------------------------------------------------
# NEW: CPV (admin-only) + Excel import
# ----------------------------------------------------------------------
@settings_bp.route("/cpv", methods=["GET", "POST"])
@login_required
@admin_required
def cpv():
    """
    Admin-only CRUD page for CPV.

    POST action:
    - create, update, delete
    """
    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        if action == "create":
            cpv_code = (request.form.get("cpv") or "").strip()
            description = (request.form.get("description") or "").strip() or None

            if not cpv_code:
                flash("Το CPV είναι υποχρεωτικό.", "danger")
                return redirect(url_for("settings.cpv"))

            if Cpv.query.filter_by(cpv=cpv_code).first():
                flash("Υπάρχει ήδη εγγραφή με αυτό το CPV.", "danger")
                return redirect(url_for("settings.cpv"))

            obj = Cpv(cpv=cpv_code, description=description)
            db.session.add(obj)
            db.session.flush()
            log_action(entity=obj, action="CREATE", before=None, after=serialize_model(obj))
            db.session.commit()

            flash("Η εγγραφή CPV προστέθηκε.", "success")
            return redirect(url_for("settings.cpv"))

        if action == "update":
            rid = _parse_optional_int((request.form.get("id") or "").strip())
            if rid is None:
                flash("Μη έγκυρη εγγραφή.", "danger")
                return redirect(url_for("settings.cpv"))

            obj = Cpv.query.get_or_404(rid)
            before = serialize_model(obj)

            cpv_code = (request.form.get("cpv") or "").strip()
            description = (request.form.get("description") or "").strip() or None

            if not cpv_code:
                flash("Το CPV είναι υποχρεωτικό.", "danger")
                return redirect(url_for("settings.cpv"))

            exists = Cpv.query.filter(Cpv.cpv == cpv_code, Cpv.id != obj.id).first()
            if exists:
                flash("Υπάρχει ήδη άλλη εγγραφή με αυτό το CPV.", "danger")
                return redirect(url_for("settings.cpv"))

            obj.cpv = cpv_code
            obj.description = description

            db.session.flush()
            log_action(entity=obj, action="UPDATE", before=before, after=serialize_model(obj))
            db.session.commit()

            flash("Η εγγραφή ενημερώθηκε.", "success")
            return redirect(url_for("settings.cpv"))

        if action == "delete":
            rid = _parse_optional_int((request.form.get("id") or "").strip())
            if rid is None:
                flash("Μη έγκυρη εγγραφή.", "danger")
                return redirect(url_for("settings.cpv"))

            obj = Cpv.query.get_or_404(rid)
            before = serialize_model(obj)

            db.session.delete(obj)
            db.session.flush()
            log_action(entity=obj, action="DELETE", before=before, after=None)
            db.session.commit()

            flash("Η εγγραφή διαγράφηκε.", "success")
            return redirect(url_for("settings.cpv"))

        flash("Μη έγκυρη ενέργεια.", "danger")
        return redirect(url_for("settings.cpv"))

    rows = Cpv.query.order_by(Cpv.cpv.asc()).all()
    return render_template("settings/cpv.html", rows=rows)


@settings_bp.route("/cpv/import", methods=["POST"])
@login_required
@admin_required
def cpv_import():
    """
    Admin-only Excel import for CPV.

    Headers (Greek preferred, accents ignored):
    - CPV (required)
    - ΠΕΡΙΓΡΑΦΗ
    """
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Δεν επιλέχθηκε αρχείο.", "danger")
        return redirect(url_for("settings.cpv"))

    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        flash("Επιτρέπεται μόνο αρχείο .xlsx", "danger")
        return redirect(url_for("settings.cpv"))

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception:
        flash("Αποτυχία ανάγνωσης Excel. Ελέγξτε το αρχείο.", "danger")
        return redirect(url_for("settings.cpv"))

    try:
        header_cells = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    except StopIteration:
        flash("Το Excel είναι κενό.", "danger")
        return redirect(url_for("settings.cpv"))

    headers = [str(h).strip() if h is not None else "" for h in header_cells]
    idx_map = {_normalize_header(h): i for i, h in enumerate(headers) if _normalize_header(h)}

    cpv_idx = idx_map.get("cpv")
    desc_idx = idx_map.get("περιγραφη", idx_map.get("description"))

    if cpv_idx is None:
        flash("Το Excel πρέπει να έχει στήλη 'CPV'.", "danger")
        return redirect(url_for("settings.cpv"))

    inserted: list[Cpv] = []
    skipped_missing = 0
    skipped_duplicate = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        def _cell(i: Optional[int]):
            if i is None or i >= len(row):
                return None
            return row[i]

        cpv_code = _safe_str(_cell(cpv_idx))
        if not cpv_code:
            skipped_missing += 1
            continue

        if Cpv.query.filter_by(cpv=cpv_code).first():
            skipped_duplicate += 1
            continue

        obj = Cpv(
            cpv=cpv_code,
            description=_safe_str(_cell(desc_idx)) or None,
        )
        db.session.add(obj)
        inserted.append(obj)

    if not inserted:
        flash("Δεν εισήχθησαν εγγραφές. Ελέγξτε required πεδία/διπλότυπα.", "warning")
        return redirect(url_for("settings.cpv"))

    db.session.flush()
    for obj in inserted:
        log_action(entity=obj, action="CREATE", before=None, after=serialize_model(obj))
    db.session.commit()

    flash(
        f"Εισαγωγή ολοκληρώθηκε: {len(inserted)} νέες εγγραφές. "
        f"Παραλείφθηκαν: {skipped_missing} (ελλιπή), {skipped_duplicate} (διπλότυπα).",
        "success",
    )
    return redirect(url_for("settings.cpv"))


# ----------------------------------------------------------------------
# OPTION VALUES + IncomeTax + Withholding + Committees (unchanged)
# ----------------------------------------------------------------------
def _options_page(key: str, label: str):
    category = _get_or_create_category(key=key, label=label)

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        if action == "create":
            value = (request.form.get("value") or "").strip()
            sort_order = _parse_optional_int((request.form.get("sort_order") or "").strip()) or 0
            is_active = bool(request.form.get("is_active") == "on")

            if not value:
                flash("Η τιμή είναι υποχρεωτική.", "danger")
                return redirect(request.path)

            ov = OptionValue(category_id=category.id, value=value, sort_order=sort_order, is_active=is_active)
            db.session.add(ov)
            db.session.flush()
            log_action(entity=ov, action="CREATE", after=serialize_model(ov))
            db.session.commit()

            flash("Η τιμή προστέθηκε.", "success")
            return redirect(request.path)

        if action == "update":
            ov_id = _parse_optional_int((request.form.get("id") or "").strip())
            if ov_id is None:
                flash("Μη έγκυρη εγγραφή.", "danger")
                return redirect(request.path)

            ov = OptionValue.query.filter_by(id=ov_id, category_id=category.id).first()
            if not ov:
                flash("Η εγγραφή δεν βρέθηκε.", "danger")
                return redirect(request.path)

            before = serialize_model(ov)

            value = (request.form.get("value") or "").strip()
            sort_order = _parse_optional_int((request.form.get("sort_order") or "").strip()) or 0
            is_active = bool(request.form.get("is_active") == "on")

            if not value:
                flash("Η τιμή είναι υποχρεωτική.", "danger")
                return redirect(request.path)

            ov.value = value
            ov.sort_order = sort_order
            ov.is_active = is_active

            db.session.flush()
            log_action(entity=ov, action="UPDATE", before=before, after=serialize_model(ov))
            db.session.commit()

            flash("Η εγγραφή ενημερώθηκε.", "success")
            return redirect(request.path)

        if action == "delete":
            ov_id = _parse_optional_int((request.form.get("id") or "").strip())
            if ov_id is None:
                flash("Μη έγκυρη εγγραφή.", "danger")
                return redirect(request.path)

            ov = OptionValue.query.filter_by(id=ov_id, category_id=category.id).first()
            if not ov:
                flash("Η εγγραφή δεν βρέθηκε.", "danger")
                return redirect(request.path)

            before = serialize_model(ov)

            db.session.delete(ov)
            db.session.flush()
            log_action(entity=ov, action="DELETE", before=before)
            db.session.commit()

            flash("Η εγγραφή διαγράφηκε.", "success")
            return redirect(request.path)

        flash("Μη έγκυρη ενέργεια.", "danger")
        return redirect(request.path)

    values = _option_values_for_category(category)
    return render_template("settings/options_values.html", category=category, values=values, page_label=label)


@settings_bp.route("/options/status", methods=["GET", "POST"])
@login_required
@admin_required
def options_status():
    return _options_page(key="KATASTASH", label="Κατάσταση")


@settings_bp.route("/options/stage", methods=["GET", "POST"])
@login_required
@admin_required
def options_stage():
    return _options_page(key="STADIO", label="Στάδιο")


@settings_bp.route("/options/allocation", methods=["GET", "POST"])
@login_required
@admin_required
def options_allocation():
    return _options_page(key="KATANOMH", label="Κατανομή")


@settings_bp.route("/options/quarterly", methods=["GET", "POST"])
@login_required
@admin_required
def options_quarterly():
    return _options_page(key="TRIMHNIAIA", label="Τριμηνιαία")


@settings_bp.route("/options/vat", methods=["GET", "POST"])
@login_required
@admin_required
def options_vat():
    return _options_page(key="FPA", label="ΦΠΑ")


@settings_bp.route("/income-tax", methods=["GET", "POST"])
@login_required
@admin_required
def income_tax_rules():
    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        if action == "create":
            desc = (request.form.get("description") or "").strip()
            rate = _parse_decimal(request.form.get("rate_percent")) or Decimal("0.00")
            threshold = _parse_decimal(request.form.get("threshold_amount")) or Decimal("150.00")
            is_active = bool(request.form.get("is_active") == "on")

            if not desc:
                flash("Η περιγραφή είναι υποχρεωτική.", "danger")
                return redirect(request.path)

            rule = IncomeTaxRule(
                description=desc,
                rate_percent=rate,
                threshold_amount=threshold,
                is_active=is_active,
            )
            db.session.add(rule)
            db.session.flush()
            log_action(rule, "CREATE", after=serialize_model(rule))
            db.session.commit()

            flash("Ο κανόνας ΦΕ προστέθηκε.", "success")
            return redirect(request.path)

        if action == "update":
            rid = _parse_optional_int((request.form.get("id") or "").strip())
            if rid is None:
                flash("Μη έγκυρη εγγραφή.", "danger")
                return redirect(request.path)

            rule = IncomeTaxRule.query.get_or_404(rid)
            before = serialize_model(rule)

            desc = (request.form.get("description") or "").strip()
            rate = _parse_decimal(request.form.get("rate_percent")) or Decimal("0.00")
            threshold = _parse_decimal(request.form.get("threshold_amount")) or Decimal("150.00")
            is_active = bool(request.form.get("is_active") == "on")

            if not desc:
                flash("Η περιγραφή είναι υποχρεωτική.", "danger")
                return redirect(request.path)

            rule.description = desc
            rule.rate_percent = rate
            rule.threshold_amount = threshold
            rule.is_active = is_active

            db.session.flush()
            log_action(rule, "UPDATE", before=before, after=serialize_model(rule))
            db.session.commit()

            flash("Ο κανόνας ΦΕ ενημερώθηκε.", "success")
            return redirect(request.path)

        if action == "delete":
            rid = _parse_optional_int((request.form.get("id") or "").strip())
            if rid is None:
                flash("Μη έγκυρη εγγραφή.", "danger")
                return redirect(request.path)

            rule = IncomeTaxRule.query.get_or_404(rid)
            before = serialize_model(rule)

            db.session.delete(rule)
            db.session.flush()
            log_action(rule, "DELETE", before=before)
            db.session.commit()

            flash("Ο κανόνας ΦΕ διαγράφηκε.", "success")
            return redirect(request.path)

        flash("Μη έγκυρη ενέργεια.", "danger")
        return redirect(request.path)

    rules = IncomeTaxRule.query.order_by(IncomeTaxRule.description.asc()).all()
    return render_template("settings/income_tax_rules.html", rules=rules)


@settings_bp.route("/withholding-profiles", methods=["GET", "POST"])
@login_required
@admin_required
def withholding_profiles():
    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        if action == "create":
            desc = (request.form.get("description") or "").strip()
            mt = _parse_decimal(request.form.get("mt_eloa_percent")) or Decimal("0.00")
            ea = _parse_decimal(request.form.get("eadhsy_percent")) or Decimal("0.00")
            k1 = _parse_decimal(request.form.get("withholding1_percent")) or Decimal("0.00")
            k2 = _parse_decimal(request.form.get("withholding2_percent")) or Decimal("0.00")
            is_active = bool(request.form.get("is_active") == "on")

            if not desc:
                flash("Η περιγραφή είναι υποχρεωτική.", "danger")
                return redirect(request.path)

            p = WithholdingProfile(
                description=desc,
                mt_eloa_percent=mt,
                eadhsy_percent=ea,
                withholding1_percent=k1,
                withholding2_percent=k2,
                is_active=is_active,
            )
            db.session.add(p)
            db.session.flush()
            log_action(p, "CREATE", after=serialize_model(p))
            db.session.commit()

            flash("Το προφίλ κρατήσεων προστέθηκε.", "success")
            return redirect(request.path)

        if action == "update":
            pid = _parse_optional_int((request.form.get("id") or "").strip())
            if pid is None:
                flash("Μη έγκυρη εγγραφή.", "danger")
                return redirect(request.path)

            p = WithholdingProfile.query.get_or_404(pid)
            before = serialize_model(p)

            desc = (request.form.get("description") or "").strip()
            mt = _parse_decimal(request.form.get("mt_eloa_percent")) or Decimal("0.00")
            ea = _parse_decimal(request.form.get("eadhsy_percent")) or Decimal("0.00")
            k1 = _parse_decimal(request.form.get("withholding1_percent")) or Decimal("0.00")
            k2 = _parse_decimal(request.form.get("withholding2_percent")) or Decimal("0.00")
            is_active = bool(request.form.get("is_active") == "on")

            if not desc:
                flash("Η περιγραφή είναι υποχρεωτική.", "danger")
                return redirect(request.path)

            p.description = desc
            p.mt_eloa_percent = mt
            p.eadhsy_percent = ea
            p.withholding1_percent = k1
            p.withholding2_percent = k2
            p.is_active = is_active

            db.session.flush()
            log_action(p, "UPDATE", before=before, after=serialize_model(p))
            db.session.commit()

            flash("Το προφίλ κρατήσεων ενημερώθηκε.", "success")
            return redirect(request.path)

        if action == "delete":
            pid = _parse_optional_int((request.form.get("id") or "").strip())
            if pid is None:
                flash("Μη έγκυρη εγγραφή.", "danger")
                return redirect(request.path)

            p = WithholdingProfile.query.get_or_404(pid)
            before = serialize_model(p)

            db.session.delete(p)
            db.session.flush()
            log_action(p, "DELETE", before=before)
            db.session.commit()

            flash("Το προφίλ κρατήσεων διαγράφηκε.", "success")
            return redirect(request.path)

        flash("Μη έγκυρη ενέργεια.", "danger")
        return redirect(request.path)

    profiles = WithholdingProfile.query.order_by(WithholdingProfile.description.asc()).all()
    return render_template("settings/withholding_profiles.html", profiles=profiles)


@settings_bp.route("/committees", methods=["GET", "POST"])
@login_required
@manager_required
def committees():
    if current_user.is_admin:
        scope_service_unit_id = _parse_optional_int((request.args.get("service_unit_id") or "").strip())
    else:
        scope_service_unit_id = current_user.service_unit_id

    service_units = ServiceUnit.query.order_by(ServiceUnit.description.asc()).all()

    committees_list = []
    personnel_list = []

    if scope_service_unit_id:
        _ensure_committee_scope_or_403(scope_service_unit_id)
        committees_list = (
            ProcurementCommittee.query
            .filter_by(service_unit_id=scope_service_unit_id)
            .order_by(ProcurementCommittee.description.asc())
            .all()
        )
        personnel_list = _active_personnel_for_dropdown(scope_service_unit_id)

    if request.method == "POST":
        su_id = _parse_optional_int((request.form.get("service_unit_id") or "").strip())
        if not su_id:
            flash("Η υπηρεσία είναι υποχρεωτική.", "danger")
            return redirect(url_for("settings.committees"))

        _ensure_committee_scope_or_403(su_id)

        allowed_ids = {p.id for p in _active_personnel_for_dropdown(su_id)}

        def _validate_member(pid: Optional[int]) -> Optional[int]:
            if pid is None:
                return None
            return pid if pid in allowed_ids else None

        action = (request.form.get("action") or "").strip()

        if action == "create":
            desc = (request.form.get("description") or "").strip()
            identity = (request.form.get("identity_text") or "").strip() or None
            president_id = _validate_member(_parse_optional_int((request.form.get("president_personnel_id") or "").strip()))
            member1_id = _validate_member(_parse_optional_int((request.form.get("member1_personnel_id") or "").strip()))
            member2_id = _validate_member(_parse_optional_int((request.form.get("member2_personnel_id") or "").strip()))
            is_active = bool(request.form.get("is_active") == "on")

            if not desc:
                flash("Η περιγραφή είναι υποχρεωτική.", "danger")
                return redirect(url_for("settings.committees", service_unit_id=su_id))

            c = ProcurementCommittee(
                service_unit_id=su_id,
                description=desc,
                identity_text=identity,
                president_personnel_id=president_id,
                member1_personnel_id=member1_id,
                member2_personnel_id=member2_id,
                is_active=is_active,
            )
            db.session.add(c)
            db.session.flush()
            log_action(c, "CREATE", after=serialize_model(c))
            db.session.commit()

            flash("Η επιτροπή προστέθηκε.", "success")
            return redirect(url_for("settings.committees", service_unit_id=su_id))

        if action == "update":
            cid = _parse_optional_int((request.form.get("id") or "").strip())
            if cid is None:
                flash("Μη έγκυρη επιτροπή.", "danger")
                return redirect(url_for("settings.committees", service_unit_id=su_id))

            c = ProcurementCommittee.query.get_or_404(cid)
            if c.service_unit_id != su_id:
                abort(403)

            before = serialize_model(c)

            desc = (request.form.get("description") or "").strip()
            identity = (request.form.get("identity_text") or "").strip() or None
            president_id = _validate_member(_parse_optional_int((request.form.get("president_personnel_id") or "").strip()))
            member1_id = _validate_member(_parse_optional_int((request.form.get("member1_personnel_id") or "").strip()))
            member2_id = _validate_member(_parse_optional_int((request.form.get("member2_personnel_id") or "").strip()))
            is_active = bool(request.form.get("is_active") == "on")

            if not desc:
                flash("Η περιγραφή είναι υποχρεωτική.", "danger")
                return redirect(url_for("settings.committees", service_unit_id=su_id))

            c.description = desc
            c.identity_text = identity
            c.president_personnel_id = president_id
            c.member1_personnel_id = member1_id
            c.member2_personnel_id = member2_id
            c.is_active = is_active

            db.session.flush()
            log_action(c, "UPDATE", before=before, after=serialize_model(c))
            db.session.commit()

            flash("Η επιτροπή ενημερώθηκε.", "success")
            return redirect(url_for("settings.committees", service_unit_id=su_id))

        if action == "delete":
            cid = _parse_optional_int((request.form.get("id") or "").strip())
            if cid is None:
                flash("Μη έγκυρη επιτροπή.", "danger")
                return redirect(url_for("settings.committees", service_unit_id=su_id))

            c = ProcurementCommittee.query.get_or_404(cid)
            if c.service_unit_id != su_id:
                abort(403)

            before = serialize_model(c)
            db.session.delete(c)
            db.session.flush()
            log_action(c, "DELETE", before=before)
            db.session.commit()

            flash("Η επιτροπή διαγράφηκε.", "success")
            return redirect(url_for("settings.committees", service_unit_id=su_id))

        flash("Μη έγκυρη ενέργεια.", "danger")
        return redirect(url_for("settings.committees", service_unit_id=su_id))

    return render_template(
        "settings/committees.html",
        service_units=service_units,
        committees=committees_list,
        personnel_list=personnel_list,
        scope_service_unit_id=scope_service_unit_id,
        is_admin=current_user.is_admin,
    )
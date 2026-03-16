"""
app/services/settings_suppliers_service.py

Focused page/use-case services for Supplier settings routes.

PURPOSE
-------
Extract non-HTTP orchestration from:

- /settings/suppliers
- /settings/suppliers/new
- /settings/suppliers/<id>/edit
- /settings/suppliers/<id>/delete
- /settings/suppliers/import

DESIGN
------
- function-first
- explicit validation paths
- routes stay thin
"""

from __future__ import annotations

from collections.abc import Mapping
from typing import Any

from ..audit import log_action, serialize_model
from ..extensions import db
from ..models import Supplier
from ..services.excel_imports import build_header_index, cell_at, normalize_header, safe_cell_str
from ..services.operation_results import FlashMessage, OperationResult


def build_suppliers_list_page_context() -> dict[str, Any]:
    """
    Build template context for the suppliers list page.
    """
    suppliers = Supplier.query.order_by(Supplier.name.asc()).all()
    return {"suppliers": suppliers}


def build_supplier_form_page_context(
    *,
    supplier: Supplier | None,
    form_title: str,
) -> dict[str, Any]:
    """
    Build template context for supplier create/edit form pages.
    """
    return {
        "supplier": supplier,
        "form_title": form_title,
    }


def execute_create_supplier(form_data: Mapping[str, Any]) -> OperationResult:
    """
    Validate and create a Supplier.
    """
    afm = (form_data.get("afm") or "").strip()
    name = (form_data.get("name") or "").strip()
    doy = (form_data.get("doy") or "").strip()
    phone = (form_data.get("phone") or "").strip()
    email = (form_data.get("email") or "").strip()
    emba = (form_data.get("emba") or "").strip()
    address = (form_data.get("address") or "").strip()
    city = (form_data.get("city") or "").strip()
    postal_code = (form_data.get("postal_code") or "").strip()
    country = (form_data.get("country") or "").strip()
    bank_name = (form_data.get("bank_name") or "").strip()
    iban = (form_data.get("iban") or "").strip()

    if not afm or len(afm) != 9 or not afm.isdigit():
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Το ΑΦΜ πρέπει να είναι 9 ψηφία.", "danger"),),
        )

    if not name:
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Η επωνυμία είναι υποχρεωτική.", "danger"),),
        )

    if Supplier.query.filter_by(afm=afm).first():
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Υπάρχει ήδη προμηθευτής με αυτό το ΑΦΜ.", "danger"),),
        )

    supplier = Supplier(
        afm=afm,
        name=name,
        doy=doy or None,
        phone=phone or None,
        email=email or None,
        emba=emba or None,
        address=address or None,
        city=city or None,
        postal_code=postal_code or None,
        country=country or None,
        bank_name=bank_name or None,
        iban=iban or None,
    )

    db.session.add(supplier)
    db.session.flush()
    log_action(entity=supplier, action="CREATE", before=None, after=serialize_model(supplier))
    db.session.commit()

    return OperationResult(
        ok=True,
        flashes=(FlashMessage("Ο προμηθευτής δημιουργήθηκε.", "success"),),
        entity_id=supplier.id,
    )


def execute_edit_supplier(
    supplier: Supplier,
    form_data: Mapping[str, Any],
) -> OperationResult:
    """
    Validate and update a Supplier.
    """
    before = serialize_model(supplier)

    afm = (form_data.get("afm") or "").strip()
    name = (form_data.get("name") or "").strip()
    doy = (form_data.get("doy") or "").strip()
    phone = (form_data.get("phone") or "").strip()
    email = (form_data.get("email") or "").strip()
    emba = (form_data.get("emba") or "").strip()
    address = (form_data.get("address") or "").strip()
    city = (form_data.get("city") or "").strip()
    postal_code = (form_data.get("postal_code") or "").strip()
    country = (form_data.get("country") or "").strip()
    bank_name = (form_data.get("bank_name") or "").strip()
    iban = (form_data.get("iban") or "").strip()

    if not afm or len(afm) != 9 or not afm.isdigit():
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Το ΑΦΜ πρέπει να είναι 9 ψηφία.", "danger"),),
        )

    if not name:
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Η επωνυμία είναι υποχρεωτική.", "danger"),),
        )

    existing_afm = Supplier.query.filter(
        Supplier.afm == afm,
        Supplier.id != supplier.id,
    ).first()
    if existing_afm:
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Υπάρχει ήδη άλλος προμηθευτής με αυτό το ΑΦΜ.", "danger"),),
        )

    supplier.afm = afm
    supplier.name = name
    supplier.doy = doy or None
    supplier.phone = phone or None
    supplier.email = email or None
    supplier.emba = emba or None
    supplier.address = address or None
    supplier.city = city or None
    supplier.postal_code = postal_code or None
    supplier.country = country or None
    supplier.bank_name = bank_name or None
    supplier.iban = iban or None

    db.session.flush()
    log_action(entity=supplier, action="UPDATE", before=before, after=serialize_model(supplier))
    db.session.commit()

    return OperationResult(
        ok=True,
        flashes=(FlashMessage("Ο προμηθευτής ενημερώθηκε.", "success"),),
        entity_id=supplier.id,
    )


def execute_delete_supplier(supplier: Supplier) -> OperationResult:
    """
    Delete a Supplier.
    """
    before = serialize_model(supplier)

    db.session.delete(supplier)
    db.session.flush()
    log_action(entity=supplier, action="DELETE", before=before, after=None)
    db.session.commit()

    return OperationResult(
        ok=True,
        flashes=(FlashMessage("Ο προμηθευτής διαγράφηκε.", "success"),),
    )


def execute_import_suppliers(file_storage: Any) -> OperationResult:
    """
    Import Suppliers from an uploaded Excel file.
    """
    if not file_storage or not getattr(file_storage, "filename", None):
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Δεν επιλέχθηκε αρχείο.", "danger"),),
        )

    filename = str(file_storage.filename or "").lower()
    if not filename.endswith(".xlsx"):
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Επιτρέπεται μόνο αρχείο .xlsx", "danger"),),
        )

    try:
        import openpyxl

        workbook = openpyxl.load_workbook(file_storage, data_only=True)
        worksheet = workbook.active
    except Exception:
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Αποτυχία ανάγνωσης Excel. Ελέγξτε το αρχείο.", "danger"),),
        )

    try:
        header_cells = [c.value for c in next(worksheet.iter_rows(min_row=1, max_row=1))]
    except StopIteration:
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Το Excel είναι κενό.", "danger"),),
        )

    def _norm(header_value: str | None) -> str:
        return normalize_header(header_value).replace(".", "")

    headers = [str(h).strip() if h is not None else "" for h in header_cells]
    idx_map = {_norm(h): i for i, h in enumerate(headers) if _norm(h)}

    afm_idx = idx_map.get("αφμ", idx_map.get("afm"))
    name_idx = idx_map.get("επωνυμια", idx_map.get("name", idx_map.get("ονομασια")))
    doy_idx = idx_map.get("δου", idx_map.get("doy", idx_map.get("δοy", idx_map.get("δοϋ"))))
    phone_idx = idx_map.get("τηλεφωνο", idx_map.get("phone", idx_map.get("tel")))
    email_idx = idx_map.get("email")
    emba_idx = idx_map.get("εμπα", idx_map.get("emba"))
    addr_idx = idx_map.get("διευθυνση", idx_map.get("address"))
    city_idx = idx_map.get("τοπος", idx_map.get("city"))
    pc_idx = idx_map.get("τκ", idx_map.get("tk", idx_map.get("postal_code")))
    country_idx = idx_map.get("χωρα", idx_map.get("country"))
    bank_idx = idx_map.get("τραπεζα", idx_map.get("bank_name"))
    iban_idx = idx_map.get("iban")

    if afm_idx is None or name_idx is None:
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Το Excel πρέπει να έχει στήλες 'ΑΦΜ' και 'ΕΠΩΝΥΜΙΑ' (ή 'name').", "danger"),),
        )

    inserted: list[Supplier] = []
    skipped_missing = 0
    skipped_invalid_afm = 0
    skipped_duplicate = 0

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        afm_raw = safe_cell_str(cell_at(row, afm_idx))
        name_raw = safe_cell_str(cell_at(row, name_idx))

        if not afm_raw or not name_raw:
            skipped_missing += 1
            continue

        afm = "".join(ch for ch in afm_raw if ch.isdigit())
        if len(afm) != 9:
            skipped_invalid_afm += 1
            continue

        if Supplier.query.filter_by(afm=afm).first():
            skipped_duplicate += 1
            continue

        supplier = Supplier(
            afm=afm,
            name=name_raw,
            doy=safe_cell_str(cell_at(row, doy_idx)) or None,
            phone=safe_cell_str(cell_at(row, phone_idx)) or None,
            email=safe_cell_str(cell_at(row, email_idx)) or None,
            emba=safe_cell_str(cell_at(row, emba_idx)) or None,
            address=safe_cell_str(cell_at(row, addr_idx)) or None,
            city=safe_cell_str(cell_at(row, city_idx)) or None,
            postal_code=safe_cell_str(cell_at(row, pc_idx)) or None,
            country=safe_cell_str(cell_at(row, country_idx)) or None,
            bank_name=safe_cell_str(cell_at(row, bank_idx)) or None,
            iban=safe_cell_str(cell_at(row, iban_idx)) or None,
        )
        db.session.add(supplier)
        inserted.append(supplier)

    if not inserted:
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Δεν εισήχθησαν εγγραφές. Ελέγξτε required πεδία/διπλότυπα/ΑΦΜ.", "warning"),),
        )

    db.session.flush()
    for supplier in inserted:
        log_action(entity=supplier, action="CREATE", before=None, after=serialize_model(supplier))
    db.session.commit()

    return OperationResult(
        ok=True,
        flashes=(
            FlashMessage(
                f"Εισαγωγή ολοκληρώθηκε: {len(inserted)} νέοι προμηθευτές. "
                f"Παραλείφθηκαν: {skipped_missing} (ελλιπή), "
                f"{skipped_invalid_afm} (μη έγκυρο ΑΦΜ), "
                f"{skipped_duplicate} (διπλότυπα).",
                "success",
            ),
        ),
    )
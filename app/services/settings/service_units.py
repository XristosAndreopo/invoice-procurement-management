"""
app/services/settings/service_units.py

Focused page/use-case services for ServiceUnit settings routes.

PURPOSE
-------
Extract non-HTTP orchestration from:

- /settings/service-units
- /settings/service-units/roles
- /settings/service-units/new
- /settings/service-units/import
- /settings/service-units/<id>/edit-info
- /settings/service-units/<id>/edit
- /settings/service-units/<id>/delete

ARCHITECTURAL INTENT
--------------------
Routes remain responsible only for:
- decorators
- reading request.form / request.files
- boundary object loads
- flashing returned messages
- render / redirect responses

This module handles:
- page-context assembly
- validation orchestration
- persistence
- audit logging
- SQLite-safe ServiceUnit deletion

DESIGN
------
- function-first
- no unnecessary classes
- explicit validation branches
"""

from __future__ import annotations

from collections.abc import Mapping
from typing import Any

from ...audit import log_action, serialize_model
from ...extensions import db
from ...models import Personnel, Procurement, ProcurementCommittee, ServiceUnit, User
from ..shared.excel_imports import build_header_index, cell_at, safe_cell_str
from ..shared.operation_results import FlashMessage, OperationResult


VALID_COMMANDER_ROLE_TYPES = {"Διοικητής", "Κυβερνήτης"}


def _active_personnel_for_dropdown() -> list[Personnel]:
    """
    Return all active Personnel ordered for dropdown usage.

    This preserves the current route behavior for ServiceUnit role assignment:
    Manager / Deputy may be selected from all active personnel.
    """
    return (
        Personnel.query.filter_by(is_active=True)
        .order_by(Personnel.last_name.asc(), Personnel.first_name.asc())
        .all()
    )


def _normalize_nullable_text(value: Any) -> str | None:
    """
    Normalize any input to trimmed nullable text.

    Empty strings become None.
    Non-string values are converted to string and stripped.
    """
    if value is None:
        return None

    text = str(value).strip()
    return text or None


def _normalize_commander_role_type(value: Any) -> str | None:
    """
    Normalize the commander/governor role type.

    Allowed persisted values:
    - Διοικητής
    - Κυβερνήτης

    Empty input becomes None.

    Raises:
        ValueError: if a non-empty value is provided but is not valid.
    """
    normalized = _normalize_nullable_text(value)
    if normalized is None:
        return None

    if normalized not in VALID_COMMANDER_ROLE_TYPES:
        raise ValueError("Μη έγκυρος τύπος Διοικητή/Κυβερνήτη.")

    return normalized


def build_service_units_list_page_context() -> dict[str, Any]:
    """
    Build template context for the ServiceUnits list page.
    """
    units = ServiceUnit.query.order_by(ServiceUnit.description.asc()).all()
    return {"units": units}


def build_service_units_roles_page_context() -> dict[str, Any]:
    """
    Build template context for the ServiceUnits role-assignment list page.
    """
    units = ServiceUnit.query.order_by(ServiceUnit.description.asc()).all()
    return {"units": units}


def build_service_unit_form_page_context(
    *,
    unit: ServiceUnit | None,
    form_title: str,
    is_create: bool,
) -> dict[str, Any]:
    """
    Build template context for create/edit ServiceUnit form pages.
    """
    return {
        "unit": unit,
        "form_title": form_title,
        "is_create": is_create,
        "commander_role_type_options": ("Διοικητής", "Κυβερνήτης"),
    }


def build_service_unit_roles_form_page_context(
    *,
    unit: ServiceUnit,
    form_title: str,
) -> dict[str, Any]:
    """
    Build template context for ServiceUnit Manager/Deputy assignment page.
    """
    return {
        "unit": unit,
        "personnel_list": _active_personnel_for_dropdown(),
        "form_title": form_title,
    }


def execute_create_service_unit(form_data: Mapping[str, Any]) -> OperationResult:
    """
    Validate and create a new ServiceUnit.
    """
    description = (form_data.get("description") or "").strip()
    code = (form_data.get("code") or "").strip()
    short_name = (form_data.get("short_name") or "").strip()
    aahit = (form_data.get("aahit") or "").strip()

    email = (form_data.get("email") or "").strip()
    address = (form_data.get("address") or "").strip()
    region = (form_data.get("region") or "").strip()
    prefecture = (form_data.get("prefecture") or "").strip()

    phone = (form_data.get("phone") or "").strip()

    commander = (form_data.get("commander") or "").strip()
    commander_role_type_raw = form_data.get("commander_role_type")

    application_administrator = (form_data.get("curator") or "").strip()
    application_admin_directory = (form_data.get("application_admin_directory") or "").strip()

    supply_officer = (form_data.get("supply_officer") or "").strip()

    if not description:
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Η περιγραφή Υπηρεσίας είναι υποχρεωτική.", "danger"),),
        )

    try:
        commander_role_type = _normalize_commander_role_type(commander_role_type_raw)
    except ValueError as exc:
        return OperationResult(
            ok=False,
            flashes=(FlashMessage(str(exc), "danger"),),
        )

    if ServiceUnit.query.filter_by(description=description).first():
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Υπάρχει ήδη Υπηρεσία με αυτή την περιγραφή.", "danger"),),
        )

    if code and ServiceUnit.query.filter_by(code=code).first():
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Υπάρχει ήδη Υπηρεσία με αυτόν τον κωδικό.", "danger"),),
        )

    if short_name and ServiceUnit.query.filter_by(short_name=short_name).first():
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Υπάρχει ήδη Υπηρεσία με αυτή τη συντομογραφία.", "danger"),),
        )

    unit = ServiceUnit(
        description=description,
        code=code or None,
        short_name=short_name or None,
        aahit=aahit or None,
        email=email or None,
        address=address or None,
        region=region or None,
        prefecture=prefecture or None,
        phone=phone or None,
        commander=commander or None,
        commander_role_type=commander_role_type,
        curator=application_administrator or None,
        application_admin_directory=application_admin_directory or None,
        supply_officer=supply_officer or None,
        manager_personnel_id=None,
        deputy_personnel_id=None,
    )

    db.session.add(unit)
    db.session.flush()
    log_action(entity=unit, action="CREATE", before=None, after=serialize_model(unit))
    db.session.commit()

    return OperationResult(
        ok=True,
        flashes=(FlashMessage("Η Υπηρεσία δημιουργήθηκε.", "success"),),
        entity_id=unit.id,
    )


def execute_edit_service_unit_info(
    unit: ServiceUnit,
    form_data: Mapping[str, Any],
) -> OperationResult:
    """
    Validate and update basic ServiceUnit information.
    """
    before = serialize_model(unit)

    description = (form_data.get("description") or "").strip()
    code = (form_data.get("code") or "").strip()
    short_name = (form_data.get("short_name") or "").strip()
    aahit = (form_data.get("aahit") or "").strip()

    email = (form_data.get("email") or "").strip()
    address = (form_data.get("address") or "").strip()
    region = (form_data.get("region") or "").strip()
    prefecture = (form_data.get("prefecture") or "").strip()

    phone = (form_data.get("phone") or "").strip()

    commander = (form_data.get("commander") or "").strip()
    commander_role_type_raw = form_data.get("commander_role_type")

    application_administrator = (form_data.get("curator") or "").strip()
    application_admin_directory = (form_data.get("application_admin_directory") or "").strip()

    supply_officer = (form_data.get("supply_officer") or "").strip()

    if not description:
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Η περιγραφή Υπηρεσίας είναι υποχρεωτική.", "danger"),),
        )

    try:
        commander_role_type = _normalize_commander_role_type(commander_role_type_raw)
    except ValueError as exc:
        return OperationResult(
            ok=False,
            flashes=(FlashMessage(str(exc), "danger"),),
        )

    duplicate_desc = ServiceUnit.query.filter(
        ServiceUnit.description == description,
        ServiceUnit.id != unit.id,
    ).first()
    if duplicate_desc:
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Υπάρχει ήδη άλλη Υπηρεσία με αυτή την περιγραφή.", "danger"),),
        )

    if code:
        duplicate_code = ServiceUnit.query.filter(
            ServiceUnit.code == code,
            ServiceUnit.id != unit.id,
        ).first()
        if duplicate_code:
            return OperationResult(
                ok=False,
                flashes=(FlashMessage("Υπάρχει ήδη άλλη Υπηρεσία με αυτόν τον κωδικό.", "danger"),),
            )

    if short_name:
        duplicate_short_name = ServiceUnit.query.filter(
            ServiceUnit.short_name == short_name,
            ServiceUnit.id != unit.id,
        ).first()
        if duplicate_short_name:
            return OperationResult(
                ok=False,
                flashes=(FlashMessage("Υπάρχει ήδη άλλη Υπηρεσία με αυτή τη συντομογραφία.", "danger"),),
            )

    unit.description = description
    unit.code = code or None
    unit.short_name = short_name or None
    unit.aahit = aahit or None

    unit.email = email or None
    unit.address = address or None
    unit.region = region or None
    unit.prefecture = prefecture or None

    unit.phone = phone or None

    unit.commander = commander or None
    unit.commander_role_type = commander_role_type

    # Business label: "Διαχειριστής Εφαρμογής"
    # Persisted storage: existing `curator` column retained intentionally.
    unit.curator = application_administrator or None
    unit.application_admin_directory = application_admin_directory or None

    unit.supply_officer = supply_officer or None

    db.session.flush()
    log_action(entity=unit, action="UPDATE", before=before, after=serialize_model(unit))
    db.session.commit()

    return OperationResult(
        ok=True,
        flashes=(FlashMessage("Η Υπηρεσία ενημερώθηκε.", "success"),),
        entity_id=unit.id,
    )


def execute_assign_service_unit_roles(
    unit: ServiceUnit,
    form_data: Mapping[str, Any],
) -> OperationResult:
    before = serialize_model(unit)

    manager_personnel_id_raw = form_data.get("manager_personnel_id")
    deputy_personnel_id_raw = form_data.get("deputy_personnel_id")

    manager_personnel_id = None
    deputy_personnel_id = None

    if manager_personnel_id_raw:
        try:
            manager_personnel_id = int(manager_personnel_id_raw)
        except (TypeError, ValueError):
            return OperationResult(
                ok=False,
                flashes=(FlashMessage("Μη έγκυρος Manager.", "danger"),),
            )

    if deputy_personnel_id_raw:
        try:
            deputy_personnel_id = int(deputy_personnel_id_raw)
        except (TypeError, ValueError):
            return OperationResult(
                ok=False,
                flashes=(FlashMessage("Μη έγκυρος Deputy.", "danger"),),
            )

    if (
        manager_personnel_id is not None
        and deputy_personnel_id is not None
        and manager_personnel_id == deputy_personnel_id
    ):
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Δεν γίνεται ο Manager και ο Deputy να είναι το ίδιο πρόσωπο.", "danger"),),
        )

    if manager_personnel_id is not None:
        manager_person = Personnel.query.filter_by(
            id=manager_personnel_id,
            is_active=True,
        ).first()
        if not manager_person:
            return OperationResult(
                ok=False,
                flashes=(FlashMessage("Ο επιλεγμένος Manager δεν είναι έγκυρο ενεργό προσωπικό.", "danger"),),
            )

        manager_used_elsewhere = ServiceUnit.query.filter(
            ServiceUnit.manager_personnel_id == manager_personnel_id,
            ServiceUnit.id != unit.id,
        ).first()
        if manager_used_elsewhere:
            return OperationResult(
                ok=False,
                flashes=(FlashMessage("Το συγκεκριμένο προσωπικό είναι ήδη Manager σε άλλη Υπηρεσία.", "danger"),),
            )

    if deputy_personnel_id is not None:
        deputy_person = Personnel.query.filter_by(
            id=deputy_personnel_id,
            is_active=True,
        ).first()
        if not deputy_person:
            return OperationResult(
                ok=False,
                flashes=(FlashMessage("Ο επιλεγμένος Deputy δεν είναι έγκυρο ενεργό προσωπικό.", "danger"),),
            )

        deputy_used_elsewhere = ServiceUnit.query.filter(
            ServiceUnit.deputy_personnel_id == deputy_personnel_id,
            ServiceUnit.id != unit.id,
        ).first()
        if deputy_used_elsewhere:
            return OperationResult(
                ok=False,
                flashes=(FlashMessage("Το συγκεκριμένο προσωπικό είναι ήδη Deputy σε άλλη Υπηρεσία.", "danger"),),
            )

    unit.manager_personnel_id = manager_personnel_id
    unit.deputy_personnel_id = deputy_personnel_id

    db.session.flush()
    log_action(entity=unit, action="UPDATE", before=before, after=serialize_model(unit))
    db.session.commit()

    return OperationResult(
        ok=True,
        flashes=(FlashMessage("Οι ρόλοι Υπηρεσίας ενημερώθηκαν.", "success"),),
        entity_id=unit.id,
    )


def execute_delete_service_unit(unit: ServiceUnit) -> OperationResult:
    """
    Delete a ServiceUnit using a defensive, SQLite-safe strategy.

    WHY THIS EXISTS
    ---------------
    In development with SQLite, cascades involving non-nullable relationships
    may behave less predictably than on PostgreSQL. To keep the project stable,
    we explicitly detach or delete related rows before deleting the ServiceUnit.

    DELETE STRATEGY
    ---------------
    1. Audit snapshot of the ServiceUnit
    2. Delete related ProcurementCommittee rows
    3. Clear nullable references from Personnel
    4. Clear manager/deputy references on User rows
    5. Clear manager/deputy references on the ServiceUnit itself
    6. Abort if Procurements still point to the ServiceUnit
    7. Delete the ServiceUnit

    IMPORTANT
    ---------
    The provided Personnel model contains only `service_unit_id` among the
    service-unit-scoped nullable references. Therefore we clear only that field
    here, avoiding references to non-existent columns.
    """
    before = serialize_model(unit)

    # Delete committees first, preserving audit logs.
    committees = ProcurementCommittee.query.filter_by(service_unit_id=unit.id).all()
    for committee in committees:
        committee_before = serialize_model(committee)
        db.session.delete(committee)
        db.session.flush()
        log_action(entity=committee, action="DELETE", before=committee_before, after=None)

    # Clear Personnel references that are nullable and scoped to this ServiceUnit.
    Personnel.query.filter_by(service_unit_id=unit.id).update(
        {"service_unit_id": None},
        synchronize_session=False,
    )

    # Clear user role pointers that may target this ServiceUnit.
    users = User.query.filter_by(service_unit_id=unit.id).all()
    for user in users:
        user.service_unit_id = None

    # Clear manager/deputy references on the unit itself before delete.
    unit.manager_personnel_id = None
    unit.deputy_personnel_id = None
    db.session.flush()

    # Defensive block: do not delete if Procurements still point to this ServiceUnit.
    procurements_exist = Procurement.query.filter_by(service_unit_id=unit.id).first() is not None
    if procurements_exist:
        db.session.rollback()
        return OperationResult(
            ok=False,
            flashes=(
                FlashMessage(
                    "Η Υπηρεσία δεν μπορεί να διαγραφεί γιατί υπάρχουν συνδεδεμένες προμήθειες.",
                    "danger",
                ),
            ),
        )

    db.session.delete(unit)
    db.session.flush()
    log_action(entity=unit, action="DELETE", before=before, after=None)
    db.session.commit()

    return OperationResult(
        ok=True,
        flashes=(FlashMessage("Η Υπηρεσία διαγράφηκε.", "success"),),
    )


def execute_import_service_units(file_storage: Any) -> OperationResult:
    """
    Import ServiceUnits from an uploaded Excel file.

    ACCEPTED HEADERS
    ----------------
    Required:
    - Περιγραφή / description

    Optional:
    - Κωδικός / code
    - Συντομογραφία / short_name
    - ΑΑΗΤ / aahit

    - Email / email / e-mail / υπηρεσιακό email
    - Διεύθυνση / address
    - Περιοχή / region
    - Νομός / prefecture
    - Τηλέφωνο / phone

    - Διοικητής/Κυβερνήτης / commander
    - Διοικητής / commander
    - Κυβερνήτης / commander

    - Τύπος Διοικητή/Κυβερνήτη / commander_role_type
    - Τύπος Διοικητή - Κυβερνήτη / commander_role_type
    - commander role type

    - Διαχειριστής Εφαρμογής / curator
    - Επιμελητής / curator
      (retained for backward-compatible import support)

    - ΔΙΕΥΘΥΝΣΗ Διαχειριστή Εφαρμογής / application_admin_directory
    - Διεύθυνση Διαχειριστή Εφαρμογής / application_admin_directory
    - application_admin_directory

    - Υπόλογος Εφοδιασμού / supply_officer

    IMPORT POLICY
    -------------
    - Only .xlsx is accepted.
    - Description remains required.
    - Duplicate checks remain:
      description, code, short_name.
    - commander_role_type is validated if provided.
    """
    if not file_storage or not getattr(file_storage, "filename", None):
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Δεν επιλέχθηκε αρχείο.", "danger"),),
        )

    filename = str(file_storage.filename or "").lower()
    if not filename.endswith(".xlsx"):
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Επιτρέπεται μόνο αρχείο .xlsx", "danger"),),
        )

    try:
        import openpyxl

        workbook = openpyxl.load_workbook(file_storage, data_only=True)
        worksheet = workbook.active
    except Exception:
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Αποτυχία ανάγνωσης Excel. Ελέγξτε το αρχείο.", "danger"),),
        )

    try:
        header_cells = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    except StopIteration:
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Το Excel είναι κενό.", "danger"),),
        )

    idx_map = build_header_index(header_cells)

    desc_idx = idx_map.get("περιγραφη", idx_map.get("description"))
    code_idx = idx_map.get("κωδικος", idx_map.get("code"))
    short_idx = idx_map.get("συντομογραφια", idx_map.get("short name", idx_map.get("short_name")))
    aahit_idx = idx_map.get("ααητ", idx_map.get("aahit"))

    email_idx = idx_map.get(
        "email",
        idx_map.get("e-mail", idx_map.get("υπηρεσιακο email", idx_map.get("υπηρεσιακο e-mail"))),
    )

    address_idx = idx_map.get("διευθυνση", idx_map.get("address"))
    region_idx = idx_map.get("περιοχη", idx_map.get("region"))
    prefecture_idx = idx_map.get("νομος", idx_map.get("prefecture"))

    phone_idx = idx_map.get("τηλεφωνο", idx_map.get("phone"))

    commander_idx = idx_map.get(
        "διοικητης/κυβερνητης",
        idx_map.get(
            "διοικητης",
            idx_map.get("κυβερνητης", idx_map.get("commander")),
        ),
    )

    commander_role_type_idx = idx_map.get(
        "τυπος διοικητη/κυβερνητη",
        idx_map.get(
            "τυπος διοικητη - κυβερνητη",
            idx_map.get(
                "τυπος διοικητη κυ-βερνητη",
                idx_map.get("commander_role_type", idx_map.get("commander role type")),
            ),
        ),
    )

    curator_idx = idx_map.get(
        "διαχειριστης εφαρμογης",
        idx_map.get("curator", idx_map.get("επιμελητης")),
    )

    application_admin_directory_idx = idx_map.get(
        "διευθυνση διαχειριστη εφαρμογης",
        idx_map.get(
            "διευθυνση διαχειριστη εφαρμογης",
            idx_map.get(
                "application_admin_directory",
                idx_map.get("διευθυνση διαχειριστη εφαρμογης"),
            ),
        ),
    )

    # Support also uppercase business-style heading "ΔΙΕΥΘΥΝΣΗ Διαχειριστή Εφαρμογής"
    if application_admin_directory_idx is None:
        application_admin_directory_idx = idx_map.get("διευθυνση διαχειριστη εφαρμογης")

    supply_officer_idx = idx_map.get("υπολογος εφοδιασμου", idx_map.get("supply_officer"))

    if desc_idx is None:
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Το Excel πρέπει να έχει στήλη 'Περιγραφή' (ή 'description').", "danger"),),
        )

    inserted_units: list[ServiceUnit] = []
    skipped_missing = 0
    skipped_duplicate = 0
    skipped_invalid_role_type = 0

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        description = safe_cell_str(cell_at(row, desc_idx))
        if not description:
            skipped_missing += 1
            continue

        code = safe_cell_str(cell_at(row, code_idx)) or None
        short_name = safe_cell_str(cell_at(row, short_idx)) or None

        duplicate_exists = ServiceUnit.query.filter_by(description=description).first() is not None

        if not duplicate_exists and code:
            duplicate_exists = ServiceUnit.query.filter_by(code=code).first() is not None

        if not duplicate_exists and short_name:
            duplicate_exists = ServiceUnit.query.filter_by(short_name=short_name).first() is not None

        if duplicate_exists:
            skipped_duplicate += 1
            continue

        commander_role_type_raw = safe_cell_str(cell_at(row, commander_role_type_idx)) or None
        try:
            commander_role_type = _normalize_commander_role_type(commander_role_type_raw)
        except ValueError:
            skipped_invalid_role_type += 1
            continue

        unit = ServiceUnit(
            description=description,
            code=code,
            short_name=short_name,
            aahit=safe_cell_str(cell_at(row, aahit_idx)) or None,
            email=safe_cell_str(cell_at(row, email_idx)) or None,
            address=safe_cell_str(cell_at(row, address_idx)) or None,
            region=safe_cell_str(cell_at(row, region_idx)) or None,
            prefecture=safe_cell_str(cell_at(row, prefecture_idx)) or None,
            phone=safe_cell_str(cell_at(row, phone_idx)) or None,
            commander=safe_cell_str(cell_at(row, commander_idx)) or None,
            commander_role_type=commander_role_type,
            curator=safe_cell_str(cell_at(row, curator_idx)) or None,
            application_admin_directory=safe_cell_str(cell_at(row, application_admin_directory_idx)) or None,
            supply_officer=safe_cell_str(cell_at(row, supply_officer_idx)) or None,
            manager_personnel_id=None,
            deputy_personnel_id=None,
        )
        db.session.add(unit)
        inserted_units.append(unit)

    if not inserted_units:
        details = []
        if skipped_missing:
            details.append(f"{skipped_missing} ελλιπείς")
        if skipped_duplicate:
            details.append(f"{skipped_duplicate} διπλότυπες")
        if skipped_invalid_role_type:
            details.append(f"{skipped_invalid_role_type} με μη έγκυρο τύπο Διοικητή/Κυβερνήτη")

        details_text = ", ".join(details) if details else "χωρίς έγκυρες εγγραφές"
        return OperationResult(
            ok=False,
            flashes=(FlashMessage(f"Δεν εισήχθησαν εγγραφές. Έλεγχος αρχείου: {details_text}.", "warning"),),
        )

    db.session.flush()
    for unit in inserted_units:
        log_action(entity=unit, action="CREATE", before=None, after=serialize_model(unit))
    db.session.commit()

    return OperationResult(
        ok=True,
        flashes=(
            FlashMessage(
                f"Εισαγωγή ολοκληρώθηκε: {len(inserted_units)} νέες Υπηρεσίες. "
                f"Παραλείφθηκαν: {skipped_missing} (ελλιπή), "
                f"{skipped_duplicate} (διπλότυπα), "
                f"{skipped_invalid_role_type} (μη έγκυρος τύπος Διοικητή/Κυβερνήτη).",
                "success",
            ),
        ),
    )
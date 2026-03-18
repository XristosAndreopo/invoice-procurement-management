"""
app/services/admin/personnel.py

Focused personnel page/use-case services for the admin blueprint.

PURPOSE
-------
This module contains person-centric administration services for:
- listing Personnel
- importing Personnel from Excel
- creating Personnel
- editing Personnel
- deleting Personnel

IMPORTANT DOMAIN RULE
---------------------
Directory/Department assignment must NOT be edited from the Personnel form page.

Organizational placement is managed centrally only from:
    /admin/organization-setup

Therefore:
- create/edit/import personnel must not accept or persist directory_id
- create/edit/import personnel must not accept or persist department_id
- the Personnel page remains person-centric
- directory/department placement belongs to the separate
  `PersonnelDepartmentAssignment` model

MODEL-COMPATIBILITY NOTE
------------------------
The actual `Personnel` ORM model supports:
- agm
- aem
- rank
- specialty
- first_name
- last_name
- is_active
- service_unit_id

It does NOT support:
- directory_id
- department_id

This module must stay aligned with that schema. Passing unsupported keyword
arguments to `Personnel(...)` or assigning missing attributes on an existing
`Personnel` instance will raise runtime errors.

ARCHITECTURAL BOUNDARY
----------------------
This module MAY:
- query ORM rows
- validate submitted form/import data
- create/update/delete Personnel rows
- flush/commit DB state
- emit structured service results

This module MUST NOT:
- render templates
- redirect
- flash directly
- implement organization placement orchestration that belongs elsewhere
"""

from __future__ import annotations

from collections.abc import Mapping
from typing import Any

from sqlalchemy.exc import IntegrityError

from ...audit import log_action, serialize_model
from ...extensions import db
from ...models import Personnel
from ..organization import (
    effective_scope_service_unit_id_for_manager_or_none,
    match_service_unit_from_text,
    service_units_for_dropdown,
    validate_service_unit_required,
)
from ..shared.excel_imports import build_header_index, cell_at, safe_cell_str
from ..shared.operation_results import FlashMessage, OperationResult
from ..shared.parsing import parse_optional_int


def build_personnel_list_page_context() -> dict[str, Any]:
    """
    Build template context for the personnel list page.

    RETURNS
    -------
    dict[str, Any]
        Template payload containing the visible Personnel rows.

    SCOPE RULE
    ----------
    - admins may see all Personnel
    - managers are restricted to their effective service-unit scope
    """
    query = Personnel.query.options(
        db.joinedload(Personnel.service_unit),
    )

    scope_service_unit_id = effective_scope_service_unit_id_for_manager_or_none()
    if scope_service_unit_id:
        query = query.filter(Personnel.service_unit_id == scope_service_unit_id)

    personnel = (
        query.order_by(
            Personnel.rank.asc(),
            Personnel.last_name.asc(),
            Personnel.first_name.asc(),
        ).all()
    )

    return {
        "personnel": personnel,
    }


def build_personnel_form_page_context(
    *,
    person: Personnel | None,
    form_title: str,
) -> dict[str, Any]:
    """
    Build template context for both create/edit personnel forms.

    PARAMETERS
    ----------
    person:
        Existing Personnel row for edit mode, or None for create mode.
    form_title:
        Page/form title to render.

    RETURNS
    -------
    dict[str, Any]
        Template context for the personnel form page.

    IMPORTANT
    ---------
    Directory / Department are no longer edited here.
    They are managed only from Organization Setup.
    """
    return {
        "person": person,
        "form_title": form_title,
        "service_units": service_units_for_dropdown(),
    }


def execute_import_personnel(file_storage: Any) -> OperationResult:
    """
    Import Personnel rows from an uploaded Excel file.

    PARAMETERS
    ----------
    file_storage:
        Uploaded Flask file object.

    RETURNS
    -------
    OperationResult
        Import outcome, including summary flash text.

    REQUIRED EXCEL COLUMNS
    ----------------------
    The first row must contain headers that map to:
    - ΑΓΜ
    - ΟΝΟΜΑ
    - ΕΠΩΝΥΜΟ

    OPTIONAL COLUMNS
    ----------------
    - ΑΕΜ
    - ΒΑΘΜΟΣ
    - ΕΙΔΙΚΟΤΗΤΑ
    - ΥΠΗΡΕΣΙΑ

    IMPORTANT DOMAIN RULE
    ---------------------
    This import is person-centric only.
    It may assign `service_unit_id` if a valid ServiceUnit match is found,
    but it must NOT attempt to assign directory/department fields on Personnel.
    """
    if not file_storage or not getattr(file_storage, "filename", None):
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Δεν επιλέχθηκε αρχείο.", "danger"),),
        )

    filename = str(file_storage.filename or "").lower()
    if not filename.endswith(".xlsx"):
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Επιτρέπεται μόνο αρχείο .xlsx", "danger"),),
        )

    try:
        import openpyxl

        workbook = openpyxl.load_workbook(file_storage, data_only=True)
        worksheet = workbook.active
    except Exception:
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Αποτυχία ανάγνωσης Excel. Ελέγξτε το αρχείο.", "danger"),),
        )

    try:
        header_cells = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    except StopIteration:
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Το Excel είναι κενό.", "danger"),),
        )

    idx_map = build_header_index(header_cells)

    agm_idx = idx_map.get("αγμ", idx_map.get("agm"))
    first_idx = idx_map.get("ονομα", idx_map.get("first name", idx_map.get("first_name")))
    last_idx = idx_map.get("επωνυμο", idx_map.get("last name", idx_map.get("last_name")))
    aem_idx = idx_map.get("αεμ", idx_map.get("aem"))
    rank_idx = idx_map.get("βαθμος", idx_map.get("rank"))
    spec_idx = idx_map.get("ειδικοτητα", idx_map.get("specialty"))
    service_idx = idx_map.get("υπηρεσια", idx_map.get("service"))

    if agm_idx is None or first_idx is None or last_idx is None:
        return OperationResult(
            ok=False,
            flashes=(
                FlashMessage(
                    "Το Excel πρέπει να έχει στήλες: ΑΓΜ, ΟΝΟΜΑ, ΕΠΩΝΥΜΟ (1η γραμμή).",
                    "danger",
                ),
            ),
        )

    inserted_people: list[Personnel] = []
    skipped_missing = 0
    skipped_duplicate = 0
    skipped_bad_service = 0

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        agm = safe_cell_str(cell_at(row, agm_idx))
        first_name = safe_cell_str(cell_at(row, first_idx))
        last_name = safe_cell_str(cell_at(row, last_idx))

        if not agm or not first_name or not last_name:
            skipped_missing += 1
            continue

        if Personnel.query.filter_by(agm=agm).first():
            skipped_duplicate += 1
            continue

        service_unit_id = None
        if service_idx is not None:
            service_val = safe_cell_str(cell_at(row, service_idx))
            if service_val:
                service_unit = match_service_unit_from_text(service_val)
                if not service_unit:
                    skipped_bad_service += 1
                    continue
                service_unit_id = service_unit.id

        # IMPORTANT:
        # Create Personnel only with fields that actually exist on the model.
        # Directory/Department placement is intentionally excluded from this
        # page/use-case and belongs to PersonnelDepartmentAssignment.
        person = Personnel(
            agm=agm,
            aem=safe_cell_str(cell_at(row, aem_idx)) or None,
            rank=safe_cell_str(cell_at(row, rank_idx)) or None,
            specialty=safe_cell_str(cell_at(row, spec_idx)) or None,
            first_name=first_name,
            last_name=last_name,
            is_active=True,
            service_unit_id=service_unit_id,
        )
        db.session.add(person)
        inserted_people.append(person)

    if not inserted_people:
        return OperationResult(
            ok=False,
            flashes=(
                FlashMessage(
                    "Δεν εισήχθησαν εγγραφές. Ελέγξτε required πεδία/διπλότυπα/Υπηρεσία.",
                    "warning",
                ),
            ),
        )

    db.session.flush()

    for person in inserted_people:
        log_action(person, "CREATE", before=None, after=serialize_model(person))

    db.session.commit()

    return OperationResult(
        ok=True,
        flashes=(
            FlashMessage(
                (
                    f"Εισαγωγή ολοκληρώθηκε: {len(inserted_people)} νέες εγγραφές. "
                    f"Παραλείφθηκαν: {skipped_missing} (ελλιπή), "
                    f"{skipped_duplicate} (διπλότυπα ΑΓΜ), "
                    f"{skipped_bad_service} (μη έγκυρη Υπηρεσία)."
                ),
                "success",
            ),
        ),
    )


def execute_create_personnel(
    form_data: Mapping[str, Any],
    *,
    is_admin: bool,
    current_service_unit_id: int | None,
) -> OperationResult:
    """
    Validate and create a Personnel row.

    PARAMETERS
    ----------
    form_data:
        Submitted form mapping.
    is_admin:
        True when the acting user is admin.
    current_service_unit_id:
        Current acting user's service-unit scope for manager-restricted create.

    RETURNS
    -------
    OperationResult
        Creation outcome.

    IMPORTANT
    ---------
    Directory / Department are no longer set from this form.
    """
    agm = (form_data.get("agm") or "").strip()
    aem = (form_data.get("aem") or "").strip()
    rank = (form_data.get("rank") or "").strip()
    specialty = (form_data.get("specialty") or "").strip()
    first_name = (form_data.get("first_name") or "").strip()
    last_name = (form_data.get("last_name") or "").strip()

    if is_admin:
        service_unit_id = parse_optional_int(form_data.get("service_unit_id"))
    else:
        service_unit_id = current_service_unit_id

    if not agm or not first_name or not last_name:
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("ΑΓΜ, Όνομα και Επώνυμο είναι υποχρεωτικά.", "danger"),),
        )

    if Personnel.query.filter_by(agm=agm).first():
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Υπάρχει ήδη προσωπικό με αυτό το ΑΓΜ.", "danger"),),
        )

    if not validate_service_unit_required(service_unit_id):
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Η Υπηρεσία είναι υποχρεωτική και πρέπει να είναι έγκυρη.", "danger"),),
        )

    # IMPORTANT:
    # Personnel is person-centric. Only persist actual model fields here.
    person = Personnel(
        agm=agm,
        aem=aem or None,
        rank=rank or None,
        specialty=specialty or None,
        first_name=first_name,
        last_name=last_name,
        is_active=True,
        service_unit_id=service_unit_id,
    )

    db.session.add(person)
    db.session.flush()
    log_action(person, "CREATE", before=None, after=serialize_model(person))
    db.session.commit()

    return OperationResult(
        ok=True,
        flashes=(FlashMessage("Το προσωπικό καταχωρήθηκε.", "success"),),
        entity_id=person.id,
    )


def execute_edit_personnel(
    person: Personnel,
    form_data: Mapping[str, Any],
    *,
    is_admin: bool,
    current_service_unit_id: int | None,
) -> OperationResult:
    """
    Validate and update an existing Personnel row.

    PARAMETERS
    ----------
    person:
        Existing Personnel row to update.
    form_data:
        Submitted form mapping.
    is_admin:
        True when the acting user is admin.
    current_service_unit_id:
        Current acting user's service-unit scope for manager-restricted edit.

    RETURNS
    -------
    OperationResult
        Update outcome.

    IMPORTANT
    ---------
    Directory / Department are no longer edited here.

    MODEL RULE
    ----------
    Since `Personnel` does not define `directory_id` / `department_id`, this
    service must not assign those attributes at all.
    """
    before_snapshot = serialize_model(person)

    agm = (form_data.get("agm") or "").strip()
    aem = (form_data.get("aem") or "").strip()
    rank = (form_data.get("rank") or "").strip()
    specialty = (form_data.get("specialty") or "").strip()
    first_name = (form_data.get("first_name") or "").strip()
    last_name = (form_data.get("last_name") or "").strip()

    if is_admin:
        service_unit_id = parse_optional_int(form_data.get("service_unit_id"))
    else:
        service_unit_id = current_service_unit_id

    is_active = bool(form_data.get("is_active"))

    if not agm or not first_name or not last_name:
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("ΑΓΜ, Όνομα και Επώνυμο είναι υποχρεωτικά.", "danger"),),
        )

    existing = Personnel.query.filter(
        Personnel.agm == agm,
        Personnel.id != person.id,
    ).first()
    if existing:
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Υπάρχει ήδη προσωπικό με αυτό το ΑΓΜ.", "danger"),),
        )

    if not validate_service_unit_required(service_unit_id):
        return OperationResult(
            ok=False,
            flashes=(FlashMessage("Η Υπηρεσία είναι υποχρεωτική και πρέπει να είναι έγκυρη.", "danger"),),
        )

    person.agm = agm
    person.aem = aem or None
    person.rank = rank or None
    person.specialty = specialty or None
    person.first_name = first_name
    person.last_name = last_name
    person.service_unit_id = service_unit_id
    person.is_active = is_active

    # IMPORTANT:
    # Do not touch directory/department placement here.
    # Placement is managed centrally through organization setup and
    # PersonnelDepartmentAssignment, not via inline Personnel fields.

    db.session.flush()
    log_action(person, "UPDATE", before=before_snapshot, after=serialize_model(person))
    db.session.commit()

    return OperationResult(
        ok=True,
        flashes=(FlashMessage("Το προσωπικό ενημερώθηκε.", "success"),),
        entity_id=person.id,
    )


def execute_delete_personnel(person: Personnel) -> OperationResult:
    """
    Delete a Personnel row if no blocking references exist.

    PARAMETERS
    ----------
    person:
        Target Personnel row.

    RETURNS
    -------
    OperationResult
        Deletion outcome.

    FAILURE MODE
    ------------
    If the row is already referenced elsewhere, the delete is rolled back and a
    user-facing error message is returned.
    """
    before = serialize_model(person)

    try:
        db.session.delete(person)
        db.session.flush()
        log_action(entity=person, action="DELETE", before=before)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return OperationResult(
            ok=False,
            flashes=(
                FlashMessage(
                    "Το προσωπικό δεν μπορεί να διαγραφεί γιατί χρησιμοποιείται ήδη αλλού στο σύστημα.",
                    "danger",
                ),
            ),
            entity_id=person.id,
        )

    return OperationResult(
        ok=True,
        flashes=(FlashMessage("Το προσωπικό διαγράφηκε.", "success"),),
        entity_id=person.id,
    )
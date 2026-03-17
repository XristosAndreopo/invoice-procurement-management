"""
app/services/settings/master_data_admin.py

Focused master-data admin services for settings routes.

PURPOSE
-------
This module extracts non-HTTP orchestration from the fat admin/master-data
routes inside `app/blueprints/settings/routes.py`, specifically:

- /settings/ale-kae
- /settings/ale-kae/import
- /settings/cpv
- /settings/cpv/import
- /settings/options/*
- /settings/income-tax
- /settings/withholding-profiles

WHY THIS FILE EXISTS
--------------------
The current settings blueprint mixes already-thin route groups
(ServiceUnits/Suppliers) with several still-fat CRUD/import flows.
Those flows perform:
- validation
- object loading
- persistence
- audit logging
- page-context assembly

That orchestration belongs in the service/use-case layer rather than in route
handlers.

DESIGN
------
- function-first
- one focused module for settings master-data administration
- no generic CRUD framework
- small explicit helper functions per sub-area
"""

from __future__ import annotations

from collections.abc import Mapping
from decimal import Decimal
from sqlalchemy.exc import IntegrityError
from typing import Any

from ...audit import log_action, serialize_model
from ...extensions import db
from ...models import AleKae, Cpv, IncomeTaxRule, OptionCategory, OptionValue, WithholdingProfile
from ..shared.excel_imports import build_header_index, cell_at, safe_cell_str
from ..master_data_service import get_all_option_rows, get_option_category_by_key
from ..shared.operation_results import FlashMessage, OperationResult
from ..shared.parsing import parse_decimal, parse_optional_int


def _get_or_create_category(key: str, label: str) -> OptionCategory:
    """
    Ensure an OptionCategory row exists and return it.

    NOTE
    ----
    This preserves the current self-healing behavior of the settings option
    pages when seed data has not been run yet.
    """
    category = get_option_category_by_key(key)
    if category:
        if category.label != label:
            category.label = label
            db.session.commit()
        return category

    category = OptionCategory(key=key, label=label)
    db.session.add(category)
    db.session.commit()
    return category


# ----------------------------------------------------------------------
# ALE-KAE
# ----------------------------------------------------------------------
def build_ale_kae_page_context() -> dict[str, Any]:
    """
    Build template context for the ALE-KAE page.
    """
    rows = AleKae.query.order_by(AleKae.ale.asc()).all()
    return {"rows": rows}


def execute_ale_kae_action(form_data: Mapping[str, object]) -> OperationResult:
    """
    Execute create/update/delete action for ALE-KAE rows.
    """
    action = (form_data.get("action") or "").strip()

    if action == "create":
        ale = (form_data.get("ale") or "").strip()
        old_kae = (form_data.get("old_kae") or "").strip() or None
        description = (form_data.get("description") or "").strip() or None
        responsibility = (form_data.get("responsibility") or "").strip() or None

        if not ale:
            return OperationResult(False, (FlashMessage("Το ΑΛΕ είναι υποχρεωτικό.", "danger"),))

        if AleKae.query.filter_by(ale=ale).first():
            return OperationResult(False, (FlashMessage("Υπάρχει ήδη εγγραφή με αυτό το ΑΛΕ.", "danger"),))

        row = AleKae(
            ale=ale,
            old_kae=old_kae,
            description=description,
            responsibility=responsibility,
        )
        db.session.add(row)
        db.session.flush()
        log_action(entity=row, action="CREATE", before=None, after=serialize_model(row))
        db.session.commit()

        return OperationResult(True, (FlashMessage("Η εγγραφή ΑΛΕ-ΚΑΕ προστέθηκε.", "success"),), entity_id=row.id)

    if action == "update":
        row_id = parse_optional_int((form_data.get("id") or "").strip())
        if row_id is None:
            return OperationResult(False, (FlashMessage("Μη έγκυρη εγγραφή.", "danger"),))

        row = AleKae.query.get_or_404(row_id)
        before = serialize_model(row)

        ale = (form_data.get("ale") or "").strip()
        old_kae = (form_data.get("old_kae") or "").strip() or None
        description = (form_data.get("description") or "").strip() or None
        responsibility = (form_data.get("responsibility") or "").strip() or None

        if not ale:
            return OperationResult(False, (FlashMessage("Το ΑΛΕ είναι υποχρεωτικό.", "danger"),))

        exists = AleKae.query.filter(AleKae.ale == ale, AleKae.id != row.id).first()
        if exists:
            return OperationResult(False, (FlashMessage("Υπάρχει ήδη άλλη εγγραφή με αυτό το ΑΛΕ.", "danger"),))

        row.ale = ale
        row.old_kae = old_kae
        row.description = description
        row.responsibility = responsibility

        db.session.flush()
        log_action(entity=row, action="UPDATE", before=before, after=serialize_model(row))
        db.session.commit()

        return OperationResult(True, (FlashMessage("Η εγγραφή ενημερώθηκε.", "success"),), entity_id=row.id)

    if action == "delete":
        row_id = parse_optional_int((form_data.get("id") or "").strip())
        if row_id is None:
            return OperationResult(False, (FlashMessage("Μη έγκυρη εγγραφή.", "danger"),))

        row = AleKae.query.get_or_404(row_id)
        before = serialize_model(row)

        db.session.delete(row)
        db.session.flush()
        log_action(entity=row, action="DELETE", before=before, after=None)
        db.session.commit()

        return OperationResult(True, (FlashMessage("Η εγγραφή διαγράφηκε.", "success"),), entity_id=row.id)

    return OperationResult(False, (FlashMessage("Μη έγκυρη ενέργεια.", "danger"),))


def execute_import_ale_kae(file_storage: Any) -> OperationResult:
    """
    Import ALE-KAE rows from an XLSX file.
    """
    file = file_storage
    if not file or not getattr(file, "filename", None):
        return OperationResult(False, (FlashMessage("Δεν επιλέχθηκε αρχείο.", "danger"),))

    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        return OperationResult(False, (FlashMessage("Επιτρέπεται μόνο αρχείο .xlsx", "danger"),))

    try:
        import openpyxl
        workbook = openpyxl.load_workbook(file, data_only=True)
        worksheet = workbook.active
    except Exception:
        return OperationResult(False, (FlashMessage("Αποτυχία ανάγνωσης Excel. Ελέγξτε το αρχείο.", "danger"),))

    try:
        header_cells = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    except StopIteration:
        return OperationResult(False, (FlashMessage("Το Excel είναι κενό.", "danger"),))

    idx_map = build_header_index(header_cells)

    ale_idx = idx_map.get("αλε", idx_map.get("ale"))
    old_kae_idx = idx_map.get("παλιος καε", idx_map.get("old kae", idx_map.get("old_kae")))
    desc_idx = idx_map.get("περιγραφη", idx_map.get("description"))
    resp_idx = idx_map.get("αρμοδιοτητας", idx_map.get("responsibility"))

    if ale_idx is None:
        return OperationResult(False, (FlashMessage("Το Excel πρέπει να έχει στήλη 'ΑΛΕ'.", "danger"),))

    inserted: list[AleKae] = []
    skipped_missing = 0
    skipped_duplicate = 0

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        ale = safe_cell_str(cell_at(row, ale_idx))
        if not ale:
            skipped_missing += 1
            continue

        if AleKae.query.filter_by(ale=ale).first():
            skipped_duplicate += 1
            continue

        obj = AleKae(
            ale=ale,
            old_kae=safe_cell_str(cell_at(row, old_kae_idx)) or None,
            description=safe_cell_str(cell_at(row, desc_idx)) or None,
            responsibility=safe_cell_str(cell_at(row, resp_idx)) or None,
        )
        db.session.add(obj)
        inserted.append(obj)

    if not inserted:
        return OperationResult(False, (FlashMessage("Δεν εισήχθησαν εγγραφές. Ελέγξτε required πεδία/διπλότυπα.", "warning"),))

    db.session.flush()
    for obj in inserted:
        log_action(entity=obj, action="CREATE", before=None, after=serialize_model(obj))
    db.session.commit()

    return OperationResult(
        True,
        (
            FlashMessage(
                f"Εισαγωγή ολοκληρώθηκε: {len(inserted)} νέες εγγραφές. "
                f"Παραλείφθηκαν: {skipped_missing} (ελλιπή), {skipped_duplicate} (διπλότυπα).",
                "success",
            ),
        ),
    )


# ----------------------------------------------------------------------
# CPV
# ----------------------------------------------------------------------
def build_cpv_page_context() -> dict[str, Any]:
    """
    Build template context for the CPV page.
    """
    rows = Cpv.query.order_by(Cpv.cpv.asc()).all()
    return {"rows": rows}


def execute_cpv_action(form_data: Mapping[str, object]) -> OperationResult:
    """
    Execute create/update/delete action for CPV rows.
    """
    action = (form_data.get("action") or "").strip()

    if action == "create":
        cpv_code = (form_data.get("cpv") or "").strip()
        description = (form_data.get("description") or "").strip() or None

        if not cpv_code:
            return OperationResult(False, (FlashMessage("Το CPV είναι υποχρεωτικό.", "danger"),))

        if Cpv.query.filter_by(cpv=cpv_code).first():
            return OperationResult(False, (FlashMessage("Υπάρχει ήδη εγγραφή με αυτό το CPV.", "danger"),))

        obj = Cpv(cpv=cpv_code, description=description)
        db.session.add(obj)
        db.session.flush()
        log_action(entity=obj, action="CREATE", before=None, after=serialize_model(obj))
        db.session.commit()

        return OperationResult(True, (FlashMessage("Η εγγραφή CPV προστέθηκε.", "success"),), entity_id=obj.id)

    if action == "update":
        obj_id = parse_optional_int((form_data.get("id") or "").strip())
        if obj_id is None:
            return OperationResult(False, (FlashMessage("Μη έγκυρη εγγραφή.", "danger"),))

        obj = Cpv.query.get_or_404(obj_id)
        before = serialize_model(obj)

        cpv_code = (form_data.get("cpv") or "").strip()
        description = (form_data.get("description") or "").strip() or None

        if not cpv_code:
            return OperationResult(False, (FlashMessage("Το CPV είναι υποχρεωτικό.", "danger"),))

        exists = Cpv.query.filter(Cpv.cpv == cpv_code, Cpv.id != obj.id).first()
        if exists:
            return OperationResult(False, (FlashMessage("Υπάρχει ήδη άλλη εγγραφή με αυτό το CPV.", "danger"),))

        obj.cpv = cpv_code
        obj.description = description

        db.session.flush()
        log_action(entity=obj, action="UPDATE", before=before, after=serialize_model(obj))
        db.session.commit()

        return OperationResult(True, (FlashMessage("Η εγγραφή ενημερώθηκε.", "success"),), entity_id=obj.id)

    if action == "delete":
        obj_id = parse_optional_int((form_data.get("id") or "").strip())
        if obj_id is None:
            return OperationResult(False, (FlashMessage("Μη έγκυρη εγγραφή.", "danger"),))

        obj = Cpv.query.get_or_404(obj_id)
        before = serialize_model(obj)

        db.session.delete(obj)
        db.session.flush()
        log_action(entity=obj, action="DELETE", before=before, after=None)
        db.session.commit()

        return OperationResult(True, (FlashMessage("Η εγγραφή διαγράφηκε.", "success"),), entity_id=obj.id)

    return OperationResult(False, (FlashMessage("Μη έγκυρη ενέργεια.", "danger"),))


def execute_import_cpv(file_storage: Any) -> OperationResult:
    """
    Import CPV rows from an XLSX file.
    """
    file = file_storage
    if not file or not getattr(file, "filename", None):
        return OperationResult(False, (FlashMessage("Δεν επιλέχθηκε αρχείο.", "danger"),))

    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        return OperationResult(False, (FlashMessage("Επιτρέπεται μόνο αρχείο .xlsx", "danger"),))

    try:
        import openpyxl
        workbook = openpyxl.load_workbook(file, data_only=True)
        worksheet = workbook.active
    except Exception:
        return OperationResult(False, (FlashMessage("Αποτυχία ανάγνωσης Excel. Ελέγξτε το αρχείο.", "danger"),))

    try:
        header_cells = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    except StopIteration:
        return OperationResult(False, (FlashMessage("Το Excel είναι κενό.", "danger"),))

    idx_map = build_header_index(header_cells)

    cpv_idx = idx_map.get("cpv")
    desc_idx = idx_map.get("περιγραφη", idx_map.get("description"))

    if cpv_idx is None:
        return OperationResult(False, (FlashMessage("Το Excel πρέπει να έχει στήλη 'CPV'.", "danger"),))

    inserted: list[Cpv] = []
    skipped_missing = 0
    skipped_duplicate = 0

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        cpv_code = safe_cell_str(cell_at(row, cpv_idx))
        if not cpv_code:
            skipped_missing += 1
            continue

        if Cpv.query.filter_by(cpv=cpv_code).first():
            skipped_duplicate += 1
            continue

        obj = Cpv(
            cpv=cpv_code,
            description=safe_cell_str(cell_at(row, desc_idx)) or None,
        )
        db.session.add(obj)
        inserted.append(obj)

    if not inserted:
        return OperationResult(False, (FlashMessage("Δεν εισήχθησαν εγγραφές. Ελέγξτε required πεδία/διπλότυπα.", "warning"),))

    db.session.flush()
    for obj in inserted:
        log_action(entity=obj, action="CREATE", before=None, after=serialize_model(obj))
    db.session.commit()

    return OperationResult(
        True,
        (
            FlashMessage(
                f"Εισαγωγή ολοκληρώθηκε: {len(inserted)} νέες εγγραφές. "
                f"Παραλείφθηκαν: {skipped_missing} (ελλιπή), {skipped_duplicate} (διπλότυπα).",
                "success",
            ),
        ),
    )


# ----------------------------------------------------------------------
# OPTION VALUES
# ----------------------------------------------------------------------
def build_option_values_page_context(*, key: str, label: str) -> dict[str, Any]:
    """
    Build template context for one generic OptionValue category page.
    """
    category = _get_or_create_category(key=key, label=label)
    values = get_all_option_rows(category.key)
    return {
        "category": category,
        "values": values,
        "page_label": label,
    }


def execute_option_value_action(*, key: str, label: str, form_data: Mapping[str, object]) -> OperationResult:
    """
    Execute create/update/delete action for one OptionValue category page.
    """
    category = _get_or_create_category(key=key, label=label)
    action = (form_data.get("action") or "").strip()

    if action == "create":
        value = (form_data.get("value") or "").strip()
        sort_order = parse_optional_int((form_data.get("sort_order") or "").strip()) or 0
        is_active = bool(form_data.get("is_active") == "on")

        if not value:
            return OperationResult(False, (FlashMessage("Η τιμή είναι υποχρεωτική.", "danger"),))

        existing = OptionValue.query.filter_by(
            category_id=category.id,
            value=value
        ).first()

        if existing:
            return OperationResult(
                False,
                (FlashMessage("Η τιμή υπάρχει ήδη σε αυτή την κατηγορία.", "warning"),),
             entity_id=existing.id,
        )

        row = OptionValue(
            category_id=category.id,
         value=value,
            is_active=is_active,
            sort_order=sort_order,
        )

        try:
            db.session.add(row)
            db.session.flush()
            log_action(entity=row, action="CREATE", after=serialize_model(row))
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            return OperationResult(
                False,
                (FlashMessage("Η τιμή υπάρχει ήδη σε αυτή την κατηγορία.", "warning"),),
            )

        return OperationResult(
            True,
            (FlashMessage("Η τιμή προστέθηκε.", "success"),),
            entity_id=row.id,
        )

    if action == "update":
        option_value_id = parse_optional_int((form_data.get("id") or "").strip())
        if option_value_id is None:
            return OperationResult(False, (FlashMessage("Μη έγκυρη εγγραφή.", "danger"),))

        row = OptionValue.query.filter_by(id=option_value_id, category_id=category.id).first()
        if not row:
            return OperationResult(False, (FlashMessage("Η εγγραφή δεν βρέθηκε.", "danger"),), not_found=True)

        before = serialize_model(row)
        value = (form_data.get("value") or "").strip()
        sort_order = parse_optional_int((form_data.get("sort_order") or "").strip()) or 0
        is_active = bool(form_data.get("is_active") == "on")

        if not value:
            return OperationResult(False, (FlashMessage("Η τιμή είναι υποχρεωτική.", "danger"),))

        duplicate = (
            OptionValue.query.filter(
                OptionValue.category_id == category.id,
                OptionValue.value == value,
                OptionValue.id != row.id,
            ).first()
        )
        if duplicate:
            return OperationResult(
                False,
                (FlashMessage("Υπάρχει ήδη άλλη εγγραφή με αυτή την τιμή.", "warning"),),
                entity_id=row.id,
            )

        row.value = value
        row.sort_order = sort_order
        row.is_active = is_active

        try:
            db.session.flush()
            log_action(entity=row, action="UPDATE", before=before, after=serialize_model(row))
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            return OperationResult(
                False,
                (FlashMessage("Υπάρχει ήδη άλλη εγγραφή με αυτή την τιμή.", "warning"),),
                entity_id=row.id,
            )

        return OperationResult(
            True,
            (FlashMessage("Η τιμή ενημερώθηκε.", "success"),),
            entity_id=row.id,
        )

    if action == "delete":
        option_value_id = parse_optional_int((form_data.get("id") or "").strip())
        if option_value_id is None:
            return OperationResult(False, (FlashMessage("Μη έγκυρη εγγραφή.", "danger"),))

        row = OptionValue.query.filter_by(id=option_value_id, category_id=category.id).first()
        if not row:
            return OperationResult(False, (FlashMessage("Η εγγραφή δεν βρέθηκε.", "danger"),), not_found=True)

        before = serialize_model(row)

        try:
            db.session.delete(row)
            db.session.flush()
            log_action(entity=row, action="DELETE", before=before)
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            return OperationResult(
                False,
                (FlashMessage("Η τιμή δεν μπορεί να διαγραφεί γιατί χρησιμοποιείται ήδη.", "danger"),),
                entity_id=row.id,
            )

        return OperationResult(
            True,
            (FlashMessage("Η τιμή διαγράφηκε.", "success"),),
            entity_id=row.id,
        )

    return OperationResult(False, (FlashMessage("Μη έγκυρη ενέργεια.", "danger"),))


# ----------------------------------------------------------------------
# INCOME TAX RULES
# ----------------------------------------------------------------------
def build_income_tax_rules_page_context() -> dict[str, Any]:
    """
    Build template context for the IncomeTaxRule page.
    """
    rules = IncomeTaxRule.query.order_by(IncomeTaxRule.description.asc()).all()
    return {"rules": rules}


def execute_income_tax_rule_action(form_data: Mapping[str, object]) -> OperationResult:
    action = (form_data.get("action") or "").strip()

    if action == "create":
        description = (form_data.get("description") or "").strip()
        rate_percent = parse_decimal((form_data.get("rate_percent") or "").strip())
        threshold_amount = parse_decimal((form_data.get("threshold_amount") or "").strip())
        is_active = bool(form_data.get("is_active") == "on")

        if not description:
            return OperationResult(False, (FlashMessage("Η περιγραφή είναι υποχρεωτική.", "danger"),))

        existing = IncomeTaxRule.query.filter_by(description=description).first()
        if existing:
            return OperationResult(
                False,
                (FlashMessage("Υπάρχει ήδη κανόνας με αυτή την περιγραφή.", "warning"),),
                entity_id=existing.id,
            )

        row = IncomeTaxRule(
            description=description,
            rate_percent=rate_percent,
            threshold_amount=threshold_amount,
            is_active=is_active,
        )

        try:
            db.session.add(row)
            db.session.flush()
            log_action(entity=row, action="CREATE", after=serialize_model(row))
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            return OperationResult(
                False,
                (FlashMessage("Υπάρχει ήδη κανόνας με αυτή την περιγραφή.", "warning"),),
            )

        return OperationResult(
            True,
            (FlashMessage("Ο κανόνας φόρου εισοδήματος προστέθηκε.", "success"),),
            entity_id=row.id,
        )

    if action == "update":
        rule_id = parse_optional_int((form_data.get("id") or "").strip())
        if rule_id is None:
            return OperationResult(False, (FlashMessage("Μη έγκυρη εγγραφή.", "danger"),))

        row = IncomeTaxRule.query.get(rule_id)
        if not row:
            return OperationResult(False, (FlashMessage("Η εγγραφή δεν βρέθηκε.", "danger"),), not_found=True)

        description = (form_data.get("description") or "").strip()
        rate_percent = parse_decimal((form_data.get("rate_percent") or "").strip())
        threshold_amount = parse_decimal((form_data.get("threshold_amount") or "").strip())
        is_active = bool(form_data.get("is_active") == "on")

        if not description:
            return OperationResult(False, (FlashMessage("Η περιγραφή είναι υποχρεωτική.", "danger"),))

        duplicate = IncomeTaxRule.query.filter(
            IncomeTaxRule.description == description,
            IncomeTaxRule.id != row.id,
        ).first()
        if duplicate:
            return OperationResult(
                False,
                (FlashMessage("Υπάρχει ήδη άλλος κανόνας με αυτή την περιγραφή.", "warning"),),
                entity_id=row.id,
            )

        before = serialize_model(row)
        row.description = description
        row.rate_percent = rate_percent
        row.threshold_amount = threshold_amount
        row.is_active = is_active

        try:
            db.session.flush()
            log_action(entity=row, action="UPDATE", before=before, after=serialize_model(row))
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            return OperationResult(
                False,
                (FlashMessage("Υπάρχει ήδη άλλος κανόνας με αυτή την περιγραφή.", "warning"),),
                entity_id=row.id,
            )

        return OperationResult(
            True,
            (FlashMessage("Ο κανόνας φόρου εισοδήματος ενημερώθηκε.", "success"),),
            entity_id=row.id,
        )

    if action == "delete":
        rule_id = parse_optional_int((form_data.get("id") or "").strip())
        if rule_id is None:
            return OperationResult(False, (FlashMessage("Μη έγκυρη εγγραφή.", "danger"),))

        row = IncomeTaxRule.query.get(rule_id)
        if not row:
            return OperationResult(False, (FlashMessage("Η εγγραφή δεν βρέθηκε.", "danger"),), not_found=True)

        before = serialize_model(row)

        try:
            db.session.delete(row)
            db.session.flush()
            log_action(entity=row, action="DELETE", before=before)
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            return OperationResult(
                False,
                (FlashMessage("Ο κανόνας δεν μπορεί να διαγραφεί γιατί χρησιμοποιείται ήδη.", "danger"),),
                entity_id=row.id,
            )

        return OperationResult(
            True,
            (FlashMessage("Ο κανόνας φόρου εισοδήματος διαγράφηκε.", "success"),),
            entity_id=row.id,
        )

    return OperationResult(False, (FlashMessage("Μη έγκυρη ενέργεια.", "danger"),))

# ----------------------------------------------------------------------
# WITHHOLDING PROFILES
# ----------------------------------------------------------------------
def build_withholding_profiles_page_context() -> dict[str, Any]:
    """
    Build template context for the WithholdingProfile page.
    """
    profiles = WithholdingProfile.query.order_by(WithholdingProfile.description.asc()).all()
    return {"profiles": profiles}


def execute_withholding_profile_action(form_data: Mapping[str, object]) -> OperationResult:
    action = (form_data.get("action") or "").strip()

    if action == "create":
        description = (form_data.get("description") or "").strip()
        mt_eloa_percent = parse_decimal((form_data.get("mt_eloa_percent") or "").strip())
        eadhsy_percent = parse_decimal((form_data.get("eadhsy_percent") or "").strip())
        withholding1_percent = parse_decimal((form_data.get("withholding1_percent") or "").strip())
        withholding2_percent = parse_decimal((form_data.get("withholding2_percent") or "").strip())
        is_active = bool(form_data.get("is_active") == "on")

        if not description:
            return OperationResult(False, (FlashMessage("Η περιγραφή είναι υποχρεωτική.", "danger"),))

        existing = WithholdingProfile.query.filter_by(description=description).first()
        if existing:
            return OperationResult(
                False,
                (FlashMessage("Υπάρχει ήδη προφίλ με αυτή την περιγραφή.", "warning"),),
                entity_id=existing.id,
            )

        row = WithholdingProfile(
            description=description,
            mt_eloa_percent=mt_eloa_percent,
            eadhsy_percent=eadhsy_percent,
            withholding1_percent=withholding1_percent,
            withholding2_percent=withholding2_percent,
            is_active=is_active,
        )

        try:
            db.session.add(row)
            db.session.flush()
            log_action(entity=row, action="CREATE", after=serialize_model(row))
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            return OperationResult(
                False,
                (FlashMessage("Υπάρχει ήδη προφίλ με αυτή την περιγραφή.", "warning"),),
            )

        return OperationResult(
            True,
            (FlashMessage("Το προφίλ κρατήσεων προστέθηκε.", "success"),),
            entity_id=row.id,
        )

    if action == "update":
        profile_id = parse_optional_int((form_data.get("id") or "").strip())
        if profile_id is None:
            return OperationResult(False, (FlashMessage("Μη έγκυρη εγγραφή.", "danger"),))

        row = WithholdingProfile.query.get(profile_id)
        if not row:
            return OperationResult(False, (FlashMessage("Η εγγραφή δεν βρέθηκε.", "danger"),), not_found=True)

        description = (form_data.get("description") or "").strip()
        mt_eloa_percent = parse_decimal((form_data.get("mt_eloa_percent") or "").strip())
        eadhsy_percent = parse_decimal((form_data.get("eadhsy_percent") or "").strip())
        withholding1_percent = parse_decimal((form_data.get("withholding1_percent") or "").strip())
        withholding2_percent = parse_decimal((form_data.get("withholding2_percent") or "").strip())
        is_active = bool(form_data.get("is_active") == "on")

        if not description:
            return OperationResult(False, (FlashMessage("Η περιγραφή είναι υποχρεωτική.", "danger"),))

        duplicate = WithholdingProfile.query.filter(
            WithholdingProfile.description == description,
            WithholdingProfile.id != row.id,
        ).first()
        if duplicate:
            return OperationResult(
                False,
                (FlashMessage("Υπάρχει ήδη άλλο προφίλ με αυτή την περιγραφή.", "warning"),),
                entity_id=row.id,
            )

        before = serialize_model(row)
        row.description = description
        row.mt_eloa_percent = mt_eloa_percent
        row.eadhsy_percent = eadhsy_percent
        row.withholding1_percent = withholding1_percent
        row.withholding2_percent = withholding2_percent
        row.is_active = is_active

        try:
            db.session.flush()
            log_action(entity=row, action="UPDATE", before=before, after=serialize_model(row))
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            return OperationResult(
                False,
                (FlashMessage("Υπάρχει ήδη άλλο προφίλ με αυτή την περιγραφή.", "warning"),),
                entity_id=row.id,
            )

        return OperationResult(
            True,
            (FlashMessage("Το προφίλ κρατήσεων ενημερώθηκε.", "success"),),
            entity_id=row.id,
        )

    if action == "delete":
        profile_id = parse_optional_int((form_data.get("id") or "").strip())
        if profile_id is None:
            return OperationResult(False, (FlashMessage("Μη έγκυρη εγγραφή.", "danger"),))

        row = WithholdingProfile.query.get(profile_id)
        if not row:
            return OperationResult(False, (FlashMessage("Η εγγραφή δεν βρέθηκε.", "danger"),), not_found=True)

        before = serialize_model(row)

        try:
            db.session.delete(row)
            db.session.flush()
            log_action(entity=row, action="DELETE", before=before)
            db.session.commit()
        except IntegrityError:
            db.session.rollback()
            return OperationResult(
                False,
                (FlashMessage("Το προφίλ δεν μπορεί να διαγραφεί γιατί χρησιμοποιείται ήδη.", "danger"),),
                entity_id=row.id,
            )

        return OperationResult(
            True,
            (FlashMessage("Το προφίλ κρατήσεων διαγράφηκε.", "success"),),
            entity_id=row.id,
        )

    return OperationResult(False, (FlashMessage("Μη έγκυρη ενέργεια.", "danger"),))


__all__ = [
    "build_ale_kae_page_context",
    "execute_ale_kae_action",
    "execute_import_ale_kae",
    "build_cpv_page_context",
    "execute_cpv_action",
    "execute_import_cpv",
    "build_option_values_page_context",
    "execute_option_value_action",
    "build_income_tax_rules_page_context",
    "execute_income_tax_rule_action",
    "build_withholding_profiles_page_context",
    "execute_withholding_profile_action",
]


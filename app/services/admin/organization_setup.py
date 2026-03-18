"""
app/services/admin/organization_setup.py

Focused page/use-case services for the consolidated organization setup flow.

PURPOSE
-------
This module extracts non-HTTP orchestration from:

    /admin/organization-setup

It moves out:
- page-context assembly
- action dispatch
- structural validation
- ORM mutation orchestration
- audit logging / commit

IMPORTANT CHANGE
----------------
Organization membership is assignment-based.

A person may belong to multiple departments/directories inside the same
ServiceUnit through `PersonnelDepartmentAssignment`.

That means:
- Personnel edit page no longer owns department/directory assignment.
- All organizational placement is centrally managed from Organization Setup.
- Procurement handler dropdown uses these assignment rows directly.

BUSINESS RULE FOR PROCUREMENT HANDLERS
--------------------------------------
Procurement handler selection is assignment-based and must preserve the exact:
- person
- directory
- department

used for a specific procurement.

Therefore, when a department role holder is assigned manually from the
organization setup screen:
- head_personnel_id
- assistant_personnel_id

the corresponding PersonnelDepartmentAssignment should also exist.

Otherwise the person is visible as a department role holder in organization
setup, but does not appear in the procurement handler dropdown, because
procurement handlers are loaded from assignment rows.

This module therefore auto-creates missing assignment rows for:
- department head
- department assistant

when department roles are updated manually.

ARCHITECTURAL INTENT
--------------------
This module is intentionally explicit and function-first.

BOUNDARY
--------
This module MAY:
- query organization entities
- validate submitted values
- mutate organization entities
- audit and commit
- return structured results for routes

This module MUST NOT:
- define routes
- call render_template(...)
- call redirect(...)
- call flash(...)
"""

from __future__ import annotations

from collections.abc import Mapping
from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from ...audit import log_action, serialize_model
from ...extensions import db
from ...models import (
    Department,
    Directory,
    Personnel,
    PersonnelDepartmentAssignment,
    ServiceUnit,
)
from ..organization import (
    active_personnel_for_service_unit,
    active_personnel_ids_for_service_unit,
    service_units_for_dropdown,
)
from ..shared.operation_results import FlashMessage
from ..shared.parsing import parse_optional_int


@dataclass(frozen=True)
class OrganizationSetupOperationResult:
    """
    Result object for organization-setup POST actions.
    """

    ok: bool
    flashes: tuple[FlashMessage, ...]
    redirect_service_unit_id: int | None = None


def build_organization_setup_page_context(
    request_args: Mapping[str, Any],
    *,
    is_admin: bool,
    current_service_unit_id: int | None,
) -> dict[str, Any]:
    """
    Build template context for the consolidated organization setup page.

    RETURNS
    -------
    dict[str, Any]
        Includes:
        - service unit scope
        - directories
        - departments
        - active personnel list
        - department membership rows grouped per department
    """
    if is_admin:
        service_unit_id = parse_optional_int(request_args.get("service_unit_id"))
    else:
        service_unit_id = current_service_unit_id

    service_units = service_units_for_dropdown()

    unit = ServiceUnit.query.get(service_unit_id) if service_unit_id else None
    directories: list[Directory] = []
    departments: list[Department] = []
    personnel_list: list[Personnel] = []
    department_memberships: dict[int, list[PersonnelDepartmentAssignment]] = {}

    if unit:
        _ensure_target_service_unit_scope(
            unit.id,
            is_admin=is_admin,
            current_service_unit_id=current_service_unit_id,
        )

        directories = (
            Directory.query.filter_by(service_unit_id=unit.id)
            .order_by(Directory.name.asc())
            .all()
        )
        departments = (
            Department.query.filter_by(service_unit_id=unit.id)
            .order_by(Department.directory_id.asc(), Department.name.asc())
            .all()
        )
        personnel_list = active_personnel_for_service_unit(unit.id)

        assignments = (
            PersonnelDepartmentAssignment.query.filter_by(service_unit_id=unit.id)
            .order_by(
                PersonnelDepartmentAssignment.department_id.asc(),
                PersonnelDepartmentAssignment.is_primary.desc(),
                PersonnelDepartmentAssignment.id.asc(),
            )
            .all()
        )

        for assignment in assignments:
            department_memberships.setdefault(assignment.department_id, []).append(assignment)

    return {
        "service_units": service_units,
        "scope_service_unit_id": (unit.id if unit else None),
        "unit": unit,
        "directories": directories,
        "departments": departments,
        "personnel_list": personnel_list,
        "department_memberships": department_memberships,
        "is_admin": is_admin,
    }


def execute_organization_setup_action(
    form_data: Mapping[str, Any],
    *,
    files: Mapping[str, Any] | None = None,
    is_admin: bool,
    current_service_unit_id: int | None,
) -> OrganizationSetupOperationResult:
    """
    Dispatch and execute one organization-setup action.

    SUPPORTED ACTIONS
    -----------------
    - import
    - create/update/delete directory
    - create/update/delete department
    - update directory director
    - update department roles
    - add_department_member
    - remove_department_member
    """
    action = (form_data.get("action") or "").strip()

    target_service_unit_id = _resolve_target_service_unit_id(
        form_data,
        is_admin=is_admin,
        current_service_unit_id=current_service_unit_id,
    )
    if target_service_unit_id is None:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Η υπηρεσία είναι υποχρεωτική.", "danger"),),
            redirect_service_unit_id=None,
        )

    _ensure_target_service_unit_scope(
        target_service_unit_id,
        is_admin=is_admin,
        current_service_unit_id=current_service_unit_id,
    )

    unit = ServiceUnit.query.get_or_404(target_service_unit_id)
    allowed_personnel_ids = active_personnel_ids_for_service_unit(unit.id)

    if action == "import":
        return _execute_import_organization_structure(
            unit,
            (files or {}).get("file"),
            allowed_personnel_ids=allowed_personnel_ids,
        )

    if action == "create_directory":
        return _execute_create_directory(unit, form_data)

    if action == "update_directory":
        return _execute_update_directory(unit, form_data)

    if action == "delete_directory":
        return _execute_delete_directory(unit, form_data)

    if action == "create_department":
        return _execute_create_department(unit, form_data)

    if action == "update_department":
        return _execute_update_department(unit, form_data)

    if action == "delete_department":
        return _execute_delete_department(unit, form_data)

    if action == "update_directory_director":
        return _execute_update_directory_director(
            unit,
            form_data,
            allowed_personnel_ids=allowed_personnel_ids,
        )

    if action == "update_department_roles":
        return _execute_update_department_roles(
            unit,
            form_data,
            allowed_personnel_ids=allowed_personnel_ids,
        )

    if action == "add_department_member":
        return _execute_add_department_member(
            unit,
            form_data,
            allowed_personnel_ids=allowed_personnel_ids,
        )

    if action == "remove_department_member":
        return _execute_remove_department_member(unit, form_data)

    return OrganizationSetupOperationResult(
        ok=False,
        flashes=(FlashMessage("Μη έγκυρη ενέργεια.", "danger"),),
        redirect_service_unit_id=unit.id,
    )


def _resolve_target_service_unit_id(
    form_data: Mapping[str, Any],
    *,
    is_admin: bool,
    current_service_unit_id: int | None,
) -> int | None:
    if is_admin:
        return parse_optional_int(form_data.get("service_unit_id"))
    return current_service_unit_id


def _ensure_target_service_unit_scope(
    service_unit_id: int,
    *,
    is_admin: bool,
    current_service_unit_id: int | None,
) -> None:
    if is_admin:
        return

    if not current_service_unit_id or service_unit_id != current_service_unit_id:
        from flask import abort
        abort(403)


def _validate_service_unit_personnel_or_none(
    raw_personnel_id: Any,
    *,
    allowed_personnel_ids: set[int],
) -> tuple[int | None, FlashMessage | None]:
    personnel_id = parse_optional_int(raw_personnel_id)
    if personnel_id is None:
        return None, None

    if personnel_id not in allowed_personnel_ids:
        return None, FlashMessage("Μη έγκυρη επιλογή προσωπικού για την υπηρεσία.", "danger")

    return personnel_id, None


def _ensure_department_assignment(
    *,
    unit: ServiceUnit,
    department: Department,
    personnel_id: int | None,
    is_primary: bool,
) -> bool:
    """
    Ensure that a PersonnelDepartmentAssignment exists for the given person and
    department.

    PARAMETERS
    ----------
    unit:
        The scoped ServiceUnit.
    department:
        The target Department.
    personnel_id:
        Personnel id to ensure membership for.
    is_primary:
        Whether the created/updated membership should be primary.

    RETURNS
    -------
    bool
        True when a new assignment row was created.
        False when no new row was needed.

    WHY THIS HELPER EXISTS
    ----------------------
    Procurement handler selection is assignment-based. If a user is assigned as
    Department Head or Assistant but no membership row exists, they will not
    appear as a selectable handler for procurements.

    This helper keeps manual role assignment consistent with the assignment-
    based organizational model.
    """
    if personnel_id is None:
        return False

    existing_assignment = PersonnelDepartmentAssignment.query.filter_by(
        personnel_id=personnel_id,
        department_id=department.id,
    ).first()

    if existing_assignment is not None:
        existing_assignment.service_unit_id = unit.id
        existing_assignment.directory_id = department.directory_id

        if is_primary and not existing_assignment.is_primary:
            existing_assignment.is_primary = True

        return False

    assignment = PersonnelDepartmentAssignment(
        personnel_id=personnel_id,
        service_unit_id=unit.id,
        directory_id=department.directory_id,
        department_id=department.id,
        is_primary=is_primary,
    )
    db.session.add(assignment)
    db.session.flush()
    log_action(entity=assignment, action="CREATE", before=None, after=serialize_model(assignment))
    return True


def _execute_create_directory(
    unit: ServiceUnit,
    form_data: Mapping[str, Any],
) -> OrganizationSetupOperationResult:
    name = (form_data.get("directory_name") or "").strip()

    if not name:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Η ονομασία Διεύθυνσης είναι υποχρεωτική.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    exists = Directory.query.filter_by(service_unit_id=unit.id, name=name).first()
    if exists:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Υπάρχει ήδη Διεύθυνση με αυτή την ονομασία στην Υπηρεσία.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    directory = Directory(
        service_unit_id=unit.id,
        name=name,
        is_active=True,
        director_personnel_id=None,
    )
    db.session.add(directory)
    db.session.flush()
    log_action(entity=directory, action="CREATE", before=None, after=serialize_model(directory))
    db.session.commit()

    return OrganizationSetupOperationResult(
        ok=True,
        flashes=(FlashMessage("Η Διεύθυνση δημιουργήθηκε.", "success"),),
        redirect_service_unit_id=unit.id,
    )


def _execute_update_directory(
    unit: ServiceUnit,
    form_data: Mapping[str, Any],
) -> OrganizationSetupOperationResult:
    directory_id = parse_optional_int(form_data.get("directory_id"))
    if directory_id is None:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Μη έγκυρη Διεύθυνση.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    directory = Directory.query.get_or_404(directory_id)
    if directory.service_unit_id != unit.id:
        from flask import abort
        abort(403)

    name = (form_data.get("directory_name") or "").strip()
    is_active = bool(form_data.get("is_active"))

    if not name:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Η ονομασία Διεύθυνσης είναι υποχρεωτική.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    exists = Directory.query.filter(
        Directory.service_unit_id == unit.id,
        Directory.name == name,
        Directory.id != directory.id,
    ).first()
    if exists:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Υπάρχει ήδη άλλη Διεύθυνση με αυτή την ονομασία.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    before = serialize_model(directory)
    directory.name = name
    directory.is_active = is_active

    db.session.flush()
    log_action(entity=directory, action="UPDATE", before=before, after=serialize_model(directory))
    db.session.commit()

    return OrganizationSetupOperationResult(
        ok=True,
        flashes=(FlashMessage("Η Διεύθυνση ενημερώθηκε.", "success"),),
        redirect_service_unit_id=unit.id,
    )


def _execute_delete_directory(
    unit: ServiceUnit,
    form_data: Mapping[str, Any],
) -> OrganizationSetupOperationResult:
    directory_id = parse_optional_int(form_data.get("directory_id"))
    if directory_id is None:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Μη έγκυρη Διεύθυνση.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    directory = Directory.query.get_or_404(directory_id)
    if directory.service_unit_id != unit.id:
        from flask import abort
        abort(403)

    before = serialize_model(directory)

    Department.query.filter_by(directory_id=directory.id).update(
        {"head_personnel_id": None, "assistant_personnel_id": None},
        synchronize_session=False,
    )

    PersonnelDepartmentAssignment.query.filter_by(directory_id=directory.id).delete(
        synchronize_session=False
    )

    departments_to_delete = Department.query.filter_by(directory_id=directory.id).all()
    for department in departments_to_delete:
        department_before = serialize_model(department)
        db.session.delete(department)
        db.session.flush()
        log_action(entity=department, action="DELETE", before=department_before, after=None)

    db.session.delete(directory)
    db.session.flush()
    log_action(entity=directory, action="DELETE", before=before, after=None)
    db.session.commit()

    return OrganizationSetupOperationResult(
        ok=True,
        flashes=(FlashMessage("Η Διεύθυνση διαγράφηκε.", "success"),),
        redirect_service_unit_id=unit.id,
    )


def _execute_create_department(
    unit: ServiceUnit,
    form_data: Mapping[str, Any],
) -> OrganizationSetupOperationResult:
    directory_id = parse_optional_int(form_data.get("directory_id"))
    name = (form_data.get("department_name") or "").strip()

    if not directory_id:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Η Διεύθυνση είναι υποχρεωτική για δημιουργία Τμήματος.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    if not name:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Η ονομασία Τμήματος είναι υποχρεωτική.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    directory = Directory.query.get(directory_id)
    if not directory or directory.service_unit_id != unit.id:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Μη έγκυρη Διεύθυνση για την Υπηρεσία.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    exists = Department.query.filter_by(directory_id=directory.id, name=name).first()
    if exists:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Υπάρχει ήδη Τμήμα με αυτή την ονομασία στη συγκεκριμένη Διεύθυνση.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    department = Department(
        service_unit_id=unit.id,
        directory_id=directory.id,
        name=name,
        is_active=True,
        head_personnel_id=None,
        assistant_personnel_id=None,
    )
    db.session.add(department)
    db.session.flush()
    log_action(entity=department, action="CREATE", before=None, after=serialize_model(department))
    db.session.commit()

    return OrganizationSetupOperationResult(
        ok=True,
        flashes=(FlashMessage("Το Τμήμα δημιουργήθηκε.", "success"),),
        redirect_service_unit_id=unit.id,
    )


def _execute_update_department(
    unit: ServiceUnit,
    form_data: Mapping[str, Any],
) -> OrganizationSetupOperationResult:
    department_id = parse_optional_int(form_data.get("department_id"))
    new_directory_id = parse_optional_int(form_data.get("directory_id"))
    name = (form_data.get("department_name") or "").strip()

    if department_id is None:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Μη έγκυρο Τμήμα.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    if new_directory_id is None:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Η Διεύθυνση είναι υποχρεωτική.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    if not name:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Η ονομασία Τμήματος είναι υποχρεωτική.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    department = Department.query.get_or_404(department_id)
    if department.service_unit_id != unit.id:
        from flask import abort
        abort(403)

    new_directory = Directory.query.get_or_404(new_directory_id)
    if new_directory.service_unit_id != unit.id:
        from flask import abort
        abort(403)

    exists = Department.query.filter(
        Department.directory_id == new_directory.id,
        Department.name == name,
        Department.id != department.id,
    ).first()
    if exists:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Υπάρχει ήδη άλλο Τμήμα με αυτή την ονομασία στη συγκεκριμένη Διεύθυνση.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    before = serialize_model(department)
    department.directory_id = new_directory.id
    department.name = name
    department.is_active = bool(form_data.get("is_active"))

    PersonnelDepartmentAssignment.query.filter_by(department_id=department.id).update(
        {"directory_id": new_directory.id},
        synchronize_session=False,
    )

    db.session.flush()
    log_action(entity=department, action="UPDATE", before=before, after=serialize_model(department))
    db.session.commit()

    return OrganizationSetupOperationResult(
        ok=True,
        flashes=(FlashMessage("Το Τμήμα ενημερώθηκε.", "success"),),
        redirect_service_unit_id=unit.id,
    )


def _execute_delete_department(
    unit: ServiceUnit,
    form_data: Mapping[str, Any],
) -> OrganizationSetupOperationResult:
    department_id = parse_optional_int(form_data.get("department_id"))
    if department_id is None:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Μη έγκυρο Τμήμα.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    department = Department.query.get_or_404(department_id)
    if department.service_unit_id != unit.id:
        from flask import abort
        abort(403)

    before = serialize_model(department)

    PersonnelDepartmentAssignment.query.filter_by(department_id=department.id).delete(
        synchronize_session=False
    )

    department.head_personnel_id = None
    department.assistant_personnel_id = None

    db.session.flush()
    db.session.delete(department)
    db.session.flush()
    log_action(entity=department, action="DELETE", before=before, after=None)
    db.session.commit()

    return OrganizationSetupOperationResult(
        ok=True,
        flashes=(FlashMessage("Το Τμήμα διαγράφηκε.", "success"),),
        redirect_service_unit_id=unit.id,
    )


def _execute_update_directory_director(
    unit: ServiceUnit,
    form_data: Mapping[str, Any],
    *,
    allowed_personnel_ids: set[int],
) -> OrganizationSetupOperationResult:
    directory_id = parse_optional_int(form_data.get("directory_id"))
    if directory_id is None:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Μη έγκυρη Διεύθυνση.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    directory = Directory.query.get_or_404(directory_id)
    if directory.service_unit_id != unit.id:
        from flask import abort
        abort(403)

    director_personnel_id, validation_flash = _validate_service_unit_personnel_or_none(
        form_data.get("director_personnel_id"),
        allowed_personnel_ids=allowed_personnel_ids,
    )

    before = serialize_model(directory)
    directory.director_personnel_id = director_personnel_id

    db.session.flush()
    log_action(entity=directory, action="UPDATE", before=before, after=serialize_model(directory))
    db.session.commit()

    flashes: list[FlashMessage] = []
    if validation_flash is not None:
        flashes.append(validation_flash)
    flashes.append(FlashMessage("Ο Διευθυντής Διεύθυνσης ενημερώθηκε.", "success"))

    return OrganizationSetupOperationResult(
        ok=True,
        flashes=tuple(flashes),
        redirect_service_unit_id=unit.id,
    )


def _execute_update_department_roles(
    unit: ServiceUnit,
    form_data: Mapping[str, Any],
    *,
    allowed_personnel_ids: set[int],
) -> OrganizationSetupOperationResult:
    """
    Update department role holders and ensure assignment-based membership exists.

    IMPORTANT
    ---------
    Procurement handler selection is built from PersonnelDepartmentAssignment.
    Therefore, assigning a Department Head / Assistant manually must also ensure
    the corresponding membership row exists, otherwise the role holder will not
    appear as a selectable procurement handler.
    """
    department_id = parse_optional_int(form_data.get("department_id"))
    if department_id is None:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Μη έγκυρο Τμήμα.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    head_personnel_id, head_flash = _validate_service_unit_personnel_or_none(
        form_data.get("head_personnel_id"),
        allowed_personnel_ids=allowed_personnel_ids,
    )
    assistant_personnel_id, assistant_flash = _validate_service_unit_personnel_or_none(
        form_data.get("assistant_personnel_id"),
        allowed_personnel_ids=allowed_personnel_ids,
    )

    if (
        head_personnel_id is not None
        and assistant_personnel_id is not None
        and head_personnel_id == assistant_personnel_id
    ):
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Ο ίδιος/η ίδια δεν μπορεί να είναι και Προϊστάμενος και Βοηθός.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    department = Department.query.get_or_404(department_id)
    if department.service_unit_id != unit.id:
        from flask import abort
        abort(403)

    before = serialize_model(department)
    department.head_personnel_id = head_personnel_id
    department.assistant_personnel_id = assistant_personnel_id

    created_memberships = 0

    if _ensure_department_assignment(
        unit=unit,
        department=department,
        personnel_id=head_personnel_id,
        is_primary=True,
    ):
        created_memberships += 1

    if _ensure_department_assignment(
        unit=unit,
        department=department,
        personnel_id=assistant_personnel_id,
        is_primary=False,
    ):
        created_memberships += 1

    db.session.flush()
    log_action(entity=department, action="UPDATE", before=before, after=serialize_model(department))
    db.session.commit()

    flashes: list[FlashMessage] = []
    if head_flash is not None:
        flashes.append(head_flash)
    if assistant_flash is not None:
        flashes.append(assistant_flash)

    if created_memberships:
        flashes.append(
            FlashMessage(
                f"Δημιουργήθηκαν αυτόματα {created_memberships} αναθέσεις μέλους για συμβατότητα με τον Χειριστή Προμήθειας.",
                "info",
            )
        )

    flashes.append(FlashMessage("Οι ρόλοι Τμήματος ενημερώθηκαν.", "success"))

    return OrganizationSetupOperationResult(
        ok=True,
        flashes=tuple(flashes),
        redirect_service_unit_id=unit.id,
    )


def _execute_add_department_member(
    unit: ServiceUnit,
    form_data: Mapping[str, Any],
    *,
    allowed_personnel_ids: set[int],
) -> OrganizationSetupOperationResult:
    """
    Add one personnel membership assignment to a department.

    VALIDATION
    ----------
    - Person must belong to same service unit scope.
    - Department must belong to same service unit.
    - Duplicate membership is prevented by validation before insert.
    """
    department_id = parse_optional_int(form_data.get("department_id"))
    personnel_id = parse_optional_int(form_data.get("personnel_id"))

    if department_id is None:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Μη έγκυρο Τμήμα.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    if personnel_id is None:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Το προσωπικό είναι υποχρεωτικό.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    if personnel_id not in allowed_personnel_ids:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Μη έγκυρη επιλογή προσωπικού για την υπηρεσία.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    department = Department.query.get_or_404(department_id)
    if department.service_unit_id != unit.id:
        from flask import abort
        abort(403)

    exists = PersonnelDepartmentAssignment.query.filter_by(
        personnel_id=personnel_id,
        department_id=department.id,
    ).first()
    if exists:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Το προσωπικό είναι ήδη καταχωρημένο στο συγκεκριμένο Τμήμα.", "warning"),),
            redirect_service_unit_id=unit.id,
        )

    assignment = PersonnelDepartmentAssignment(
        personnel_id=personnel_id,
        service_unit_id=unit.id,
        directory_id=department.directory_id,
        department_id=department.id,
        is_primary=False,
    )

    db.session.add(assignment)
    db.session.flush()
    log_action(entity=assignment, action="CREATE", before=None, after=serialize_model(assignment))
    db.session.commit()

    return OrganizationSetupOperationResult(
        ok=True,
        flashes=(FlashMessage("Το μέλος προστέθηκε στο Τμήμα.", "success"),),
        redirect_service_unit_id=unit.id,
    )


def _execute_remove_department_member(
    unit: ServiceUnit,
    form_data: Mapping[str, Any],
) -> OrganizationSetupOperationResult:
    """
    Remove one personnel membership assignment from a department.
    """
    assignment_id = parse_optional_int(form_data.get("assignment_id"))
    if assignment_id is None:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Μη έγκυρη ανάθεση μέλους.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    assignment = PersonnelDepartmentAssignment.query.get_or_404(assignment_id)
    if assignment.service_unit_id != unit.id:
        from flask import abort
        abort(403)

    before = serialize_model(assignment)
    db.session.delete(assignment)
    db.session.flush()
    log_action(entity=assignment, action="DELETE", before=before, after=None)
    db.session.commit()

    return OrganizationSetupOperationResult(
        ok=True,
        flashes=(FlashMessage("Το μέλος αφαιρέθηκε από το Τμήμα.", "success"),),
        redirect_service_unit_id=unit.id,
    )


def _clean_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_agm(value: Any) -> str:
    raw = _clean_cell(value)
    if not raw:
        return ""
    if raw.endswith(".0"):
        raw = raw[:-2]
    return raw.strip()


def _personnel_id_by_agm_for_service_unit(service_unit_id: int, agm: str) -> int | None:
    agm_value = _normalize_agm(agm)
    if not agm_value:
        return None

    person = (
        Personnel.query.filter(
            Personnel.service_unit_id == service_unit_id,
            Personnel.agm == agm_value,
            Personnel.is_active.is_(True),
        )
        .first()
    )
    return person.id if person else None


def _execute_import_organization_structure(
    unit,
    file_storage,
    *,
    allowed_personnel_ids: set[int] | None = None,
) -> OrganizationSetupOperationResult:
    """
    Import organization structure from Excel.

    IMPORTANT IMPORT BEHAVIOR
    -------------------------
    - Creates missing Directories
    - Creates missing Departments
    - Assigns directory/department role holders when AGM matches
    - Creates membership assignments for matched personnel into the department
    """
    if file_storage is None or not getattr(file_storage, "filename", ""):
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Δεν επιλέχθηκε αρχείο Excel.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    filename = (file_storage.filename or "").lower()
    if not filename.endswith(".xlsx"):
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Υποστηρίζονται μόνο αρχεία .xlsx.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    try:
        file_bytes = file_storage.read()
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
        sheet = workbook.active
    except Exception:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Αποτυχία ανάγνωσης του αρχείου Excel.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage("Το Excel είναι κενό.", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    header_row = rows[0]
    headers = {_clean_cell(cell).upper(): idx for idx, cell in enumerate(header_row)}

    required_headers = {
        "ΔΙΕΥΘΥΝΣΗ",
        "ΔΙΕΥΘΥΝΤΗΣ_ΑΓΜ",
        "ΤΜΗΜΑ",
        "ΠΡΟΙΣΤΑΜΕΝΟΣ_ΑΓΜ",
        "ΒΟΗΘΟΣ_ΑΓΜ",
    }

    missing = [h for h in required_headers if h not in headers]
    if missing:
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(
                FlashMessage(
                    "Λείπουν υποχρεωτικές στήλες: " + ", ".join(missing),
                    "danger",
                ),
            ),
            redirect_service_unit_id=unit.id,
        )

    created_directories = 0
    created_departments = 0
    created_memberships = 0
    assigned_directors = 0
    assigned_managers = 0
    assigned_deputies = 0
    skipped_role_assignments = 0

    try:
        for excel_row in rows[1:]:
            if excel_row is None:
                continue

            directory_name = _clean_cell(excel_row[headers["ΔΙΕΥΘΥΝΣΗ"]])
            director_agm = _normalize_agm(excel_row[headers["ΔΙΕΥΘΥΝΤΗΣ_ΑΓΜ"]])
            department_name = _clean_cell(excel_row[headers["ΤΜΗΜΑ"]])
            manager_agm = _normalize_agm(excel_row[headers["ΠΡΟΙΣΤΑΜΕΝΟΣ_ΑΓΜ"]])
            deputy_agm = _normalize_agm(excel_row[headers["ΒΟΗΘΟΣ_ΑΓΜ"]])

            if not any([directory_name, director_agm, department_name, manager_agm, deputy_agm]):
                continue

            if not directory_name:
                continue

            directory = (
                Directory.query.filter(
                    Directory.service_unit_id == unit.id,
                    Directory.name == directory_name,
                )
                .first()
            )
            if directory is None:
                directory = Directory(
                    service_unit_id=unit.id,
                    name=directory_name,
                )
                db.session.add(directory)
                db.session.flush()
                created_directories += 1

            if director_agm:
                director_personnel_id = _personnel_id_by_agm_for_service_unit(unit.id, director_agm)
                if (
                    director_personnel_id is not None
                    and (
                        allowed_personnel_ids is None
                        or director_personnel_id in allowed_personnel_ids
                    )
                ):
                    if getattr(directory, "director_personnel_id", None) != director_personnel_id:
                        directory.director_personnel_id = director_personnel_id
                        assigned_directors += 1
                else:
                    skipped_role_assignments += 1

            if not department_name:
                continue

            department = (
                Department.query.filter(
                    Department.service_unit_id == unit.id,
                    Department.directory_id == directory.id,
                    Department.name == department_name,
                )
                .first()
            )
            if department is None:
                department = Department(
                    service_unit_id=unit.id,
                    directory_id=directory.id,
                    name=department_name,
                )
                db.session.add(department)
                db.session.flush()
                created_departments += 1

            if manager_agm:
                manager_personnel_id = _personnel_id_by_agm_for_service_unit(unit.id, manager_agm)
                if (
                    manager_personnel_id is not None
                    and (
                        allowed_personnel_ids is None
                        or manager_personnel_id in allowed_personnel_ids
                    )
                ):
                    if getattr(department, "head_personnel_id", None) != manager_personnel_id:
                        department.head_personnel_id = manager_personnel_id
                        assigned_managers += 1

                    membership_exists = PersonnelDepartmentAssignment.query.filter_by(
                        personnel_id=manager_personnel_id,
                        department_id=department.id,
                    ).first()
                    if membership_exists is None:
                        db.session.add(
                            PersonnelDepartmentAssignment(
                                personnel_id=manager_personnel_id,
                                service_unit_id=unit.id,
                                directory_id=directory.id,
                                department_id=department.id,
                                is_primary=True,
                            )
                        )
                        created_memberships += 1
                else:
                    skipped_role_assignments += 1

            if deputy_agm:
                deputy_personnel_id = _personnel_id_by_agm_for_service_unit(unit.id, deputy_agm)
                if (
                    deputy_personnel_id is not None
                    and (
                        allowed_personnel_ids is None
                        or deputy_personnel_id in allowed_personnel_ids
                    )
                ):
                    if getattr(department, "assistant_personnel_id", None) != deputy_personnel_id:
                        department.assistant_personnel_id = deputy_personnel_id
                        assigned_deputies += 1

                    membership_exists = PersonnelDepartmentAssignment.query.filter_by(
                        personnel_id=deputy_personnel_id,
                        department_id=department.id,
                    ).first()
                    if membership_exists is None:
                        db.session.add(
                            PersonnelDepartmentAssignment(
                                personnel_id=deputy_personnel_id,
                                service_unit_id=unit.id,
                                directory_id=directory.id,
                                department_id=department.id,
                                is_primary=False,
                            )
                        )
                        created_memberships += 1
                else:
                    skipped_role_assignments += 1

        db.session.commit()

    except Exception as exc:
        db.session.rollback()
        return OrganizationSetupOperationResult(
            ok=False,
            flashes=(FlashMessage(f"Αποτυχία import: {exc}", "danger"),),
            redirect_service_unit_id=unit.id,
        )

    summary = (
        f"Το import ολοκληρώθηκε. "
        f"Νέες Διευθύνσεις: {created_directories}, "
        f"Νέα Τμήματα: {created_departments}, "
        f"Νέες συμμετοχές προσωπικού: {created_memberships}, "
        f"Διευθυντές: {assigned_directors}, "
        f"Προϊστάμενοι: {assigned_managers}, "
        f"Βοηθοί: {assigned_deputies}."
    )

    flashes = [FlashMessage(summary, "success")]
    if skipped_role_assignments:
        flashes.append(
            FlashMessage(
                f"{skipped_role_assignments} αναθέσεις ρόλων παραλείφθηκαν "
                f"επειδή δεν βρέθηκε ενεργό προσωπικό της ίδιας Υπηρεσίας.",
                "warning",
            )
        )

    return OrganizationSetupOperationResult(
        ok=True,
        flashes=tuple(flashes),
        redirect_service_unit_id=unit.id,
    )
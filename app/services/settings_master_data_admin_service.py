"""
app/services/settings_master_data_admin_service.py

Focused master-data admin services for settings routes.

PURPOSE
-------
This module extracts non-HTTP orchestration from the fat admin/master-data
routes inside `app/blueprints/settings/routes.py`, specifically:

- /settings/ale-kae
- /settings/ale-kae/import
- /settings/cpv
- /settings/cpv/import
- /settings/options/*
- /settings/income-tax
- /settings/withholding-profiles

WHY THIS FILE EXISTS
--------------------
The current settings blueprint mixes already-thin route groups
(ServiceUnits/Suppliers) with several still-fat CRUD/import flows.
Those flows perform:
- validation
- object loading
- persistence
- audit logging
- page-context assembly

That orchestration belongs in the service/use-case layer rather than in route
handlers.

DESIGN
------
- function-first
- one focused module for settings master-data administration
- no generic CRUD framework
- small explicit helper functions per sub-area
"""

from __future__ import annotations

from collections.abc import Mapping
from decimal import Decimal
from typing import Any

from ..audit import log_action, serialize_model
from ..extensions import db
from ..models import AleKae, Cpv, IncomeTaxRule, OptionCategory, OptionValue, WithholdingProfile
from .excel_imports import build_header_index, cell_at, safe_cell_str
from .master_data_service import get_all_option_rows, get_option_category_by_key
from .operation_results import FlashMessage, OperationResult
from .parsing import parse_decimal, parse_optional_int


def _get_or_create_category(key: str, label: str) -> OptionCategory:
    """
    Ensure an OptionCategory row exists and return it.

    NOTE
    ----
    This preserves the current self-healing behavior of the settings option
    pages when seed data has not been run yet.
    """
    category = get_option_category_by_key(key)
    if category:
        if category.label != label:
            category.label = label
            db.session.commit()
        return category

    category = OptionCategory(key=key, label=label)
    db.session.add(category)
    db.session.commit()
    return category


# ----------------------------------------------------------------------
# ALE-KAE
# ----------------------------------------------------------------------
def build_ale_kae_page_context() -> dict[str, Any]:
    """
    Build template context for the ALE-KAE page.
    """
    rows = AleKae.query.order_by(AleKae.ale.asc()).all()
    return {"rows": rows}


def execute_ale_kae_action(form_data: Mapping[str, object]) -> OperationResult:
    """
    Execute create/update/delete action for ALE-KAE rows.
    """
    action = (form_data.get("action") or "").strip()

    if action == "create":
        ale = (form_data.get("ale") or "").strip()
        old_kae = (form_data.get("old_kae") or "").strip() or None
        description = (form_data.get("description") or "").strip() or None
        responsibility = (form_data.get("responsibility") or "").strip() or None

        if not ale:
            return OperationResult(False, (FlashMessage("Το ΑΛΕ είναι υποχρεωτικό.", "danger"),))

        if AleKae.query.filter_by(ale=ale).first():
            return OperationResult(False, (FlashMessage("Υπάρχει ήδη εγγραφή με αυτό το ΑΛΕ.", "danger"),))

        row = AleKae(
            ale=ale,
            old_kae=old_kae,
            description=description,
            responsibility=responsibility,
        )
        db.session.add(row)
        db.session.flush()
        log_action(entity=row, action="CREATE", before=None, after=serialize_model(row))
        db.session.commit()

        return OperationResult(True, (FlashMessage("Η εγγραφή ΑΛΕ-ΚΑΕ προστέθηκε.", "success"),), entity_id=row.id)

    if action == "update":
        row_id = parse_optional_int((form_data.get("id") or "").strip())
        if row_id is None:
            return OperationResult(False, (FlashMessage("Μη έγκυρη εγγραφή.", "danger"),))

        row = AleKae.query.get_or_404(row_id)
        before = serialize_model(row)

        ale = (form_data.get("ale") or "").strip()
        old_kae = (form_data.get("old_kae") or "").strip() or None
        description = (form_data.get("description") or "").strip() or None
        responsibility = (form_data.get("responsibility") or "").strip() or None

        if not ale:
            return OperationResult(False, (FlashMessage("Το ΑΛΕ είναι υποχρεωτικό.", "danger"),))

        exists = AleKae.query.filter(AleKae.ale == ale, AleKae.id != row.id).first()
        if exists:
            return OperationResult(False, (FlashMessage("Υπάρχει ήδη άλλη εγγραφή με αυτό το ΑΛΕ.", "danger"),))

        row.ale = ale
        row.old_kae = old_kae
        row.description = description
        row.responsibility = responsibility

        db.session.flush()
        log_action(entity=row, action="UPDATE", before=before, after=serialize_model(row))
        db.session.commit()

        return OperationResult(True, (FlashMessage("Η εγγραφή ενημερώθηκε.", "success"),), entity_id=row.id)

    if action == "delete":
        row_id = parse_optional_int((form_data.get("id") or "").strip())
        if row_id is None:
            return OperationResult(False, (FlashMessage("Μη έγκυρη εγγραφή.", "danger"),))

        row = AleKae.query.get_or_404(row_id)
        before = serialize_model(row)

        db.session.delete(row)
        db.session.flush()
        log_action(entity=row, action="DELETE", before=before, after=None)
        db.session.commit()

        return OperationResult(True, (FlashMessage("Η εγγραφή διαγράφηκε.", "success"),), entity_id=row.id)

    return OperationResult(False, (FlashMessage("Μη έγκυρη ενέργεια.", "danger"),))


def execute_import_ale_kae(file_storage: Any) -> OperationResult:
    """
    Import ALE-KAE rows from an XLSX file.
    """
    file = file_storage
    if not file or not getattr(file, "filename", None):
        return OperationResult(False, (FlashMessage("Δεν επιλέχθηκε αρχείο.", "danger"),))

    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        return OperationResult(False, (FlashMessage("Επιτρέπεται μόνο αρχείο .xlsx", "danger"),))

    try:
        import openpyxl
        workbook = openpyxl.load_workbook(file, data_only=True)
        worksheet = workbook.active
    except Exception:
        return OperationResult(False, (FlashMessage("Αποτυχία ανάγνωσης Excel. Ελέγξτε το αρχείο.", "danger"),))

    try:
        header_cells = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    except StopIteration:
        return OperationResult(False, (FlashMessage("Το Excel είναι κενό.", "danger"),))

    idx_map = build_header_index(header_cells)

    ale_idx = idx_map.get("αλε", idx_map.get("ale"))
    old_kae_idx = idx_map.get("παλιος καε", idx_map.get("old kae", idx_map.get("old_kae")))
    desc_idx = idx_map.get("περιγραφη", idx_map.get("description"))
    resp_idx = idx_map.get("αρμοδιοτητας", idx_map.get("responsibility"))

    if ale_idx is None:
        return OperationResult(False, (FlashMessage("Το Excel πρέπει να έχει στήλη 'ΑΛΕ'.", "danger"),))

    inserted: list[AleKae] = []
    skipped_missing = 0
    skipped_duplicate = 0

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        ale = safe_cell_str(cell_at(row, ale_idx))
        if not ale:
            skipped_missing += 1
            continue

        if AleKae.query.filter_by(ale=ale).first():
            skipped_duplicate += 1
            continue

        obj = AleKae(
            ale=ale,
            old_kae=safe_cell_str(cell_at(row, old_kae_idx)) or None,
            description=safe_cell_str(cell_at(row, desc_idx)) or None,
            responsibility=safe_cell_str(cell_at(row, resp_idx)) or None,
        )
        db.session.add(obj)
        inserted.append(obj)

    if not inserted:
        return OperationResult(False, (FlashMessage("Δεν εισήχθησαν εγγραφές. Ελέγξτε required πεδία/διπλότυπα.", "warning"),))

    db.session.flush()
    for obj in inserted:
        log_action(entity=obj, action="CREATE", before=None, after=serialize_model(obj))
    db.session.commit()

    return OperationResult(
        True,
        (
            FlashMessage(
                f"Εισαγωγή ολοκληρώθηκε: {len(inserted)} νέες εγγραφές. "
                f"Παραλείφθηκαν: {skipped_missing} (ελλιπή), {skipped_duplicate} (διπλότυπα).",
                "success",
            ),
        ),
    )


# ----------------------------------------------------------------------
# CPV
# ----------------------------------------------------------------------
def build_cpv_page_context() -> dict[str, Any]:
    """
    Build template context for the CPV page.
    """
    rows = Cpv.query.order_by(Cpv.cpv.asc()).all()
    return {"rows": rows}


def execute_cpv_action(form_data: Mapping[str, object]) -> OperationResult:
    """
    Execute create/update/delete action for CPV rows.
    """
    action = (form_data.get("action") or "").strip()

    if action == "create":
        cpv_code = (form_data.get("cpv") or "").strip()
        description = (form_data.get("description") or "").strip() or None

        if not cpv_code:
            return OperationResult(False, (FlashMessage("Το CPV είναι υποχρεωτικό.", "danger"),))

        if Cpv.query.filter_by(cpv=cpv_code).first():
            return OperationResult(False, (FlashMessage("Υπάρχει ήδη εγγραφή με αυτό το CPV.", "danger"),))

        obj = Cpv(cpv=cpv_code, description=description)
        db.session.add(obj)
        db.session.flush()
        log_action(entity=obj, action="CREATE", before=None, after=serialize_model(obj))
        db.session.commit()

        return OperationResult(True, (FlashMessage("Η εγγραφή CPV προστέθηκε.", "success"),), entity_id=obj.id)

    if action == "update":
        obj_id = parse_optional_int((form_data.get("id") or "").strip())
        if obj_id is None:
            return OperationResult(False, (FlashMessage("Μη έγκυρη εγγραφή.", "danger"),))

        obj = Cpv.query.get_or_404(obj_id)
        before = serialize_model(obj)

        cpv_code = (form_data.get("cpv") or "").strip()
        description = (form_data.get("description") or "").strip() or None

        if not cpv_code:
            return OperationResult(False, (FlashMessage("Το CPV είναι υποχρεωτικό.", "danger"),))

        exists = Cpv.query.filter(Cpv.cpv == cpv_code, Cpv.id != obj.id).first()
        if exists:
            return OperationResult(False, (FlashMessage("Υπάρχει ήδη άλλη εγγραφή με αυτό το CPV.", "danger"),))

        obj.cpv = cpv_code
        obj.description = description

        db.session.flush()
        log_action(entity=obj, action="UPDATE", before=before, after=serialize_model(obj))
        db.session.commit()

        return OperationResult(True, (FlashMessage("Η εγγραφή ενημερώθηκε.", "success"),), entity_id=obj.id)

    if action == "delete":
        obj_id = parse_optional_int((form_data.get("id") or "").strip())
        if obj_id is None:
            return OperationResult(False, (FlashMessage("Μη έγκυρη εγγραφή.", "danger"),))

        obj = Cpv.query.get_or_404(obj_id)
        before = serialize_model(obj)

        db.session.delete(obj)
        db.session.flush()
        log_action(entity=obj, action="DELETE", before=before, after=None)
        db.session.commit()

        return OperationResult(True, (FlashMessage("Η εγγραφή διαγράφηκε.", "success"),), entity_id=obj.id)

    return OperationResult(False, (FlashMessage("Μη έγκυρη ενέργεια.", "danger"),))


def execute_import_cpv(file_storage: Any) -> OperationResult:
    """
    Import CPV rows from an XLSX file.
    """
    file = file_storage
    if not file or not getattr(file, "filename", None):
        return OperationResult(False, (FlashMessage("Δεν επιλέχθηκε αρχείο.", "danger"),))

    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        return OperationResult(False, (FlashMessage("Επιτρέπεται μόνο αρχείο .xlsx", "danger"),))

    try:
        import openpyxl
        workbook = openpyxl.load_workbook(file, data_only=True)
        worksheet = workbook.active
    except Exception:
        return OperationResult(False, (FlashMessage("Αποτυχία ανάγνωσης Excel. Ελέγξτε το αρχείο.", "danger"),))

    try:
        header_cells = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
    except StopIteration:
        return OperationResult(False, (FlashMessage("Το Excel είναι κενό.", "danger"),))

    idx_map = build_header_index(header_cells)

    cpv_idx = idx_map.get("cpv")
    desc_idx = idx_map.get("περιγραφη", idx_map.get("description"))

    if cpv_idx is None:
        return OperationResult(False, (FlashMessage("Το Excel πρέπει να έχει στήλη 'CPV'.", "danger"),))

    inserted: list[Cpv] = []
    skipped_missing = 0
    skipped_duplicate = 0

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        cpv_code = safe_cell_str(cell_at(row, cpv_idx))
        if not cpv_code:
            skipped_missing += 1
            continue

        if Cpv.query.filter_by(cpv=cpv_code).first():
            skipped_duplicate += 1
            continue

        obj = Cpv(
            cpv=cpv_code,
            description=safe_cell_str(cell_at(row, desc_idx)) or None,
        )
        db.session.add(obj)
        inserted.append(obj)

    if not inserted:
        return OperationResult(False, (FlashMessage("Δεν εισήχθησαν εγγραφές. Ελέγξτε required πεδία/διπλότυπα.", "warning"),))

    db.session.flush()
    for obj in inserted:
        log_action(entity=obj, action="CREATE", before=None, after=serialize_model(obj))
    db.session.commit()

    return OperationResult(
        True,
        (
            FlashMessage(
                f"Εισαγωγή ολοκληρώθηκε: {len(inserted)} νέες εγγραφές. "
                f"Παραλείφθηκαν: {skipped_missing} (ελλιπή), {skipped_duplicate} (διπλότυπα).",
                "success",
            ),
        ),
    )


# ----------------------------------------------------------------------
# OPTION VALUES
# ----------------------------------------------------------------------
def build_option_values_page_context(*, key: str, label: str) -> dict[str, Any]:
    """
    Build template context for one generic OptionValue category page.
    """
    category = _get_or_create_category(key=key, label=label)
    values = get_all_option_rows(category.key)
    return {
        "category": category,
        "values": values,
        "page_label": label,
    }


def execute_option_value_action(*, key: str, label: str, form_data: Mapping[str, object]) -> OperationResult:
    """
    Execute create/update/delete action for one OptionValue category page.
    """
    category = _get_or_create_category(key=key, label=label)
    action = (form_data.get("action") or "").strip()

    if action == "create":
        value = (form_data.get("value") or "").strip()
        sort_order = parse_optional_int((form_data.get("sort_order") or "").strip()) or 0
        is_active = bool(form_data.get("is_active") == "on")

        if not value:
            return OperationResult(False, (FlashMessage("Η τιμή είναι υποχρεωτική.", "danger"),))

        row = OptionValue(
            category_id=category.id,
            value=value,
            sort_order=sort_order,
            is_active=is_active,
        )
        db.session.add(row)
        db.session.flush()
        log_action(entity=row, action="CREATE", after=serialize_model(row))
        db.session.commit()

        return OperationResult(True, (FlashMessage("Η τιμή προστέθηκε.", "success"),), entity_id=row.id)

    if action == "update":
        option_value_id = parse_optional_int((form_data.get("id") or "").strip())
        if option_value_id is None:
            return OperationResult(False, (FlashMessage("Μη έγκυρη εγγραφή.", "danger"),))

        row = OptionValue.query.filter_by(id=option_value_id, category_id=category.id).first()
        if not row:
            return OperationResult(False, (FlashMessage("Η εγγραφή δεν βρέθηκε.", "danger"),), not_found=True)

        before = serialize_model(row)
        value = (form_data.get("value") or "").strip()
        sort_order = parse_optional_int((form_data.get("sort_order") or "").strip()) or 0
        is_active = bool(form_data.get("is_active") == "on")

        if not value:
            return OperationResult(False, (FlashMessage("Η τιμή είναι υποχρεωτική.", "danger"),))

        row.value = value
        row.sort_order = sort_order
        row.is_active = is_active

        db.session.flush()
        log_action(entity=row, action="UPDATE", before=before, after=serialize_model(row))
        db.session.commit()

        return OperationResult(True, (FlashMessage("Η εγγραφή ενημερώθηκε.", "success"),), entity_id=row.id)

    if action == "delete":
        option_value_id = parse_optional_int((form_data.get("id") or "").strip())
        if option_value_id is None:
            return OperationResult(False, (FlashMessage("Μη έγκυρη εγγραφή.", "danger"),))

        row = OptionValue.query.filter_by(id=option_value_id, category_id=category.id).first()
        if not row:
            return OperationResult(False, (FlashMessage("Η εγγραφή δεν βρέθηκε.", "danger"),), not_found=True)

        before = serialize_model(row)
        db.session.delete(row)
        db.session.flush()
        log_action(entity=row, action="DELETE", before=before)
        db.session.commit()

        return OperationResult(True, (FlashMessage("Η εγγραφή διαγράφηκε.", "success"),), entity_id=row.id)

    return OperationResult(False, (FlashMessage("Μη έγκυρη ενέργεια.", "danger"),))


# ----------------------------------------------------------------------
# INCOME TAX RULES
# ----------------------------------------------------------------------
def build_income_tax_rules_page_context() -> dict[str, Any]:
    """
    Build template context for the IncomeTaxRule page.
    """
    rules = IncomeTaxRule.query.order_by(IncomeTaxRule.description.asc()).all()
    return {"rules": rules}


def execute_income_tax_rule_action(form_data: Mapping[str, object]) -> OperationResult:
    """
    Execute create/update/delete action for IncomeTaxRule rows.
    """
    action = (form_data.get("action") or "").strip()

    if action == "create":
        description = (form_data.get("description") or "").strip()
        rate = parse_decimal(form_data.get("rate_percent")) or Decimal("0.00")
        threshold = parse_decimal(form_data.get("threshold_amount")) or Decimal("150.00")
        is_active = bool(form_data.get("is_active") == "on")

        if not description:
            return OperationResult(False, (FlashMessage("Η περιγραφή είναι υποχρεωτική.", "danger"),))

        rule = IncomeTaxRule(
            description=description,
            rate_percent=rate,
            threshold_amount=threshold,
            is_active=is_active,
        )
        db.session.add(rule)
        db.session.flush()
        log_action(rule, "CREATE", after=serialize_model(rule))
        db.session.commit()

        return OperationResult(True, (FlashMessage("Ο κανόνας ΦΕ προστέθηκε.", "success"),), entity_id=rule.id)

    if action == "update":
        rule_id = parse_optional_int((form_data.get("id") or "").strip())
        if rule_id is None:
            return OperationResult(False, (FlashMessage("Μη έγκυρη εγγραφή.", "danger"),))

        rule = IncomeTaxRule.query.get_or_404(rule_id)
        before = serialize_model(rule)

        description = (form_data.get("description") or "").strip()
        rate = parse_decimal(form_data.get("rate_percent")) or Decimal("0.00")
        threshold = parse_decimal(form_data.get("threshold_amount")) or Decimal("150.00")
        is_active = bool(form_data.get("is_active") == "on")

        if not description:
            return OperationResult(False, (FlashMessage("Η περιγραφή είναι υποχρεωτική.", "danger"),))

        rule.description = description
        rule.rate_percent = rate
        rule.threshold_amount = threshold
        rule.is_active = is_active

        db.session.flush()
        log_action(rule, "UPDATE", before=before, after=serialize_model(rule))
        db.session.commit()

        return OperationResult(True, (FlashMessage("Ο κανόνας ΦΕ ενημερώθηκε.", "success"),), entity_id=rule.id)

    if action == "delete":
        rule_id = parse_optional_int((form_data.get("id") or "").strip())
        if rule_id is None:
            return OperationResult(False, (FlashMessage("Μη έγκυρη εγγραφή.", "danger"),))

        rule = IncomeTaxRule.query.get_or_404(rule_id)
        before = serialize_model(rule)

        db.session.delete(rule)
        db.session.flush()
        log_action(rule, "DELETE", before=before)
        db.session.commit()

        return OperationResult(True, (FlashMessage("Ο κανόνας ΦΕ διαγράφηκε.", "success"),), entity_id=rule.id)

    return OperationResult(False, (FlashMessage("Μη έγκυρη ενέργεια.", "danger"),))


# ----------------------------------------------------------------------
# WITHHOLDING PROFILES
# ----------------------------------------------------------------------
def build_withholding_profiles_page_context() -> dict[str, Any]:
    """
    Build template context for the WithholdingProfile page.
    """
    profiles = WithholdingProfile.query.order_by(WithholdingProfile.description.asc()).all()
    return {"profiles": profiles}


def execute_withholding_profile_action(form_data: Mapping[str, object]) -> OperationResult:
    """
    Execute create/update/delete action for WithholdingProfile rows.
    """
    action = (form_data.get("action") or "").strip()

    if action == "create":
        description = (form_data.get("description") or "").strip()
        mt = parse_decimal(form_data.get("mt_eloa_percent")) or Decimal("0.00")
        ea = parse_decimal(form_data.get("eadhsy_percent")) or Decimal("0.00")
        withholding1 = parse_decimal(form_data.get("withholding1_percent")) or Decimal("0.00")
        withholding2 = parse_decimal(form_data.get("withholding2_percent")) or Decimal("0.00")
        is_active = bool(form_data.get("is_active") == "on")

        if not description:
            return OperationResult(False, (FlashMessage("Η περιγραφή είναι υποχρεωτική.", "danger"),))

        profile = WithholdingProfile(
            description=description,
            mt_eloa_percent=mt,
            eadhsy_percent=ea,
            withholding1_percent=withholding1,
            withholding2_percent=withholding2,
            is_active=is_active,
        )
        db.session.add(profile)
        db.session.flush()
        log_action(profile, "CREATE", after=serialize_model(profile))
        db.session.commit()

        return OperationResult(True, (FlashMessage("Το προφίλ κρατήσεων προστέθηκε.", "success"),), entity_id=profile.id)

    if action == "update":
        profile_id = parse_optional_int((form_data.get("id") or "").strip())
        if profile_id is None:
            return OperationResult(False, (FlashMessage("Μη έγκυρη εγγραφή.", "danger"),))

        profile = WithholdingProfile.query.get_or_404(profile_id)
        before = serialize_model(profile)

        description = (form_data.get("description") or "").strip()
        mt = parse_decimal(form_data.get("mt_eloa_percent")) or Decimal("0.00")
        ea = parse_decimal(form_data.get("eadhsy_percent")) or Decimal("0.00")
        withholding1 = parse_decimal(form_data.get("withholding1_percent")) or Decimal("0.00")
        withholding2 = parse_decimal(form_data.get("withholding2_percent")) or Decimal("0.00")
        is_active = bool(form_data.get("is_active") == "on")

        if not description:
            return OperationResult(False, (FlashMessage("Η περιγραφή είναι υποχρεωτική.", "danger"),))

        profile.description = description
        profile.mt_eloa_percent = mt
        profile.eadhsy_percent = ea
        profile.withholding1_percent = withholding1
        profile.withholding2_percent = withholding2
        profile.is_active = is_active

        db.session.flush()
        log_action(profile, "UPDATE", before=before, after=serialize_model(profile))
        db.session.commit()

        return OperationResult(True, (FlashMessage("Το προφίλ κρατήσεων ενημερώθηκε.", "success"),), entity_id=profile.id)

    if action == "delete":
        profile_id = parse_optional_int((form_data.get("id") or "").strip())
        if profile_id is None:
            return OperationResult(False, (FlashMessage("Μη έγκυρη εγγραφή.", "danger"),))

        profile = WithholdingProfile.query.get_or_404(profile_id)
        before = serialize_model(profile)

        db.session.delete(profile)
        db.session.flush()
        log_action(profile, "DELETE", before=before)
        db.session.commit()

        return OperationResult(True, (FlashMessage("Το προφίλ κρατήσεων διαγράφηκε.", "success"),), entity_id=profile.id)

    return OperationResult(False, (FlashMessage("Μη έγκυρη ενέργεια.", "danger"),))


__all__ = [
    "build_ale_kae_page_context",
    "execute_ale_kae_action",
    "execute_import_ale_kae",
    "build_cpv_page_context",
    "execute_cpv_action",
    "execute_import_cpv",
    "build_option_values_page_context",
    "execute_option_value_action",
    "build_income_tax_rules_page_context",
    "execute_income_tax_rule_action",
    "build_withholding_profiles_page_context",
    "execute_withholding_profile_action",
]

"""
app/blueprints/admin/routes.py

Admin Routes – Enterprise Administration Module

Includes:
- Personnel Management (organizational directory)
- Excel import for Personnel (admin-only)
- Consolidated Service Unit organization management

UPDATED ORGANIZATIONAL MODEL:
- Personnel belongs to:
  - one ServiceUnit (required for new/updated non-neutral entries)
  - optional Directory (Διεύθυνση) of that ServiceUnit
  - optional Department (Τμήμα) of that Directory + ServiceUnit

CONSOLIDATED SETUP PAGE:
- /admin/organization-setup
- Per ServiceUnit, manages:
  - Directories
  - Departments
  - Directory director assignments
  - Department head / assistant assignments
  - Excel import for structure + role assignments

PERMISSIONS (server-side, UI never trusted):
- Admin: manage all
- Manager (ServiceUnit manager only): manage ONLY their own ServiceUnit
- Viewers / deputies: never mutate here

VALIDATION (server-side):
- Directory must belong to selected ServiceUnit
- Department must belong to selected Directory AND ServiceUnit
- Assigned personnel must be active and belong to the same ServiceUnit
- Manager cannot operate on another ServiceUnit

IMPORTANT:
- Any UI filtering is convenience only. All checks are server-side.
- Audit pattern:
    db.session.flush() -> log_action(...) -> db.session.commit()
- SQLite is used in dev, but logic remains PostgreSQL-ready.
"""

from __future__ import annotations

from functools import wraps
from typing import Optional, Set
import unicodedata

from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    abort,
)
from flask_login import login_required, current_user

from ...extensions import db
from ...models import Personnel, ServiceUnit, Directory, Department
from ...audit import log_action, serialize_model

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")


# -------------------------------------------------------
# PERMISSION DECORATORS
# -------------------------------------------------------
def admin_or_manager_required(func):
    """
    Allow admin OR ServiceUnit manager (NOT deputy).

    SECURITY:
    - Server-side enforcement (UI never trusted).
    - Manager is determined via current_user.is_manager().
    """
    @wraps(func)
    def wrapper(*args, **kwargs):
        if not current_user.is_authenticated:
            abort(403)

        if getattr(current_user, "is_admin", False):
            return func(*args, **kwargs)

        is_mgr = getattr(current_user, "is_manager", None)
        if callable(is_mgr) and is_mgr():
            return func(*args, **kwargs)

        abort(403)

    return wrapper


def admin_required(func):
    """Admin-only (server-side)."""
    @wraps(func)
    def wrapper(*args, **kwargs):
        if not current_user.is_authenticated or not getattr(current_user, "is_admin", False):
            abort(403)
        return func(*args, **kwargs)

    return wrapper


# -------------------------------------------------------
# HELPERS
# -------------------------------------------------------
def _parse_optional_int(value: str | None) -> Optional[int]:
    """Parse optional integer from form/query. Returns None if empty/invalid."""
    if not value:
        return None
    try:
        return int(str(value).strip())
    except ValueError:
        return None


def _service_units_for_dropdown():
    """ServiceUnits for dropdown selection (sorted)."""
    return ServiceUnit.query.order_by(ServiceUnit.description.asc()).all()


def _directories_for_dropdown():
    """All Directories (sorted). UI may filter client-side; server validates."""
    return Directory.query.order_by(Directory.service_unit_id.asc(), Directory.name.asc()).all()


def _departments_for_dropdown():
    """All Departments (sorted). UI may filter client-side; server validates."""
    return Department.query.order_by(
        Department.service_unit_id.asc(),
        Department.directory_id.asc(),
        Department.name.asc(),
    ).all()


def _normalize_header(text: str) -> str:
    """
    Normalize Excel headers:
    - lowercase
    - trim spaces
    - remove diacritics (e.g. Περιγραφή == Περιγραφη)
    """
    if text is None:
        return ""
    s = " ".join(str(text).strip().lower().split())
    s = "".join(ch for ch in unicodedata.normalize("NFD", s) if unicodedata.category(ch) != "Mn")
    return s


def _safe_str(v) -> str:
    """Convert excel cell to trimmed string."""
    if v is None:
        return ""
    return str(v).strip()


def _match_service_unit(service_value: str) -> Optional[ServiceUnit]:
    """
    Match service unit from a text value.

    Tries (case-insensitive):
    - code
    - short_name
    - description
    """
    sv = (service_value or "").strip()
    if not sv:
        return None

    q = ServiceUnit.query

    su = q.filter(ServiceUnit.code.isnot(None)).filter(ServiceUnit.code.ilike(sv)).first()
    if su:
        return su

    su = q.filter(ServiceUnit.short_name.isnot(None)).filter(ServiceUnit.short_name.ilike(sv)).first()
    if su:
        return su

    su = q.filter(ServiceUnit.description.ilike(sv)).first()
    if su:
        return su

    return None


def _effective_scope_service_unit_id_for_manager_or_none() -> Optional[int]:
    """
    For non-admin manager: they are scoped to their own service unit.
    For admin: return None (no restriction).
    """
    if getattr(current_user, "is_admin", False):
        return None
    return getattr(current_user, "service_unit_id", None)


def _validate_service_unit_required(service_unit_id: Optional[int]) -> bool:
    """
    ServiceUnit is required for new/updated entries.

    SECURITY:
    - Prevents forging non-existent service_unit_id.
    """
    if service_unit_id is None:
        return False
    return ServiceUnit.query.get(service_unit_id) is not None


def _validate_directory_for_service_unit(directory_id: Optional[int], service_unit_id: int) -> bool:
    """
    Validate that directory exists and belongs to the given service unit.

    Rules:
    - directory_id can be None => OK
    - else Directory.service_unit_id must match
    """
    if directory_id is None:
        return True
    d = Directory.query.get(directory_id)
    return bool(d and d.service_unit_id == service_unit_id)


def _validate_department_for_directory_and_service_unit(
    department_id: Optional[int],
    directory_id: Optional[int],
    service_unit_id: int,
) -> bool:
    """
    Validate that department exists and belongs to:
    - the given directory_id
    - the given service_unit_id

    Rules:
    - department_id can be None => OK
    - If department_id is provided, directory_id MUST also be provided.
    """
    if department_id is None:
        return True

    if directory_id is None:
        return False

    dep = Department.query.get(department_id)
    if not dep:
        return False

    return bool(dep.service_unit_id == service_unit_id and dep.directory_id == directory_id)


def _active_personnel_ids_for_service_unit(service_unit_id: int) -> Set[int]:
    """Active personnel IDs of a service unit (for server-side validation)."""
    rows = Personnel.query.filter_by(is_active=True, service_unit_id=service_unit_id).all()
    return {p.id for p in rows}


def _active_personnel_for_service_unit(service_unit_id: int):
    """Active personnel list of a service unit (for dropdowns)."""
    return (
        Personnel.query.filter_by(is_active=True, service_unit_id=service_unit_id)
        .order_by(Personnel.last_name.asc(), Personnel.first_name.asc())
        .all()
    )


def _get_posted_service_unit_id_or_abort() -> int:
    """
    Resolve target ServiceUnit for POST actions.

    Admin:
    - may post any service_unit_id

    Manager:
    - forced to their own service unit
    """
    if getattr(current_user, "is_admin", False):
        posted_su_id = _parse_optional_int(request.form.get("service_unit_id"))
    else:
        posted_su_id = getattr(current_user, "service_unit_id", None)

    if not posted_su_id:
        flash("Η υπηρεσία είναι υποχρεωτική.", "danger")
        raise ValueError("missing service_unit_id")

    if not getattr(current_user, "is_admin", False):
        if posted_su_id != getattr(current_user, "service_unit_id", None):
            abort(403)

    return posted_su_id


# -------------------------------------------------------
# PERSONNEL LIST
# -------------------------------------------------------
@admin_bp.route("/personnel")
@login_required
@admin_or_manager_required
def personnel_list():
    """
    List organizational personnel.

    Admin: sees all
    Manager: sees only personnel of their ServiceUnit
    """
    q = Personnel.query

    scope_su_id = _effective_scope_service_unit_id_for_manager_or_none()
    if scope_su_id:
        q = q.filter(Personnel.service_unit_id == scope_su_id)

    personnel = (
        q.order_by(
            Personnel.rank.asc(),
            Personnel.last_name.asc(),
            Personnel.first_name.asc(),
        ).all()
    )

    return render_template("admin/personnel_list.html", personnel=personnel)


# -------------------------------------------------------
# IMPORT PERSONNEL (EXCEL) - ADMIN ONLY
# -------------------------------------------------------
@admin_bp.route("/personnel/import", methods=["POST"])
@login_required
@admin_required
def import_personnel():
    """
    Import Personnel from Excel (admin-only).

    Required columns:
    - ΑΓΜ
    - ΟΝΟΜΑ
    - ΕΠΩΝΥΜΟ

    Optional columns:
    - ΑΕΜ
    - ΒΑΘΜΟΣ
    - ΕΙΔΙΚΟΤΗΤΑ
    - ΥΠΗΡΕΣΙΑ (matches ServiceUnit by code OR short_name OR description)

    NOTE:
    - Organizational structure (Directory/Department) is NOT imported here.
      That is handled from the consolidated organization setup page.
    """
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Δεν επιλέχθηκε αρχείο.", "danger")
        return redirect(url_for("admin.personnel_list"))

    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        flash("Επιτρέπεται μόνο αρχείο .xlsx", "danger")
        return redirect(url_for("admin.personnel_list"))

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception:
        flash("Αποτυχία ανάγνωσης Excel. Ελέγξτε το αρχείο.", "danger")
        return redirect(url_for("admin.personnel_list"))

    try:
        header_cells = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    except StopIteration:
        flash("Το Excel είναι κενό.", "danger")
        return redirect(url_for("admin.personnel_list"))

    headers = [str(h).strip() if h is not None else "" for h in header_cells]
    idx_map = {_normalize_header(h): i for i, h in enumerate(headers) if _normalize_header(h)}

    agm_idx = idx_map.get("αγμ", idx_map.get("agm"))
    first_idx = idx_map.get("ονομα", idx_map.get("first name", idx_map.get("first_name")))
    last_idx = idx_map.get("επωνυμο", idx_map.get("last name", idx_map.get("last_name")))

    aem_idx = idx_map.get("αεμ", idx_map.get("aem"))
    rank_idx = idx_map.get("βαθμος", idx_map.get("rank"))
    spec_idx = idx_map.get("ειδικοτητα", idx_map.get("specialty"))
    service_idx = idx_map.get("υπηρεσια", idx_map.get("service"))

    if agm_idx is None or first_idx is None or last_idx is None:
        flash("Το Excel πρέπει να έχει στήλες: ΑΓΜ, ΟΝΟΜΑ, ΕΠΩΝΥΜΟ (1η γραμμή).", "danger")
        return redirect(url_for("admin.personnel_list"))

    inserted_people: list[Personnel] = []
    skipped_missing = 0
    skipped_duplicate = 0
    skipped_bad_service = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        def _cell(i: Optional[int]):
            if i is None:
                return None
            if i >= len(row):
                return None
            return row[i]

        agm = _safe_str(_cell(agm_idx))
        first_name = _safe_str(_cell(first_idx))
        last_name = _safe_str(_cell(last_idx))

        if not agm or not first_name or not last_name:
            skipped_missing += 1
            continue

        if Personnel.query.filter_by(agm=agm).first():
            skipped_duplicate += 1
            continue

        aem = _safe_str(_cell(aem_idx)) or None
        rank = _safe_str(_cell(rank_idx)) or None
        specialty = _safe_str(_cell(spec_idx)) or None

        service_unit_id = None
        if service_idx is not None:
            service_val = _safe_str(_cell(service_idx))
            if service_val:
                su = _match_service_unit(service_val)
                if not su:
                    skipped_bad_service += 1
                    continue
                service_unit_id = su.id

        person = Personnel(
            agm=agm,
            aem=aem,
            rank=rank,
            specialty=specialty,
            first_name=first_name,
            last_name=last_name,
            is_active=True,
            service_unit_id=service_unit_id,
            directory_id=None,
            department_id=None,
        )
        db.session.add(person)
        inserted_people.append(person)

    if not inserted_people:
        flash(
            "Δεν εισήχθησαν εγγραφές. Ελέγξτε required πεδία/διπλότυπα/Υπηρεσία.",
            "warning",
        )
        return redirect(url_for("admin.personnel_list"))

    db.session.flush()
    for person in inserted_people:
        log_action(person, "CREATE", before=None, after=serialize_model(person))
    db.session.commit()

    flash(
        f"Εισαγωγή ολοκληρώθηκε: {len(inserted_people)} νέες εγγραφές. "
        f"Παραλείφθηκαν: {skipped_missing} (ελλιπή), {skipped_duplicate} (διπλότυπα ΑΓΜ), "
        f"{skipped_bad_service} (μη έγκυρη Υπηρεσία).",
        "success",
    )
    return redirect(url_for("admin.personnel_list"))


# -------------------------------------------------------
# CREATE PERSONNEL
# -------------------------------------------------------
@admin_bp.route("/personnel/new", methods=["GET", "POST"])
@login_required
@admin_or_manager_required
def create_personnel():
    """
    Create new Personnel record.

    Admin:
    - can set any service unit / directory / department
    Manager:
    - can set ONLY their own service unit (forced server-side)
    """
    service_units = _service_units_for_dropdown()
    directories = _directories_for_dropdown()
    departments = _departments_for_dropdown()

    if request.method == "POST":
        agm = (request.form.get("agm") or "").strip()
        aem = (request.form.get("aem") or "").strip()
        rank = (request.form.get("rank") or "").strip()
        specialty = (request.form.get("specialty") or "").strip()
        first_name = (request.form.get("first_name") or "").strip()
        last_name = (request.form.get("last_name") or "").strip()

        directory_id = _parse_optional_int(request.form.get("directory_id"))
        department_id = _parse_optional_int(request.form.get("department_id"))

        if getattr(current_user, "is_admin", False):
            service_unit_id = _parse_optional_int(request.form.get("service_unit_id"))
        else:
            service_unit_id = getattr(current_user, "service_unit_id", None)

        if not agm or not first_name or not last_name:
            flash("ΑΓΜ, Όνομα και Επώνυμο είναι υποχρεωτικά.", "danger")
            return redirect(url_for("admin.create_personnel"))

        if Personnel.query.filter_by(agm=agm).first():
            flash("Υπάρχει ήδη προσωπικό με αυτό το ΑΓΜ.", "danger")
            return redirect(url_for("admin.create_personnel"))

        if not _validate_service_unit_required(service_unit_id):
            flash("Η Υπηρεσία είναι υποχρεωτική και πρέπει να είναι έγκυρη.", "danger")
            return redirect(url_for("admin.create_personnel"))

        if not _validate_directory_for_service_unit(directory_id, service_unit_id):
            flash("Μη έγκυρη Διεύθυνση για την επιλεγμένη Υπηρεσία.", "danger")
            return redirect(url_for("admin.create_personnel"))

        if not _validate_department_for_directory_and_service_unit(department_id, directory_id, service_unit_id):
            flash("Μη έγκυρο Τμήμα για την επιλεγμένη Διεύθυνση/Υπηρεσία.", "danger")
            return redirect(url_for("admin.create_personnel"))

        person = Personnel(
            agm=agm,
            aem=aem or None,
            rank=rank or None,
            specialty=specialty or None,
            first_name=first_name,
            last_name=last_name,
            is_active=True,
            service_unit_id=service_unit_id,
            directory_id=directory_id,
            department_id=department_id,
        )

        db.session.add(person)
        db.session.flush()
        log_action(person, "CREATE", before=None, after=serialize_model(person))
        db.session.commit()

        flash("Το προσωπικό καταχωρήθηκε.", "success")
        return redirect(url_for("admin.personnel_list"))

    return render_template(
        "admin/personnel_form.html",
        person=None,
        form_title="Νέο Προσωπικό",
        service_units=service_units,
        directories=directories,
        departments=departments,
    )


# -------------------------------------------------------
# EDIT PERSONNEL
# -------------------------------------------------------
@admin_bp.route("/personnel/<int:personnel_id>/edit", methods=["GET", "POST"])
@login_required
@admin_or_manager_required
def edit_personnel(personnel_id: int):
    """
    Edit Personnel.

    Admin:
    - can edit any Personnel
    Manager:
    - can edit ONLY Personnel of their own ServiceUnit
    """
    person = Personnel.query.get_or_404(personnel_id)

    if not getattr(current_user, "is_admin", False):
        scope_su_id = getattr(current_user, "service_unit_id", None)
        if not scope_su_id or person.service_unit_id != scope_su_id:
            abort(403)

    service_units = _service_units_for_dropdown()
    directories = _directories_for_dropdown()
    departments = _departments_for_dropdown()

    if request.method == "POST":
        before_snapshot = serialize_model(person)

        agm = (request.form.get("agm") or "").strip()
        aem = (request.form.get("aem") or "").strip()
        rank = (request.form.get("rank") or "").strip()
        specialty = (request.form.get("specialty") or "").strip()
        first_name = (request.form.get("first_name") or "").strip()
        last_name = (request.form.get("last_name") or "").strip()

        directory_id = _parse_optional_int(request.form.get("directory_id"))
        department_id = _parse_optional_int(request.form.get("department_id"))

        if getattr(current_user, "is_admin", False):
            service_unit_id = _parse_optional_int(request.form.get("service_unit_id"))
        else:
            service_unit_id = getattr(current_user, "service_unit_id", None)

        is_active = bool(request.form.get("is_active"))

        if not agm or not first_name or not last_name:
            flash("ΑΓΜ, Όνομα και Επώνυμο είναι υποχρεωτικά.", "danger")
            return redirect(url_for("admin.edit_personnel", personnel_id=person.id))

        existing = Personnel.query.filter(Personnel.agm == agm, Personnel.id != person.id).first()
        if existing:
            flash("Υπάρχει ήδη προσωπικό με αυτό το ΑΓΜ.", "danger")
            return redirect(url_for("admin.edit_personnel", personnel_id=person.id))

        if not _validate_service_unit_required(service_unit_id):
            flash("Η Υπηρεσία είναι υποχρεωτική και πρέπει να είναι έγκυρη.", "danger")
            return redirect(url_for("admin.edit_personnel", personnel_id=person.id))

        if not _validate_directory_for_service_unit(directory_id, service_unit_id):
            flash("Μη έγκυρη Διεύθυνση για την επιλεγμένη Υπηρεσία.", "danger")
            return redirect(url_for("admin.edit_personnel", personnel_id=person.id))

        if not _validate_department_for_directory_and_service_unit(department_id, directory_id, service_unit_id):
            flash("Μη έγκυρο Τμήμα για την επιλεγμένη Διεύθυνση/Υπηρεσία.", "danger")
            return redirect(url_for("admin.edit_personnel", personnel_id=person.id))

        if not getattr(current_user, "is_admin", False):
            scope_su_id = getattr(current_user, "service_unit_id", None)
            if not scope_su_id or service_unit_id != scope_su_id:
                abort(403)

        person.agm = agm
        person.aem = aem or None
        person.rank = rank or None
        person.specialty = specialty or None
        person.first_name = first_name
        person.last_name = last_name
        person.service_unit_id = service_unit_id
        person.directory_id = directory_id
        person.department_id = department_id
        person.is_active = is_active

        db.session.flush()
        log_action(person, "UPDATE", before=before_snapshot, after=serialize_model(person))
        db.session.commit()

        flash("Το προσωπικό ενημερώθηκε.", "success")
        return redirect(url_for("admin.personnel_list"))

    return render_template(
        "admin/personnel_form.html",
        person=person,
        form_title="Επεξεργασία Προσωπικού",
        service_units=service_units,
        directories=directories,
        departments=departments,
    )


# -------------------------------------------------------
# CONSOLIDATED ORGANIZATION PAGE
# -------------------------------------------------------
@admin_bp.route("/organization-setup", methods=["GET", "POST"])
@login_required
@admin_or_manager_required
def organization_setup():
    """
    Consolidated ServiceUnit organization management page.

    Admin:
    - can select any service unit

    Manager:
    - forced to their own service unit (server-side)

    Supported actions:
    - create_directory
    - update_directory
    - delete_directory
    - create_department
    - update_department
    - delete_department
    - update_directory_director
    - update_department_roles
    - import

    Excel import behavior:
    - Can CREATE missing Directories / Departments
    - Can UPDATE role assignments if AGM values match active personnel
    - Rows that are invalid are skipped
    """
    if getattr(current_user, "is_admin", False):
        service_unit_id = _parse_optional_int(request.args.get("service_unit_id"))
    else:
        service_unit_id = getattr(current_user, "service_unit_id", None)

    service_units = _service_units_for_dropdown()

    unit = ServiceUnit.query.get(service_unit_id) if service_unit_id else None
    directories = []
    departments = []
    personnel_list = []

    if unit:
        if not getattr(current_user, "is_admin", False):
            if unit.id != getattr(current_user, "service_unit_id", None):
                abort(403)

        directories = (
            Directory.query.filter_by(service_unit_id=unit.id)
            .order_by(Directory.name.asc())
            .all()
        )
        departments = (
            Department.query.filter_by(service_unit_id=unit.id)
            .order_by(Department.directory_id.asc(), Department.name.asc())
            .all()
        )
        personnel_list = _active_personnel_for_service_unit(unit.id)

    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        try:
            posted_su_id = _get_posted_service_unit_id_or_abort()
        except ValueError:
            return redirect(url_for("admin.organization_setup"))

        unit = ServiceUnit.query.get_or_404(posted_su_id)
        allowed_personnel_ids = _active_personnel_ids_for_service_unit(unit.id)

        def _validate_personnel(pid: Optional[int]) -> Optional[int]:
            """
            Return the personnel ID only if:
            - it exists in the active personnel set of the same service unit
            - otherwise return None
            """
            if pid is None:
                return None
            return pid if pid in allowed_personnel_ids else None

        # ---------------------------------------------------
        # DIRECTORY CREATE
        # ---------------------------------------------------
        if action == "create_directory":
            name = (request.form.get("directory_name") or "").strip()

            if not name:
                flash("Η ονομασία Διεύθυνσης είναι υποχρεωτική.", "danger")
                return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

            exists = Directory.query.filter_by(service_unit_id=unit.id, name=name).first()
            if exists:
                flash("Υπάρχει ήδη Διεύθυνση με αυτή την ονομασία στην Υπηρεσία.", "danger")
                return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

            d = Directory(
                service_unit_id=unit.id,
                name=name,
                is_active=True,
                director_personnel_id=None,
            )
            db.session.add(d)
            db.session.flush()
            log_action(entity=d, action="CREATE", before=None, after=serialize_model(d))
            db.session.commit()

            flash("Η Διεύθυνση δημιουργήθηκε.", "success")
            return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

        # ---------------------------------------------------
        # DIRECTORY UPDATE
        # ---------------------------------------------------
        if action == "update_directory":
            directory_id = _parse_optional_int(request.form.get("directory_id"))
            if directory_id is None:
                flash("Μη έγκυρη Διεύθυνση.", "danger")
                return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

            d = Directory.query.get_or_404(directory_id)
            if d.service_unit_id != unit.id:
                abort(403)

            before = serialize_model(d)

            name = (request.form.get("directory_name") or "").strip()
            is_active = bool(request.form.get("is_active"))

            if not name:
                flash("Η ονομασία Διεύθυνσης είναι υποχρεωτική.", "danger")
                return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

            exists = Directory.query.filter(
                Directory.service_unit_id == unit.id,
                Directory.name == name,
                Directory.id != d.id,
            ).first()
            if exists:
                flash("Υπάρχει ήδη άλλη Διεύθυνση με αυτή την ονομασία.", "danger")
                return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

            d.name = name
            d.is_active = is_active

            db.session.flush()
            log_action(entity=d, action="UPDATE", before=before, after=serialize_model(d))
            db.session.commit()

            flash("Η Διεύθυνση ενημερώθηκε.", "success")
            return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

        # ---------------------------------------------------
        # DIRECTORY DELETE
        # ---------------------------------------------------
        if action == "delete_directory":
            directory_id = _parse_optional_int(request.form.get("directory_id"))
            if directory_id is None:
                flash("Μη έγκυρη Διεύθυνση.", "danger")
                return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

            d = Directory.query.get_or_404(directory_id)
            if d.service_unit_id != unit.id:
                abort(403)

            before = serialize_model(d)

            # Defensive cleanup:
            # do not rely only on FK ondelete behavior, especially in SQLite dev.
            Personnel.query.filter_by(directory_id=d.id).update({"directory_id": None}, synchronize_session=False)
            Personnel.query.filter_by(department_id=None, service_unit_id=unit.id)

            for dep in Department.query.filter_by(directory_id=d.id).all():
                Personnel.query.filter_by(department_id=dep.id).update({"department_id": None}, synchronize_session=False)

            db.session.delete(d)
            db.session.flush()
            log_action(entity=d, action="DELETE", before=before, after=None)
            db.session.commit()

            flash("Η Διεύθυνση διαγράφηκε.", "success")
            return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

        # ---------------------------------------------------
        # DEPARTMENT CREATE
        # ---------------------------------------------------
        if action == "create_department":
            directory_id = _parse_optional_int(request.form.get("directory_id"))
            name = (request.form.get("department_name") or "").strip()

            if not directory_id:
                flash("Η Διεύθυνση είναι υποχρεωτική για δημιουργία Τμήματος.", "danger")
                return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

            if not name:
                flash("Η ονομασία Τμήματος είναι υποχρεωτική.", "danger")
                return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

            d = Directory.query.get(directory_id)
            if not d or d.service_unit_id != unit.id:
                flash("Μη έγκυρη Διεύθυνση για την Υπηρεσία.", "danger")
                return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

            exists = Department.query.filter_by(directory_id=d.id, name=name).first()
            if exists:
                flash("Υπάρχει ήδη Τμήμα με αυτή την ονομασία στη συγκεκριμένη Διεύθυνση.", "danger")
                return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

            dep = Department(
                service_unit_id=unit.id,
                directory_id=d.id,
                name=name,
                is_active=True,
                head_personnel_id=None,
                assistant_personnel_id=None,
            )
            db.session.add(dep)
            db.session.flush()
            log_action(entity=dep, action="CREATE", before=None, after=serialize_model(dep))
            db.session.commit()

            flash("Το Τμήμα δημιουργήθηκε.", "success")
            return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

        # ---------------------------------------------------
        # DEPARTMENT UPDATE
        # ---------------------------------------------------
        if action == "update_department":
            department_id = _parse_optional_int(request.form.get("department_id"))
            new_directory_id = _parse_optional_int(request.form.get("directory_id"))
            name = (request.form.get("department_name") or "").strip()

            if department_id is None:
                flash("Μη έγκυρο Τμήμα.", "danger")
                return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

            if new_directory_id is None:
                flash("Η Διεύθυνση είναι υποχρεωτική.", "danger")
                return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

            if not name:
                flash("Η ονομασία Τμήματος είναι υποχρεωτική.", "danger")
                return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

            dep = Department.query.get_or_404(department_id)
            if dep.service_unit_id != unit.id:
                abort(403)

            new_directory = Directory.query.get_or_404(new_directory_id)
            if new_directory.service_unit_id != unit.id:
                abort(403)

            before = serialize_model(dep)

            exists = Department.query.filter(
                Department.directory_id == new_directory.id,
                Department.name == name,
                Department.id != dep.id,
            ).first()
            if exists:
                flash("Υπάρχει ήδη άλλο Τμήμα με αυτή την ονομασία στη συγκεκριμένη Διεύθυνση.", "danger")
                return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

            dep.directory_id = new_directory.id
            dep.name = name
            dep.is_active = bool(request.form.get("is_active"))

            db.session.flush()
            log_action(entity=dep, action="UPDATE", before=before, after=serialize_model(dep))
            db.session.commit()

            flash("Το Τμήμα ενημερώθηκε.", "success")
            return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

        # ---------------------------------------------------
        # DEPARTMENT DELETE
        # ---------------------------------------------------
        if action == "delete_department":
            department_id = _parse_optional_int(request.form.get("department_id"))
            if department_id is None:
                flash("Μη έγκυρο Τμήμα.", "danger")
                return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

            dep = Department.query.get_or_404(department_id)
            if dep.service_unit_id != unit.id:
                abort(403)

            before = serialize_model(dep)

            # Defensive cleanup for SQLite dev and UI consistency.
            Personnel.query.filter_by(department_id=dep.id).update({"department_id": None}, synchronize_session=False)

            db.session.delete(dep)
            db.session.flush()
            log_action(entity=dep, action="DELETE", before=before, after=None)
            db.session.commit()

            flash("Το Τμήμα διαγράφηκε.", "success")
            return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

        # ---------------------------------------------------
        # UPDATE DIRECTORY DIRECTOR
        # ---------------------------------------------------
        if action == "update_directory_director":
            directory_id = _parse_optional_int(request.form.get("directory_id"))
            director_pid = _validate_personnel(_parse_optional_int(request.form.get("director_personnel_id")))

            if directory_id is None:
                flash("Μη έγκυρη Διεύθυνση.", "danger")
                return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

            d = Directory.query.get_or_404(directory_id)
            if d.service_unit_id != unit.id:
                abort(403)

            before = serialize_model(d)
            d.director_personnel_id = director_pid

            db.session.flush()
            log_action(entity=d, action="UPDATE", before=before, after=serialize_model(d))
            db.session.commit()

            flash("Ο Διευθυντής Διεύθυνσης ενημερώθηκε.", "success")
            return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

        # ---------------------------------------------------
        # UPDATE DEPARTMENT HEAD / ASSISTANT
        # ---------------------------------------------------
        if action == "update_department_roles":
            department_id = _parse_optional_int(request.form.get("department_id"))
            head_pid = _validate_personnel(_parse_optional_int(request.form.get("head_personnel_id")))
            assistant_pid = _validate_personnel(_parse_optional_int(request.form.get("assistant_personnel_id")))

            if department_id is None:
                flash("Μη έγκυρο Τμήμα.", "danger")
                return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

            if head_pid and assistant_pid and head_pid == assistant_pid:
                flash("Ο ίδιος/η ίδια δεν μπορεί να είναι και Προϊστάμενος και Βοηθός.", "danger")
                return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

            dep = Department.query.get_or_404(department_id)
            if dep.service_unit_id != unit.id:
                abort(403)

            before = serialize_model(dep)
            dep.head_personnel_id = head_pid
            dep.assistant_personnel_id = assistant_pid

            db.session.flush()
            log_action(entity=dep, action="UPDATE", before=before, after=serialize_model(dep))
            db.session.commit()

            flash("Οι ρόλοι Τμήματος ενημερώθηκαν.", "success")
            return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

        # ---------------------------------------------------
        # EXCEL IMPORT (STRUCTURE + ROLES)
        # ---------------------------------------------------
        if action == "import":
            file = request.files.get("file")
            if not file or not file.filename:
                flash("Δεν επιλέχθηκε αρχείο.", "danger")
                return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

            filename = (file.filename or "").lower()
            if not filename.endswith(".xlsx"):
                flash("Επιτρέπεται μόνο αρχείο .xlsx", "danger")
                return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

            try:
                import openpyxl
                wb = openpyxl.load_workbook(file, data_only=True)
                ws = wb.active
            except Exception:
                flash("Αποτυχία ανάγνωσης Excel. Ελέγξτε το αρχείο.", "danger")
                return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

            try:
                header_cells = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            except StopIteration:
                flash("Το Excel είναι κενό.", "danger")
                return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

            headers = [str(h).strip() if h is not None else "" for h in header_cells]
            idx_map = {_normalize_header(h): i for i, h in enumerate(headers) if _normalize_header(h)}

            # Expected headers (Greek/English aliases):
            # ΔΙΕΥΘΥΝΣΗ, ΔΙΕΥΘΥΝΤΗΣ_ΑΓΜ
            # ΤΜΗΜΑ, ΠΡΟΙΣΤΑΜΕΝΟΣ_ΑΓΜ, ΒΟΗΘΟΣ_ΑΓΜ
            dir_name_idx = idx_map.get("διευθυνση", idx_map.get("directory"))
            dir_director_agm_idx = idx_map.get("διευθυντης_αγμ", idx_map.get("director_agm"))

            dep_name_idx = idx_map.get("τμημα", idx_map.get("department"))
            dep_head_agm_idx = idx_map.get("προισταμενος_αγμ", idx_map.get("head_agm"))
            dep_assist_agm_idx = idx_map.get("βοηθος_αγμ", idx_map.get("assistant_agm"))

            if dir_name_idx is None and dep_name_idx is None:
                flash(
                    "Το Excel πρέπει να έχει τουλάχιστον μία από τις στήλες: "
                    "ΔΙΕΥΘΥΝΣΗ ή ΤΜΗΜΑ.",
                    "danger",
                )
                return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

            directories_by_name = {
                d.name.strip(): d
                for d in Directory.query.filter_by(service_unit_id=unit.id).all()
            }
            departments_by_key = {
                (dep.directory_id, dep.name.strip()): dep
                for dep in Department.query.filter_by(service_unit_id=unit.id).all()
            }
            personnel_by_agm = {
                p.agm.strip(): p
                for p in Personnel.query.filter_by(service_unit_id=unit.id, is_active=True).all()
            }

            created_dirs = 0
            created_deps = 0
            updated_dirs = 0
            updated_deps = 0
            skipped = 0

            def _cell(row_vals, i: Optional[int]):
                if i is None:
                    return None
                if i >= len(row_vals):
                    return None
                return row_vals[i]

            for row in ws.iter_rows(min_row=2, values_only=True):
                did_something = False

                dir_name = _safe_str(_cell(row, dir_name_idx)) if dir_name_idx is not None else ""
                director_agm = _safe_str(_cell(row, dir_director_agm_idx)) if dir_director_agm_idx is not None else ""

                dep_name = _safe_str(_cell(row, dep_name_idx)) if dep_name_idx is not None else ""
                head_agm = _safe_str(_cell(row, dep_head_agm_idx)) if dep_head_agm_idx is not None else ""
                assistant_agm = _safe_str(_cell(row, dep_assist_agm_idx)) if dep_assist_agm_idx is not None else ""

                directory_obj: Directory | None = None

                # Create/find Directory
                if dir_name:
                    directory_obj = directories_by_name.get(dir_name)
                    if not directory_obj:
                        directory_obj = Directory(
                            service_unit_id=unit.id,
                            name=dir_name,
                            is_active=True,
                            director_personnel_id=None,
                        )
                        db.session.add(directory_obj)
                        db.session.flush()
                        log_action(entity=directory_obj, action="CREATE", before=None, after=serialize_model(directory_obj))
                        directories_by_name[dir_name] = directory_obj
                        created_dirs += 1
                        did_something = True

                    # Optional director update
                    if director_agm:
                        p = personnel_by_agm.get(director_agm)
                        if p and directory_obj.director_personnel_id != p.id:
                            before = serialize_model(directory_obj)
                            directory_obj.director_personnel_id = p.id
                            db.session.flush()
                            log_action(entity=directory_obj, action="UPDATE", before=before, after=serialize_model(directory_obj))
                            updated_dirs += 1
                            did_something = True

                # Create/find Department
                if dep_name:
                    # A department import row must resolve to a directory.
                    if not directory_obj:
                        skipped += 1
                        continue

                    dep_key = (directory_obj.id, dep_name)
                    dep_obj = departments_by_key.get(dep_key)

                    if not dep_obj:
                        dep_obj = Department(
                            service_unit_id=unit.id,
                            directory_id=directory_obj.id,
                            name=dep_name,
                            is_active=True,
                            head_personnel_id=None,
                            assistant_personnel_id=None,
                        )
                        db.session.add(dep_obj)
                        db.session.flush()
                        log_action(entity=dep_obj, action="CREATE", before=None, after=serialize_model(dep_obj))
                        departments_by_key[dep_key] = dep_obj
                        created_deps += 1
                        did_something = True

                    head_pid = personnel_by_agm.get(head_agm).id if head_agm and head_agm in personnel_by_agm else None
                    assistant_pid = (
                        personnel_by_agm.get(assistant_agm).id if assistant_agm and assistant_agm in personnel_by_agm else None
                    )

                    if head_pid and assistant_pid and head_pid == assistant_pid:
                        skipped += 1
                        continue

                    # Update department roles only if at least one AGM column is present.
                    if head_agm or assistant_agm:
                        if dep_obj.head_personnel_id != head_pid or dep_obj.assistant_personnel_id != assistant_pid:
                            before = serialize_model(dep_obj)
                            dep_obj.head_personnel_id = head_pid
                            dep_obj.assistant_personnel_id = assistant_pid
                            db.session.flush()
                            log_action(entity=dep_obj, action="UPDATE", before=before, after=serialize_model(dep_obj))
                            updated_deps += 1
                            did_something = True

                if not did_something:
                    skipped += 1

            db.session.commit()

            flash(
                "Import ολοκληρώθηκε. "
                f"Δημιουργήθηκαν: {created_dirs} Διευθύνσεις, {created_deps} Τμήματα. "
                f"Ενημερώθηκαν: {updated_dirs} Διευθύνσεις, {updated_deps} Τμήματα. "
                f"Παραλείφθηκαν: {skipped} γραμμές.",
                "success",
            )
            return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

        flash("Μη έγκυρη ενέργεια.", "danger")
        return redirect(url_for("admin.organization_setup", service_unit_id=unit.id))

    return render_template(
        "admin/organization_setup.html",
        service_units=service_units,
        scope_service_unit_id=(unit.id if unit else None),
        unit=unit,
        directories=directories,
        departments=departments,
        personnel_list=personnel_list,
        is_admin=getattr(current_user, "is_admin", False),
    )